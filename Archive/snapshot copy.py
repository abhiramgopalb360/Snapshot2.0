import string
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from vyper.utils.tools import StatisticalTools as st

import preprocessing
from explorer import DataProfiler  # 100522 customized explorer

warnings.simplefilter(action="ignore", category=pd.core.common.SettingWithCopyWarning)


class Snapshot:
    def __init__(
        self,
        profile_data: pd.DataFrame,
        segment_var: str,
        baseline: str = "",
        segments: list = [],
        bins_path: str = "",
        nbins: int = 6,
        variable_order: list = [],
        include: list = [],
        exclude: list = [],
        continuous: list = [],
        exclude_other: bool = False,
        na_drop_threshold: float = 0.95,
        epsilon: bool = False,
        acxiom: bool = False,
        dictionary_path: str = ""
    ) -> None:

        self.profile_data = profile_data
        self.segment_var = segment_var
        self.nbins = nbins
        self.variable_order = variable_order
        self.include = include
        self.exclude = exclude
        self.bins_path = bins_path
        self.continuous = continuous
        self.exclude_other = exclude_other
        # TODO: add logic to drop variables above na_drop_threshold
        self.na_drop_threshold = na_drop_threshold
        self.epsilon = epsilon
        self.acxiom = acxiom
        self.dictionary_path = dictionary_path

        if not segments:
            segments = profile_data[segment_var].unique().tolist()
        else:
            all_segs = profile_data[segment_var].unique().tolist()
            for seg in segments:
                if seg not in all_segs:
                    raise ValueError(f"Segment '{seg}' not in '{segment_var}'")
        self.num_segments = len(segments)

        if baseline:
            if baseline not in segments:
                raise ValueError(f"Baseline '{baseline}' not in '{segment_var}'")
            segments.insert(0, segments.pop(segments.index(baseline)))

        self.mapping_dict = {}
        for i, seg in enumerate(segments):
            if i == 0:
                self.mapping_dict["Baseline"] = seg
            else:
                self.mapping_dict[f"Segment_{i}"] = seg

        # get custom bounds for continuous vars
        # TODO: add exception if no or invalid path provided
        df_continuous = pd.read_csv(bins_path)
        # Drop variables that do not have pre-defined bins
        df_continuous.dropna(inplace=True)
        # create bin dictionary for run_profiler()
        self.continuous_var_bounds = {}
        for _, row in df_continuous.iterrows():
            # Read in the row as string and convert to list of floats
            self.continuous_var_bounds[row["Attribute"]] = [float(i) for i in row["Bin"].split(sep=",")]

        self.categories = {}
        for _, row in df_continuous.iterrows():
            # Read in the row as string and convert to list of floats
            self.categories[row["Snowflake Field names"]] = row["Category"]
        self.unique_categories = sorted(set(self.categories.values()))

    def run_profiler(self) -> None:

        self.__preprocess()

        for col in self.profile_data:
            try:
                self.profile_data[col] = pd.to_numeric(self.profile_data[col])
            except Exception:
                continue

        varclass = pd.DataFrame()
        for col in self.profile_data.columns:
            varclass = pd.concat([varclass.reset_index(drop=True), pd.DataFrame([st.classify_variable(self.profile_data[col])], columns=[col])], axis=1)
        varclass[varclass.columns in self.continuous] = "continuous"

        if self.include:
            self.exclude = varclass.columns[varclass.columns not in self.include]
        # Variables to exclude
        varclass[varclass.columns in self.exclude] = "exclude"

        if self.exclude_other:
            varclass[varclass.columns == "Categorical+other"] = "exclude"
        continuous_var_cuts = dict()
        self.varclass = varclass

        # binning continuous variable to bins in the binning file
        for variable in varclass.columns[varclass.iloc[0] == "continuous"]:
            # TODO: add condition for epsilon
            # if variable.startswith("MT_") or variable.startswith("PROPENSITY_") or variable.startswith("LIKELY_") or variable == "TGT_PRE_MOVER_20_MODEL":
            #     continuous_var_cuts[variable] = pd.cut(self.profile_data[variable], bins=[0, 5, 25, 45, 65, 85, 99])
            if variable in self.continuous_var_bounds:
                cut_bins = self.continuous_var_bounds[variable]
                continuous_var_cuts[variable] = pd.cut(self.profile_data[variable], bins=cut_bins)
            else:
                continuous_var_cuts[variable] = st.get_breaks(self.profile_data.loc[:, variable], nbins=self.nbins, squash_extremes=False)

        col_order = []
        index_order = []
        # TODO: array([2, 0, 1]) Always encode 0 as US
        # Change the US population as BASELINE
        for seg in np.sort(list(self.mapping_dict.keys())):
            profile_1 = DataProfiler(self.profile_data[self.profile_data[self.segment_var] == self.mapping_dict[seg]]).create_profile(
                number_of_bins=self.nbins, cts_cuts=continuous_var_cuts
            )

            if seg == "Baseline":
                profile = profile_1
                profile.columns = ["Variable", "Category", "Count " + str(self.mapping_dict[seg]), "Percent " + str(self.mapping_dict[seg])]
                for cols in profile.columns:
                    col_order.append(cols)

            else:
                profile_1.columns = ["Variable", "Category", "Count " + str(self.mapping_dict[seg]), "Percent " + str(self.mapping_dict[seg])]
                col_order.append("Count " + str(self.mapping_dict[seg]))
                col_order.append("Percent " + str(self.mapping_dict[seg]))
                profile = pd.merge(profile, profile_1, on=["Variable", "Category"], how="outer")
                profile[str(self.mapping_dict[seg]) + " vs " + self.mapping_dict["Baseline"]] = (
                    profile["Percent " + str(self.mapping_dict[seg])] / profile["Percent " + self.mapping_dict["Baseline"]]
                ) * 100
                index_order.append(str(self.mapping_dict[seg]) + " vs " + self.mapping_dict["Baseline"])

        profile = profile[col_order + index_order]

        for i in profile.columns:

            if "Count" in i:
                profile[i][profile[i].isna()] = 0
            if "Percent" in i:
                profile[i][profile[i].isna()] = 0

        # Defining PSI
        psi = {}
        uscol = ""
        NumPSI = []
        for i in profile.columns:
            if ("Percent" in i) & (self.mapping_dict["Baseline"] not in i):
                NumPSI.append(i)
            elif ("Percent" in i) & (self.mapping_dict["Baseline"] in i):
                uscol = i

        for _, row in profile.iterrows():
            if row["Variable"] not in psi:
                psi[row["Variable"]] = {}
            if row["Category"] not in psi[row["Variable"]]:
                psi[row["Variable"]][row["Category"]] = {}

            for i in NumPSI:

                try:
                    psi[row["Variable"]][row["Category"]][i + "_PSI"] = (row[uscol] - row[i]) * np.log(row[uscol] / row[i])
                    if psi[row["Variable"]][row["Category"]][i + "_PSI"] == np.inf:
                        psi[row["Variable"]][row["Category"]][i + "_PSI"] = 0
                    psi[row["Variable"]][i + "_Overall"] = 0

                except Exception:
                    psi[row["Variable"]][row["Category"]][i + "_PSI"] = 0

        overall = {}
        for variable in psi:

            for label in psi[variable]:
                if "Overall" in str(label):
                    continue

                for scores in psi[variable][label]:
                    if variable not in overall:
                        overall[variable] = {}
                    if scores not in overall[variable]:
                        overall[variable][scores] = 0
                    overall[variable][scores] += psi[variable][label][scores]

        # Variable overall contains all the overall PSI Scores
        self.overall = overall

        for col in profile.columns[profile.columns.str.contains("Percent")]:
            profile[col] = profile[col] / 100

        profile["Category"] = profile["Category"].astype(str)
        self.profile = profile

        if self.epsilon:
            self.add_extra()

    def add_extra(self) -> None:

        profile = self.profile.copy(deep=True)

        Field_dict = pd.read_excel(self.dictionary_path)
        Field_dict = Field_dict.loc[:, "NAME":"Category"]
        Field_dict["NAME"].fillna(method="ffill", inplace=True)
        Field_dict["Snowflake"].fillna(method="ffill", inplace=True)
        Field_dict["Value"].fillna("", inplace=True)
        Field_dict.columns = ["NAME", "Description", "RATEID", "Value", "Value Description", "Current Count", "Current %", "Snowflake", "Category"]
        # Confirm numpy is imported

        prev_dict = {}

        def toDict(Field, Value, VD, snowflake, FDdescription):
            Field = str(Field)
            Value = str(Value)
            VD = str(VD)
            snowflake = str(snowflake)
            FDdescription = str(FDdescription)

            if Value == "":
                return 0

            if snowflake not in prev_dict:
                prev_dict[snowflake] = {}
                prev_dict[snowflake]["Field"] = Field
                prev_dict[snowflake]["Field Description"] = FDdescription
            if Value not in prev_dict[snowflake]:
                prev_dict[snowflake][Value] = {}
            prev_dict[snowflake][Value]["desp"] = VD
            # prev_dict[snowflake][Value]['Current Count'] =  CC
            # prev_dict[snowflake][Value]['Current %'] = CP
            return prev_dict

        Field_dict.apply(lambda x: toDict(x["NAME"], x["Value"], x["Value Description"], x["Snowflake"], x["Description"]), axis=1)
        # Add the Field Name
        def addFieldName(x):
            if x in prev_dict.keys():
                return prev_dict[x]["Field"]

        profile.insert(1, "Label", "")
        profile["Label"] = profile["Variable"].apply(lambda x: addFieldName(x))

        # Add Field Description
        def addFieldDesc(x):
            if x in prev_dict.keys():
                return prev_dict[x]["Field Description"]

        profile.insert(2, "Definition", "")
        profile["Definition"] = profile["Variable"].apply(lambda x: addFieldDesc(x))

        # Add Value Description
        def addValuedescription(snowflake, value):

            try:
                if snowflake in prev_dict.keys():
                    if value in prev_dict[snowflake].keys():
                        return prev_dict[snowflake][value]["desp"]
                    elif snowflake.startswith("MT_") or snowflake.startswith("PROPENSITY_") or snowflake.startswith("LIKELY_") or snowflake == "TGT_PRE_MOVER_20_MODEL":
                        return prev_dict[snowflake][str(value)]["desp"]
                    else:
                        try:
                            value = int(value)
                            if value in prev_dict[snowflake].keys():
                                return prev_dict[snowflake][value]["desp"]
                        except Exception:
                            pass
            except:
                pass

        profile.insert(4, "Description", "")
        profile["Description"] = profile.apply(lambda x: addValuedescription(x["Variable"], str(x["Category"])), axis=1)

        self.profile_extra = profile

    def create_profile(self, filename, split_category=False) -> None:

        wb = Workbook()
        del wb["Sheet"]

        if split_category:
            for cat in self.unique_categories[::-1]:
                snowflake_vars = [key for key, value in self.categories.items() if value == cat]
                if self.epsilon:
                    self.add_extra()
                    subset_profile = self.profile_extra[self.profile_extra["Variable"].isin(snowflake_vars)]
                else:
                    subset_profile = self.profile[self.profile["Variable"].isin(snowflake_vars)]

                if subset_profile.empty:
                    continue

                profiling_ws = wb.create_sheet(cat, 0)
                profiling_ws = wb.active

                self.__allformat(profiling_ws, subset_profile)

        profiling_ws = wb.create_sheet("Profile", 0)
        profiling_ws = wb.active
        if self.epsilon:
            self.add_extra()
            final_profile = self.profile_extra
        else:
            final_profile = self.profile
        
        self.__allformat(profiling_ws, final_profile)

        ws2 = wb.create_sheet("PSI")

        psi_df = pd.DataFrame(self.overall).transpose().reset_index()
        psi_df.rename({"index": "Category"}, axis=1, inplace=True)

        rows = dataframe_to_rows(psi_df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)

        # TODO: raise error if save file is not .xlsx
        wb.save(filename)

    def create_visual(self, filename, plot_index=True, data_table=False, show_axes=True, show_na=True) -> None:

        wb = Workbook()
        del wb["Sheet"]

        profile = self.profile.copy(deep=True)

        profile["Variable"] = profile["Variable"].apply(lambda name: name[:31] if len(name) >= 31 else name)

        for var in profile["Variable"].unique()[::-1]:
            subset_profile = profile[profile["Variable"] == var]

            if subset_profile.empty:
                continue

            # create sheet
            profiling_ws = wb.create_sheet(var, 0)
            profiling_ws = wb.active

            self.__allformat(profiling_ws, subset_profile, show_na)
            self.__visual(profiling_ws, plot_index, data_table, show_axes)

        # TODO: format Index page
        all_var_profiling_ws1 = wb.create_sheet("Index", 0)
        # make new sheet the active sheet we are working on
        wb.active = wb["Index"]
        all_var_profiling_ws1 = wb.active
        for i in wb.sheetnames:
            all_var_profiling_ws1.append([i])

        # TODO: raise error if save file is not .xlsx
        wb.save(filename)

    def __allformat(self, ws, profile, show_na=True) -> None:
        


        # remove gridlines from the sheet
        ws.sheet_view.showGridLines = False

        # set border line thickness
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")

        if not show_na:
            profile = profile[~profile["Category"].isin(["NA", "99", 99, "Z"])]

        # get max rows and cols of profiling df
        max_rows = profile.shape[0]
        max_cols = profile.shape[1]

        input_rows = range(4, max_rows + 4)
        input_cols = list(string.ascii_uppercase)[1: max_cols + 1]

        for y, i in enumerate(input_cols):
            x = i + "3"

            ws[x].font = Font(bold=True)
            current_cell = ws[x]
            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")

            if y == 0:
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif y == max_cols - 1:
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
            else:
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

            ws[x] = profile.columns[y]

        counter = []
        counter_val = 1


        for index, i in enumerate(input_cols[2::]):
            x = i + "2"

            if index >= self.num_segments * 2:
                break

            if index % 2 == 1:
                ws.merge_cells(f"{counter[-1]}:{x}")
                ws[counter[-1]].font = Font(bold=True)
                current_cell = ws[counter[-1]]
                current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                continue

            if i == "D":
                ws[x] = "BASELINE"
                ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
                counter.append(x)
                continue
            ws[x] = "SEGMENT " + str(counter_val)
            counter_val += 1
            counter.append(x)

        border = Border(top=thick, left=thick, right=thick, bottom=thick)

        def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):

            """
            Apply styles to a range of cells as if they were a single cell.

            :param ws:  Excel worksheet instance
            :param range: An excel range to style (e.g. A1:F20)
            :param border: An openpyxl Border
            :param fill: An openpyxl PatternFill or GradientFill
            :param font: An openpyxl Font object
            """

            top = Border(top=border.top)
            left = Border(left=border.left)
            right = Border(right=border.right)
            bottom = Border(bottom=border.bottom)

            first_cell = ws[cell_range.split(":")[0]]
            if alignment:
                ws.merge_cells(cell_range)
                first_cell.alignment = alignment

            rows = ws[cell_range]
            if font:
                first_cell.font = font

            for cell in rows[0]:
                cell.border = cell.border + top
            for cell in rows[-1]:
                cell.border = cell.border + bottom

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = l.border + left
                r.border = r.border + right
                if fill:
                    for c in row:
                        c.fill = fill

        for range1 in ws.merged_cells.ranges:
            style_range(ws, str(range1), border=border)

        for y, i in enumerate(input_cols):
            for z, j in enumerate(input_rows):
                ws.row_dimensions[j].height = 25
                c = i + str(j)
                current_cell = ws[c]
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                    current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

                ws[c] = profile.iloc[z, y]

                if y == 0:
                    current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
                    if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                        current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                elif y == max_cols - 1:
                    current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
                    if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                        current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)

        char = "C"
        for _ in range(self.num_segments):
            char = chr(ord(char) + 2)
            for col in ws[char]:
                col.number_format = "0%"

        start = chr(ord(char) + 1)
        for _ in range(1, self.num_segments):
            char = chr(ord(char) + 1)
            for col in ws[char]:
                col.number_format = "#,#0"

        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 32
        for c in range(ord("D"), ord(char) + 1):
            ws.column_dimensions[chr(c)].width = 16

        # Headers Alignment
        for row in ws.iter_rows(min_col=2, max_col=11, min_row=3, max_row=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Columns
        for row in ws.iter_rows(min_col=2, max_col=14):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

            # Color scale
        red = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
        yellow = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")
        green = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type="solid")
        # white = PatternFill(start_color="00EEEEEE", end_color="00EEEEEE", fill_type="solid")

        dxf1 = DifferentialStyle(fill=red)
        dxf2 = DifferentialStyle(fill=yellow)
        dxf3 = DifferentialStyle(fill=green)
        # dxf4 = DifferentialStyle(fill=white)

        rule1 = Rule(type="cellIs", operator="between", formula=[0.0, 85.0], dxf=dxf1)
        rule2 = Rule(type="cellIs", operator="between", formula=[85.0, 115.0], dxf=dxf2)
        rule3 = Rule(type="cellIs", operator="between", formula=[115.0, 1000000.0], dxf=dxf3)
        # rule4 = Rule(type="cellIs", operator="equal", formula=[0], dxf=dxf4)

        final_row = ws.max_row

        rule_string = f"{start}4:{char}{final_row}"

        ws.conditional_formatting.add(rule_string, rule1)
        ws.conditional_formatting.add(rule_string, rule2)
        ws.conditional_formatting.add(rule_string, rule3)

        # merge first column cells
        self.__merge(ws)

    def __visual(self, ws, plot_index, data_table, show_axes) -> None:
        def color_palette(n):
            if n == 2:
                return ["003f5c", "ffa600"]
            elif n == 3:
                return ["003f5c", "bc5090", "ffa600"]
            elif n == 4:
                return ["003f5c", "7a5195", "ef5675", "ffa600"]
            elif n == 5:
                return ["003f5c", "58508d", "bc5090", "ff6361", "ffa600"]
            else:
                box = []
                while len(box) < n:
                    box += ["003f5c", "7a5195", "ef5675", "ffa600"]
                return box[:n]

        c1 = BarChart()
        c1.height = 16  # default is 7.5
        c1.width = 30  # default is 15
        if data_table:
            c1.plot_area.dTable = DataTable()
            c1.plot_area.dTable.showHorzBorder = True
            c1.plot_area.dTable.showVertBorder = True
            c1.plot_area.dTable.showOutline = True
            c1.plot_area.dTable.showKeys = True
            c1.legend = None
        else:
            c1.dataLabels = DataLabelList()
            c1.dataLabels.showVal = True
            c1.legend.position = "b"

        colors = color_palette(self.num_segments)

        for seg in range(self.num_segments):
            x1 = seg * 2 + 5
            data = Reference(ws, min_col=x1, min_row=3, max_col=x1, max_row=ws.max_row - 1)
            cats = Reference(ws, min_col=3, min_row=4, max_col=3, max_row=ws.max_row - 1)
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(cats)
            c1.shape = 4
            c1.series[seg].graphicalProperties.solidFill = colors[seg]

        if show_axes:
            c1.x_axis.title = "Categories"
            c1.y_axis.title = "Percentage"
        c1.y_axis.majorGridlines = None
        c1.title = ws["B4"].value

        # Create a second chart
        if plot_index:
            c2 = LineChart()

            for seg in range(1, self.num_segments):
                x2 = seg + x1

                data = Reference(ws, min_col=x2, min_row=3, max_col=x2, max_row=ws.max_row - 1)
                cats = Reference(ws, min_col=3, min_row=4, max_col=3, max_row=ws.max_row - 1)

                c2.add_data(data, titles_from_data=True)
                c2.set_categories(cats)

            c2.y_axis.axId = 200
            c2.y_axis.title = "Index"

            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            c2.y_axis.crosses = "max"
            c1 += c2

        ws.add_chart(c1, f"B{ws.max_row + 5}")

    def __merge(self, ws, merge_columns=[2]) -> None:

        # Merge cells
        key_column = 2
        start_row = 4
        max_row = ws.max_row + 1
        key = None

        # Iterate all rows in key_column
        for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
            if key != row_cells[0].value or row == max_row:
                if key is not None:
                    for merge_column in merge_columns:
                        ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row - 1, end_column=merge_column)
                        ws.cell(row=start_row, column=merge_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    start_row = row
                key = row_cells[0].value
            if row == max_row:
                row += 1

    def __preprocess(self) -> None:
        self.profile_data = pd.DataFrame(self.profile_data)

        if self.epsilon:
            self.profile_data = preprocessing.epsilon_preprocess(self.profile_data)
        elif self.acxiom:
            # TODO add preprocessing for acxiom
            pass
