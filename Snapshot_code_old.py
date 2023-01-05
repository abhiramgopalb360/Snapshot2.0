# %%
from functools import reduce
import vyper
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from vyper.user import Model
from vyper.utils.tools import StatisticalTools as st
from vyper.user import Model

# from vyper.user.explorer import DataProfiler
from explorer import DataProfiler  # 100522 customized explorer
from sklearn.utils import shuffle
from sklearn.model_selection import train_test_split
from pandas.api.types import is_string_dtype
from openpyxl import Workbook
import pandas as pd
import numpy as np
import re
import statsmodels.api as sm
import math
import warnings
import string
import xlwings as xw
import openpyxl

# %%
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import colors
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart.plotarea import DataTable
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

from openpyxl.styles import Alignment
from openpyxl.styles import colors
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, Color
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill
from openpyxl import Workbook

from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# %%

## Snapshot functions
def addextra(epsilonpath, profile):
    # Read in the data and prepare the dataframe
    Field_dict = pd.read_excel(epsilonpath)
    Field_dict = Field_dict.loc[:, "NAME":"Category"]
    Field_dict["NAME"].fillna(method="ffill", inplace=True)
    Field_dict["Snowflake"].fillna(method="ffill", inplace=True)
    Field_dict["Value"].fillna("", inplace=True)
    Field_dict.columns = ["NAME", "Description", "RATEID", "Value", "Value Description", "Current Count", "Current %", "Snowflake", "Category"]
    # Confirm numpy is imported
    import numpy as np

    prev_dict = {}

    def toDict(Field, Value, VD, snowflake, FDdescription):

        if Value == "":
            return 0

        if snowflake not in prev_dict:
            prev_dict[snowflake] = {}
            prev_dict[snowflake]["Field"] = Field
            prev_dict[snowflake]["Field Description"] = FDdescription
        if Value not in prev_dict[snowflake]:
            prev_dict[snowflake][Value] = {}
        prev_dict[snowflake][Value]["desp"] = VD
        # prev_dict[snowflake][Value]['Current Count'] =  CC
        # prev_dict[snowflake][Value]['Current %'] = CP
        return prev_dict

    k = Field_dict.apply(lambda x: toDict(x["NAME"], x["Value"], x["Value Description"], x["Snowflake"], x["Description"]), axis=1)
    # Add the Field Name
    def addFieldName(x):
        if x in prev_dict.keys():
            return prev_dict[x]["Field"]

    profile.insert(1, "Label", "")
    profile["Label"] = profile["Variable"].apply(lambda x: addFieldName(x))
    # print(profile[profile['Variable']=='MT_CONSISTENT_RELIGIOUS_DONORS'])
    # Add Field Description
    def addFieldDesc(x):
        if x in prev_dict.keys():
            return prev_dict[x]["Field Description"]

    profile.insert(2, "Definition", "")
    profile["Definition"] = profile["Variable"].apply(lambda x: addFieldDesc(x))
    # print(profile[profile['Variable']=='ETHNIC_GROUP_CODE3'])
    # print(profile[profile['Variable']=='MT_CONSISTENT_RELIGIOUS_DONORS'])
    # Add Value Description
    def addValuedescription(snowflake, value):
        if snowflake in prev_dict.keys():
            if value in prev_dict[snowflake].keys():
                return prev_dict[snowflake][value]["desp"]
            elif snowflake.startswith("MT_") or snowflake.startswith("PROPENSITY_") or snowflake.startswith("LIKELY_") or snowflake == "TGT_PRE_MOVER_20_MODEL":
                return prev_dict[snowflake][str(value)]["desp"]
            else:
                try:
                    value = int(value)
                    if value in prev_dict[snowflake].keys():
                        return prev_dict[snowflake][value]["desp"]
                except:
                    pass

    profile.insert(4, "Description", "")
    profile["Description"] = profile.apply(lambda x: addValuedescription(x["Variable"], x["Category"]), axis=1)
    # print(prev_dict['ETHNIC_GROUP_CODE3'])

    return profile


def report2(profile_extra, overall, savepath, continuous_path):

    df_categories = pd.read_csv(continuous_path)
    # Drop variables that do not have pre-defined bins
    unique_categories = df_categories["Category"].unique().tolist()

    profile = profile_extra

    wb = Workbook()

    # set border line thickness
    thick = Side(border_style="thick", color="000000")
    # medium = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")

    # get max rows and cols of profiling df
    max_rows = profile.shape[0]
    max_cols = profile.shape[1]
    from openpyxl.utils.dataframe import dataframe_to_rows

    ### Create dataframes for each category
    all_var = []
    for cat in unique_categories:
        ## Find variables under category
        snowflake_vars = df_categories[df_categories["Category"] == cat]["Snowflake Field names"].tolist()
        ## Subset dataframe
        subset_df = profile[profile["Variable"].isin(snowflake_vars)]
        all_var.append(snowflake_vars)
        ## Creating new sheet for each category
        all_var_profiling_ws = wb.create_sheet(cat, 0)
        # make new sheet the active sheet we are working on
        all_var_profiling_ws = wb.active

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thick, left=thick, right=thick, bottom=thick)

        rows = dataframe_to_rows(subset_df)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                if type(value) == pd._libs.interval.Interval:
                    value = str(value)
                all_var_profiling_ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                all_var_profiling_ws.cell(row=r_idx + 1, column=c_idx + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # remove gridlines from the sheet
        all_var_profiling_ws.sheet_view.showGridLines = False

        # set border line thickness
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # get max rows and cols of profiling df
        max_rows = profile.shape[0]
        max_cols = profile.shape[1]

        input_rows = range(4, max_rows + 4)
        input_cols = list(string.ascii_uppercase)[1 : max_cols + 1]

        y = 0
        for i in input_cols:
            x = "3"
            x = i + x

            all_var_profiling_ws[x].font = Font(bold=True)
            current_cell = all_var_profiling_ws[x]
            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 20221011
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")

            if y == 0:
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif y == max_cols - 1:
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
            else:
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

            all_var_profiling_ws[x] = profile.columns[y]
            y = y + 1

        counter = []
        counter_val = 1

        for index, i in enumerate(input_cols[5::]):
            x = "2"
            x = i + x
            if index % 2 == 1:
                all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
                all_var_profiling_ws[counter[-1]].font = Font(bold=True)
                current_cell = all_var_profiling_ws[counter[-1]]
                current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 20221011
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                continue

            if (i == input_cols[-2]) or (i == input_cols[-1]):
                break
            if i == "G":
                all_var_profiling_ws[x] = "BASELINE"
                all_var_profiling_ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
                counter.append(x)
                continue
            all_var_profiling_ws[x] = "Segment " + str(counter_val)
            counter_val += 1
            counter.append(x)

        def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):

            """
            Apply styles to a range of cells as if they were a single cell.

            :param ws:  Excel worksheet instance
            :param range: An excel range to style (e.g. A1:F20)
            :param border: An openpyxl Border
            :param fill: An openpyxl PatternFill or GradientFill
            :param font: An openpyxl Font object
            """

            top = Border(top=border.top)
            left = Border(left=border.left)
            right = Border(right=border.right)
            bottom = Border(bottom=border.bottom)

            first_cell = ws[cell_range.split(":")[0]]
            if alignment:
                ws.merge_cells(cell_range)
                first_cell.alignment = alignment

            rows = ws[cell_range]
            if font:
                first_cell.font = font

            for cell in rows[0]:
                cell.border = cell.border + top
            for cell in rows[-1]:
                cell.border = cell.border + bottom

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = l.border + left
                r.border = r.border + right
                if fill:
                    for c in row:
                        c.fill = fill

                    # medium = Side(border_style="medium", color="000000")

        for range1 in all_var_profiling_ws.merged_cells.ranges:
            style_range(all_var_profiling_ws, str(range1), border=border)

        all_var_profiling_ws.move_range("B3:Z3", cols=1)
        all_var_profiling_ws.delete_rows(2)
        all_var_profiling_ws.delete_cols(2)

        all_var_profiling_ws.insert_rows(2)

        counter = []
        counter_val = 1

        for index, i in enumerate(input_cols[5::]):
            x = "2"
            x = i + x
            if index % 2 == 1:
                all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
                all_var_profiling_ws[counter[-1]].font = Font(bold=True)
                current_cell = all_var_profiling_ws[counter[-1]]
                current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                continue

            if (i == input_cols[-2]) or (i == input_cols[-1]):
                break
            if i == "G":
                all_var_profiling_ws[x] = "BASELINE"
                all_var_profiling_ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
                counter.append(x)
                continue
            all_var_profiling_ws[x] = "Segment " + str(counter_val)
            counter_val += 1
            counter.append(x)

        # all_var_profiling_ws.delete_cols(2)
        # all_var_profiling_ws.delete_rows(3)

        # Merge cells A and B test
        key_column = 2
        merge_columns = [2, 3, 4]
        start_row = 4
        max_row = all_var_profiling_ws.max_row
        key = None

        # Iterate all rows in `key_colum`
        for row, row_cells in enumerate(all_var_profiling_ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
            if key != row_cells[0].value or row == max_row:
                if not key is None:
                    for merge_column in merge_columns:
                        all_var_profiling_ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row - 1, end_column=merge_column)
                        all_var_profiling_ws.cell(row=start_row, column=merge_column).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )  # 1011
                    start_row = row
                key = row_cells[0].value
            if row == max_row:
                row += 1

        if profile.shape[1] > 10:
            allformat3(all_var_profiling_ws)
        else:
            allformat2(all_var_profiling_ws)

    ws2 = wb.create_sheet("Allcategory", 0)

    psi_df = pd.DataFrame(overall).transpose().reset_index()
    psi_df.rename({"index": "Category"}, axis=1, inplace=True)

    def flatten(xss):
        return [x for xs in xss for x in xs]

    all_var2 = flatten(all_var)
    psi_df1 = psi_df[psi_df["Category"].isin(all_var2)]

    rows = dataframe_to_rows(psi_df1, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    wb.save(savepath)


def CreateXLProfile_Snap(profile, overall, savepath):
    wb = Workbook()

    profile["Category"] = profile["Category"].astype(str)

    # create sheet
    all_var_profiling_ws = wb.create_sheet("profile", 0)

    # make new sheet the active sheet we are working on
    all_var_profiling_ws = wb.active

    # remove gridlines from the sheet
    all_var_profiling_ws.sheet_view.showGridLines = False

    # set border line thickness
    thick = Side(border_style="thick", color="000000")
    # medium = Side(border_style="medium", color="000000")
    thin = Side(border_style="thin", color="000000")

    # get max rows and cols of profiling df
    max_rows = profile.shape[0]
    max_cols = profile.shape[1]

    input_rows = range(4, max_rows + 4)
    input_cols = list(string.ascii_uppercase)[1 : max_cols + 1]

    y = 0
    for i in input_cols:
        x = "3"
        x = i + x

        all_var_profiling_ws[x].font = Font(bold=True)
        current_cell = all_var_profiling_ws[x]
        current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_cell.fill = PatternFill("solid", fgColor="A9C4FE")

        if y == 0:
            current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
        elif y == max_cols - 1:
            current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
        else:
            current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

        all_var_profiling_ws[x] = profile.columns[y]
        y = y + 1

    counter = []
    counter_val = 1

    for index, i in enumerate(input_cols[5::]):
        x = "2"
        x = i + x
        if index % 2 == 1:
            all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
            all_var_profiling_ws[counter[-1]].font = Font(bold=True)
            current_cell = all_var_profiling_ws[counter[-1]]
            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            continue

        if (i == input_cols[-2]) or (i == input_cols[-1]):
            break
        if i == "G":
            all_var_profiling_ws[x] = "BASELINE"
            all_var_profiling_ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
            counter.append(x)
            continue
        all_var_profiling_ws[x] = "Segment " + str(counter_val)
        counter_val += 1
        counter.append(x)
        # all_var_profiling_ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=5)
        # all_var_profiling_ws.merge_cells(f'{x}:D2')

    def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):

        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

    border = Border(top=thick, left=thick, right=thick, bottom=thick)

    for range1 in all_var_profiling_ws.merged_cells.ranges:
        style_range(all_var_profiling_ws, str(range1), border=border)

    y = 0
    for i in input_cols:
        z = 0
        for j in input_rows:
            all_var_profiling_ws.row_dimensions[j].height = 25
            x = str(j)
            x = i + x
            # print(x)
            current_cell = all_var_profiling_ws[x]
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

            all_var_profiling_ws[x] = profile.iloc[z, y]

            if y == 0:
                current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
                if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                    current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif y == max_cols - 1:
                current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
                if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                    current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)

            z = z + 1
        y = y + 1
    if profile.shape[1] > 10:
        allformat3(all_var_profiling_ws)
    else:
        allformat2(all_var_profiling_ws)

    ws2 = wb.create_sheet("Allcategory")

    psi_df = pd.DataFrame(overall).transpose().reset_index()
    psi_df.rename({"index": "Category"}, axis=1, inplace=True)

    rows = dataframe_to_rows(psi_df, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    if savepath != "":
        wb.save(savepath)
        merge(savepath)


# %%
def Snapshot_Profile(
    profile_data,
    segment_var,
    segments,
    segment_names,
    include,
    continuous_path,
    variable_order,
    other_segment=False,
    file="Profile",
    exclude="",
    PPT=True,
    continuous=[],
    excludeother=False,
    mapping_dict={},
) -> pd.DataFrame:

    print("20221025")
    # Function to read in a csv with containing the variable names and the respective bin cutoffs
    def continuous_bins(continuous_path):
        # Read in file
        df_continuous = pd.read_csv(continuous_path)
        # Drop variables that do not have pre-defined bins
        df_continuous.dropna(inplace=True)

        # create bin dictionary for Snapshot_Profile()
        continuous_bounds = {}
        for index, row in df_continuous.iterrows():
            # Read in the row as string and convert to list of floats
            continuous_bounds[row["Snowflake Field names"]] = [float(i) for i in row["Bin"].split(sep=",")]

        return continuous_bounds

    continuous_var_bounds = continuous_bins(continuous_path)
    profile_data = pd.DataFrame(profile_data)
    profile_data = preprocess(profile_data)
    profile_data["ETHNIC_GROUP_CODE1"].nunique()
    for col in profile_data:
        try:
            profile_data[col] = pd.to_numeric(profile_data[col])
        except:
            continue

    if segments is None:
        segments = pd.unique(profile_data[segment_var])
        other_segment = False
        segment_names = str(segments)
    if not (isinstance(other_segment, bool)):
        raise ValueError("other must be True or False.")
    if segment_names is None:
        segment_names = str(segments)
        if other_segment:
            segment_names = [segment_names, "other"]
    if len(segments) + other_segment != len(segment_names):
        if len(segments) == len(segment_names):
            segment_names = [segment_names, "other"]
        else:
            warnings.warn("Incorrect length of names, replacing by segments")
            segment_names = str(segments)
        if other_segment:
            segment_names = [segment_names, "other"]
    ## output -> [1 2 0]
    profile_data["ETHNIC_GROUP_CODE1"].unique()
    varclass = pd.DataFrame()
    # 1005, raise threshold for OCCU AND DMA and
    for col in profile_data.columns:
        if col in ["OCCUPATION"]:
            varclass = pd.concat(
                [varclass.reset_index(drop=True), pd.DataFrame([st.classify_variable(profile_data[col], distint_values_threshold=30)], columns=[col])], axis=1
            )
        elif col in ["DMA"]:
            varclass = pd.concat(
                [
                    varclass.reset_index(drop=True),
                    pd.DataFrame([st.classify_variable(profile_data[col], distint_values_threshold=400, max_categories=500)], columns=[col]),
                ],
                axis=1,
            )
        elif col in ["ETHNIC_GROUP_CODE1"]:
            varclass = pd.concat(
                [
                    varclass.reset_index(drop=True),
                    pd.DataFrame([st.classify_variable(profile_data[col], distint_values_threshold=400, max_categories=500)], columns=[col]),
                ],
                axis=1,
            )
        else:
            varclass = pd.concat([varclass.reset_index(drop=True), pd.DataFrame([st.classify_variable(profile_data[col])], columns=[col])], axis=1)
    varclass[varclass.columns in continuous] = "continuous"
    varclass
    if not (include is None):
        exclude = varclass.columns[varclass.columns not in include]
    # Variables to exclude
    varclass[varclass.columns in exclude] = "exclude"
    if excludeother:
        varclass[varclass.columns == "Categorical+other"] = "exclude"
    continuous_var_cuts = dict()

    # binning continuous variable to bins in the binning file
    for variable in varclass.columns[varclass.iloc[0] == "continuous"]:
        if not continuous_var_bounds:
            continuous_var_bounds = dict()
        if variable.startswith("MT_") or variable.startswith("PROPENSITY_") or variable.startswith("LIKELY_") or variable == "TGT_PRE_MOVER_20_MODEL":
            continuous_var_cuts[variable] = pd.cut(profile_data[variable], bins=[0, 5, 25, 45, 65, 85, 99])
        elif variable in continuous_var_bounds:
            cut_bins = continuous_var_bounds[variable]
            continuous_var_cuts[variable] = pd.cut(profile_data[variable], bins=cut_bins)

        else:
            continuous_var_cuts[variable] = st.get_breaks(profile_data.loc[:, variable], nbins=10, squash_extremes=False)

    col_order = []
    index_order = []
    ### TODO: array([2, 0, 1]) Always encode 0 as US
    ### Change the US population as BASELINE
    mapping_dict = mapping_dict
    for seg in np.sort(list(mapping_dict.keys())):
        profile_1 = DataProfiler(profile_data[profile_data[segment_var] == mapping_dict[seg]]).create_profile(number_of_bins=10, cts_cuts=continuous_var_cuts)

        if seg == "Baseline":
            profile = profile_1
            profile.columns = ["Variable", "Category", "Count " + str(mapping_dict[seg]), "Percent " + str(mapping_dict[seg])]
            for cols in profile.columns:
                col_order.append(cols)

        else:
            profile_1.columns = ["Variable", "Category", "Count " + str(mapping_dict[seg]), "Percent " + str(mapping_dict[seg])]
            col_order.append("Count " + str(mapping_dict[seg]))
            col_order.append("Percent " + str(mapping_dict[seg]))
            profile = pd.merge(profile, profile_1, on=["Variable", "Category"], how="outer")
            profile[str(mapping_dict[seg]) + " vs " + mapping_dict["Baseline"]] = (
                profile["Percent " + str(mapping_dict[seg])] / profile["Percent " + mapping_dict["Baseline"]]
            ) * 100
            index_order.append(str(mapping_dict[seg]) + " vs " + mapping_dict["Baseline"])

    profile = profile[col_order + index_order]

    profile[profile["Variable"] == "ETHNIC_GROUP_CODE1"]
    # print(profile)
    for i in profile.columns:

        if "Count" in i:
            profile[i][profile[i].isna()] = 0
        if "Percent" in i:
            profile[i][profile[i].isna()] = 0

    epsilon_path = "Data/Epsilon_Final.xlsx"
    profile_extra = addextra(epsilon_path, profile)
    ADV_TGT_INCOME_30 = {
        "1": "$ 0 - $ 14,999",
        "2": "$15,000 - $19,999",
        "3": "$20,000 - $29,999",
        "4": "$30,000 - $39,999",
        "5": "$40,000 - $49,999",
        "6": "$50,000 - $74,999",
        "7": "$75,000 - $99,999",
        "8": "$100,000 - $124,999",
        "9": "$125,000 - $149,999",
        "A": "$150,000 - $174,999",
        "B": "$175,000 - 199,999",
        "C": "$200,000 - 249,999",
        "D": "$250,000 or More",
    }
    profile_extra.loc[profile_extra["Variable"] == "ADV_TGT_INCOME_30", "Description"] = profile_extra[profile_extra["Variable"] == "ADV_TGT_INCOME_30"][
        "Category"
    ].map(ADV_TGT_INCOME_30)

    ### Defining PSI

    PSI = {}
    uscol = ""
    NumPSI = []
    for i in profile_extra.columns:
        # print(i)
        if ("Percent" in i) & (mapping_dict["Baseline"] not in i):
            NumPSI.append(i)
        elif ("Percent" in i) & (mapping_dict["Baseline"] in i):
            uscol = i

    for index, row in profile_extra.iterrows():
        if row["Variable"] not in PSI:
            PSI[row["Variable"]] = {}
        if row["Category"] not in PSI[row["Variable"]]:
            PSI[row["Variable"]][row["Category"]] = {}

        for i in NumPSI:

            try:
                PSI[row["Variable"]][row["Category"]][i + "_PSI"] = (row[uscol] - row[i]) * np.log(row[uscol] / row[i])
                if PSI[row["Variable"]][row["Category"]][i + "_PSI"] == np.inf:
                    PSI[row["Variable"]][row["Category"]][i + "_PSI"] = 0
                PSI[row["Variable"]][i + "_Overall"] = 0

            except:
                PSI[row["Variable"]][row["Category"]][i + "_PSI"] = 0

    overall = {}
    for variable in PSI:

        # print(variable)
        for label in PSI[variable]:
            if "Overall" in str(label):
                continue

            for scores in PSI[variable][label]:
                if variable not in overall:
                    overall[variable] = {}
                if scores not in overall[variable]:
                    overall[variable][scores] = 0
                overall[variable][scores] += PSI[variable][label][scores]

    ### Variable overall contains all the overall PSI Scores
    ####

    for col in profile_extra.columns[profile_extra.columns.str.contains("Percent")]:
        profile_extra[col] = profile_extra[col] / 100
    profile_extra[profile_extra["Variable"] == "ETHNIC_GROUP_CODE1"]
    if file:
        import os

        # Make Profile ####
        filesave = file + ".xlsx"
        filesave2 = file + "Category" + ".xlsx"
        if os.path.exists(filesave):
            os.remove(filesave)
        report2(profile_extra, overall=overall, savepath=filesave2, continuous_path=continuous_path)
        CreateXLProfile_Snap(profile_extra, overall, savepath=filesave)

        if PPT:

            filesave = file + "PPT" + ".xlsx"
            lcn = str(profile.columns[-1])
            profile_sliced = profile[profile[lcn] > 0]
            wb = Workbook()

            def shorten_name(name):
                if len(name) >= 31:
                    name = name[0:31]
                return name

            profile_sliced[profile_sliced["Variable"] == "VEHICLE_MAKE_4"]
            profile_sliced["Variable"] = profile_sliced["Variable"].apply(lambda x: shorten_name(x))
            profile_sliced["Category"] = profile_sliced["Category"].astype(str)

            all_var_profiling_ws = wb.create_sheet(profile_sliced["Variable"].iloc[0], 0)
            # make new sheet the active sheet we are working on
            all_var_profiling_ws = wb.active

            # remove gridlines from the sheet
            all_var_profiling_ws.sheet_view.showGridLines = False

            # set border line thickness
            thick = Side(border_style="thick", color="000000")
            # medium = Side(border_style="medium", color="000000")
            thin = Side(border_style="thin", color="000000")

            # get max rows and cols of profiling df
            max_rows = profile_sliced.shape[0]
            max_cols = profile_sliced.shape[1]

            input_cols = list(string.ascii_uppercase)
            from openpyxl.utils.dataframe import dataframe_to_rows

            Allrows = dataframe_to_rows(profile)

            j = 4
            for z in range(0, len(profile_sliced)):
                current_cat = profile["Variable"].iloc[z]

                if z != 0:
                    if profile_sliced["Variable"].iloc[z] != profile_sliced["Variable"].iloc[z - 1]:
                        # create sheet
                        all_var_profiling_ws = wb.create_sheet(profile_sliced["Variable"].iloc[z], 0)
                        # make new sheet the active sheet we are working on
                        all_var_profiling_ws = wb.active

                        for y in range(len(profile_sliced.columns)):
                            current_cell = all_var_profiling_ws.cell(row=3, column=y + 2)

                            current_cell.font = Font(bold=True)
                            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                            current_cell.value = profile_sliced.columns[y]
                            if y == 0:
                                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                            elif y == max_cols - 1:
                                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
                            else:
                                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
                        j = 4
                    else:
                        incol = 2
                        for y in range(len(profile_sliced.columns)):
                            current_cell = all_var_profiling_ws.cell(row=3, column=incol)
                            incol = incol + 1
                            current_cell.font = Font(bold=True)
                            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                            current_cell.value = profile.columns[y]
                            if y == 0:
                                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                            elif y == max_cols - 1:
                                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
                            else:
                                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            for col in wb.sheetnames:
                if col == "Sheet":
                    continue
                temp_df = profile_sliced.loc[profile_sliced["Variable"] == col, :]
                ### To code for thresold 3%
                percent_cols = []
                for i in temp_df.columns:
                    if "Percent" in i:
                        percent_cols.append(i)
                ##or min
                temp_df["FLAG"] = (temp_df[percent_cols] >= 0.001).any(axis=1).astype(bool)
                temp_df = temp_df[temp_df["FLAG"]]
                temp_df = temp_df[~temp_df["Category"].isin(["NA", "99", 99, "Z"])]
                if len(temp_df) == 0:
                    # delete that sheet
                    std = wb.get_sheet_by_name(col)
                    wb.remove_sheet(std)
                    print(f"{std} removed")
                    continue
                ### To code if temp_df contains NA in

                temp_df = temp_df.drop("FLAG", axis=1)

                ###
                rows = dataframe_to_rows(temp_df)
                wb.active = wb[col]
                ws = wb.active

                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                        ws.cell(row=r_idx + 1, column=c_idx + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

                all_var_profiling_ws = ws
                all_var_profiling_ws.move_range("B3:Z3", cols=1)
                all_var_profiling_ws.delete_rows(2)
                all_var_profiling_ws.delete_cols(2)
                all_var_profiling_ws.insert_rows(2)

                counter = []
                counter_val = 1
                input_cols = list(string.ascii_uppercase)[1 : max_cols + 1]
                for index, i in enumerate(input_cols[5::]):
                    x = "2"
                    x = i + x
                    if index % 2 == 1:
                        all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
                        all_var_profiling_ws[counter[-1]].font = Font(bold=True)
                        current_cell = all_var_profiling_ws[counter[-1]]
                        current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                        current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                        continue

                    if (i == input_cols[-2]) or (i == input_cols[-1]):
                        break
                    if i == "G":
                        all_var_profiling_ws[x] = "BASELINE"
                        all_var_profiling_ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
                        counter.append(x)
                        continue
                    all_var_profiling_ws[x] = "Segment " + str(counter_val)
                    counter_val += 1
                    counter.append(x)

                    # Merge cells A and B test
                key_column = 2
                merge_columns = [2, 3, 4]
                start_row = 4
                max_row = ws.max_row
                key = None

                # Iterate all rows in `key_colum`
                for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
                    if key != row_cells[0].value or row == max_row:
                        if not key is None:
                            for merge_column in merge_columns:
                                ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row, end_column=merge_column)
                                ws.cell(row=start_row, column=merge_column).alignment = Alignment(
                                    horizontal="center", vertical="center", wrap_text=True
                                )  # 1011
                            start_row = row
                        key = row_cells[0].value
                    if row == max_row:
                        row += 1

                if profile_sliced.shape[1] == 10:
                    visual2(ws)
                else:
                    visual3(ws)

                if profile_sliced.shape[1] > 10:
                    allformat3(ws)
                else:
                    allformat2(ws)

            wb.save(filesave)

    print("110322")
    return profile_extra


# us_state = pd.read_csv('EPS_OPP_UNLOCK_08152022.csv')
# unlock = pd.read_csv('unlock_append_06082022.csv')
# # Read in the Epsilon Data Dictionary we created

# unlock = unlock[unlock['MATCH_LEVEL'].isin(['IND P1','IND P2','HH'])]

# unlock = unlock[unlock['SEGMENT_O']=='Application Completed']

# # Pull list of Selected Attributes from Snowflake Headers
# variable_list = us_state.columns.tolist()

# # # Selected Attributes only
# us_state1 = us_state[variable_list]
# unlock1 = unlock[variable_list]

# us_state1['category'] = 'US_Population'
# unlock1['category'] = 'Application Completed'

# new_df = pd.concat([us_state1,unlock1])

# new_df2 = new_df.reset_index()

# ## Change the names here to edit titles headers
# mapping_dict = {'Baseline':'US_Population','Segment_1':'Application Completed'}

# # Set up variables for snapshot
# file_name = 'Unlock_08052022_Application Completed'
# seg_var = 'category'
# epsilon_path = 'Data/Epsilon_Final.xlsx'
# bin_vars_path = 'Epsilon_attributes_binning_2.csv'


# profile_data = new_df2
# segment_var=seg_var
# continuous_path = bin_vars_path
# segments = None
# segment_names = None
# include = None
# variable_order = None
# other_segment = False
# file = file_name
# exclude = []
# PPT = True
# continuous = []
# excludeother = False
# mapping_dict=mapping_dict


# %%
# Formatting functions


def allformat2(sheet):

    # Selecting active sheet
    ws = sheet

    #! Amateur Hour begins
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16

    for col in ws["H"]:
        col.number_format = "0%"

    for col in ws["J"]:
        col.number_format = "0%"

    for col in ws["K"]:
        col.number_format = "#,##0"

        # Headers Alignment
    for row in ws.iter_rows(min_col=2, max_col=11, min_row=3, max_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Columns B, C and D
    for row in ws.iter_rows(min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    # Columns E an F
    for row in ws.iter_rows(min_col=5, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Columns G, H, I, J and K
    for row in ws.iter_rows(min_col=7, max_col=14):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Color scale
    # 10/11 need to chang it to proggrssing color GradientFill
    yellow = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")  # "00FFCC00"

    red1 = PatternFill(start_color="00FFcccc", end_color="00FFcccc", fill_type="solid")  # "00FF0000"
    red2 = PatternFill(start_color="00FF9999", end_color="00FF9999", fill_type="solid")  #
    red3 = PatternFill(start_color="00FF7C80", end_color="00FF7C80", fill_type="solid")  #
    red4 = PatternFill(start_color="00FF5050", end_color="00FF5050", fill_type="solid")  #
    red5 = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")  #

    green1 = PatternFill(start_color="00e2efda", end_color="00e2efda", fill_type="solid")
    green2 = PatternFill(start_color="00c6e0b4", end_color="00c6e0b4", fill_type="solid")  # "0099CC00"
    green3 = PatternFill(start_color="00a9d08e", end_color="00a9d08e", fill_type="solid")
    green4 = PatternFill(start_color="0070ad47", end_color="0070ad47", fill_type="solid")
    green5 = PatternFill(start_color="00548235", end_color="00548235", fill_type="solid")

    white = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")

    # GradientFill(start_color="00FF0000", end_color='990000', fill_type='solid')
    #'005C7A00'
    dxf1 = DifferentialStyle(fill=yellow)
    dxf2_1 = DifferentialStyle(fill=red1)
    dxf2_2 = DifferentialStyle(fill=red2)
    dxf2_3 = DifferentialStyle(fill=red3)
    dxf2_4 = DifferentialStyle(fill=red4)
    dxf2_5 = DifferentialStyle(fill=red5)
    dxf3_1 = DifferentialStyle(fill=green1)
    dxf3_2 = DifferentialStyle(fill=green2)
    dxf3_3 = DifferentialStyle(fill=green3)
    dxf3_4 = DifferentialStyle(fill=green4)
    dxf3_5 = DifferentialStyle(fill=green5)
    dxf4 = DifferentialStyle(fill=white)
    # rule1 = Rule(type='cellIs', operator='between', formula=[0.0, 95.0], dxf= dxf2)
    rule1_1 = Rule(type="cellIs", operator="between", formula=[70.0, 87.0], dxf=dxf2_1)
    rule1_2 = Rule(type="cellIs", operator="between", formula=[65.0, 70.0], dxf=dxf2_2)
    rule1_3 = Rule(type="cellIs", operator="between", formula=[50.0, 65.0], dxf=dxf2_3)
    rule1_4 = Rule(type="cellIs", operator="between", formula=[35.0, 50.0], dxf=dxf2_4)
    rule1_5 = Rule(type="cellIs", operator="between", formula=[0.0, 35.0], dxf=dxf2_5)

    rule2 = Rule(type="cellIs", operator="between", formula=[87.0, 115.0], dxf=dxf1)

    # rule3 = Rule(type='cellIs', operator='between', formula=[115.0, 1000000.0], dxf= dxf3_1)
    rule3_1 = Rule(type="cellIs", operator="between", formula=[115.0, 150.0], dxf=dxf3_1)
    rule3_2 = Rule(type="cellIs", operator="between", formula=[150.0, 200.0], dxf=dxf3_2)
    rule3_3 = Rule(type="cellIs", operator="between", formula=[200.0, 250.0], dxf=dxf3_3)
    rule3_4 = Rule(type="cellIs", operator="between", formula=[250.0, 300.0], dxf=dxf3_4)
    rule3_5 = Rule(type="cellIs", operator="between", formula=[300.0, 100000.0], dxf=dxf3_5)

    # rule4 = Rule(type='cellIs', operator='', formula=None, dxf= dxf4)
    try:
        final_row = ws.max_row

        rule_string = f"K4:K{final_row}"
        ws.conditional_formatting.add(rule_string, rule1_1)
        ws.conditional_formatting.add(rule_string, rule1_2)
        ws.conditional_formatting.add(rule_string, rule1_3)
        ws.conditional_formatting.add(rule_string, rule1_4)
        ws.conditional_formatting.add(rule_string, rule1_5)

        ws.conditional_formatting.add(rule_string, rule2)

        ws.conditional_formatting.add(rule_string, rule3_1)
        ws.conditional_formatting.add(rule_string, rule3_2)
        ws.conditional_formatting.add(rule_string, rule3_3)
        ws.conditional_formatting.add(rule_string, rule3_4)
        ws.conditional_formatting.add(rule_string, rule3_5)
    except:
        ## so you found this code huh.. this is the pinacle of optimisation btw
        True == True


def allformat3(sheet):

    # Selecting active sheet
    ws = sheet

    #! Amateur Hour begins
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16

    for col in ws["H"]:
        col.number_format = "0%"

    for col in ws["J"]:
        col.number_format = "0%"

    for col in ws["L"]:
        col.number_format = "0%"

    for col in ws["M"]:
        col.number_format = "#,##0"

    for col in ws["N"]:
        col.number_format = "#,##0"

        # Headers Alignment
    for row in ws.iter_rows(min_col=2, max_col=11, min_row=3, max_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Columns B, C and D
    for row in ws.iter_rows(min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    # Columns E an F
    for row in ws.iter_rows(min_col=5, max_col=6):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Columns G, H, I, J and K
    for row in ws.iter_rows(min_col=7, max_col=14):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Color scale
    red = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")
    yellow = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
    green = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type="solid")
    white = PatternFill(start_color="00FFFFFF", end_color="00FFFFFF", fill_type="solid")

    dxf1 = DifferentialStyle(fill=yellow)
    dxf2 = DifferentialStyle(fill=red)
    dxf3 = DifferentialStyle(fill=green)
    dxf4 = DifferentialStyle(fill=white)

    rule1 = Rule(type="cellIs", operator="between", formula=[0.0, 95.0], dxf=dxf1)
    rule2 = Rule(type="cellIs", operator="between", formula=[95.0, 115.0], dxf=dxf2)
    rule3 = Rule(type="cellIs", operator="between", formula=[115.0, 1000000.0], dxf=dxf3)
    # rule4 = Rule(type='cellIs', operator='', formula=None, dxf= dxf4)

    try:
        final_row = ws.max_row

        rule_string = f"K4:K{final_row}"
        ws.conditional_formatting.add(rule_string, rule1)
        ws.conditional_formatting.add(rule_string, rule2)
        ws.conditional_formatting.add(rule_string, rule3)
    except:
        ## so you found this code huh.. this is the pinacle of optimisation btw
        True == True


def visual3(worksheet):

    ws = worksheet

    c1 = BarChart()
    c1.height = 19.05  # default is 7.5
    c1.width = 33.85  # default is 15
    c1.plot_area.dTable = DataTable()
    c1.plot_area.dTable.showHorzBorder = True
    c1.plot_area.dTable.showVertBorder = True
    c1.plot_area.dTable.showOutline = True
    c1.plot_area.dTable.showKeys = True

    data = Reference(ws, min_col=8, min_row=3, max_row=ws.max_row, max_col=8)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4

    data = Reference(ws, min_col=10, min_row=3, max_row=ws.max_row, max_col=10)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4

    data = Reference(ws, min_col=12, min_row=3, max_row=ws.max_row, max_col=12)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4

    c1.x_axis.title = "Categories"
    c1.y_axis.title = "Percentage"
    c1.y_axis.majorGridlines = None
    c1.title = ws["C4"].value

    openpyxl.chart.legend.Legend(legendEntry=())

    # Create a second chart
    c2 = LineChart()

    data = Reference(ws, min_col=13, min_row=3, max_row=ws.max_row, max_col=13)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c2.add_data(data, titles_from_data=True)
    c2.set_categories(cats)

    data = Reference(ws, min_col=14, min_row=3, max_row=ws.max_row, max_col=14)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c2.add_data(data, titles_from_data=True)
    c2.set_categories(cats)

    c2.y_axis.axId = 200
    c2.y_axis.title = "Index"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c2.y_axis.crosses = "max"
    c1 += c2

    c1.legend = None

    ws.add_chart(c1, "D15")


def visual2(worksheet):
    # change size 111222

    ws = worksheet

    c1 = BarChart()
    c1.height = 19.05  # default is 7.5
    c1.width = 33.85  # default is 15
    c1.plot_area.dTable = DataTable()
    c1.plot_area.dTable.showHorzBorder = True
    c1.plot_area.dTable.showVertBorder = True
    c1.plot_area.dTable.showOutline = True
    c1.plot_area.dTable.showKeys = True

    data = Reference(ws, min_col=8, min_row=3, max_row=ws.max_row, max_col=8)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4

    data = Reference(ws, min_col=10, min_row=3, max_row=ws.max_row, max_col=10)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.shape = 4

    c1.x_axis.title = "Categories"
    c1.y_axis.title = "Percentage"
    c1.y_axis.majorGridlines = None
    c1.title = ws["C4"].value

    openpyxl.chart.legend.Legend(legendEntry=())

    # Create a second chart
    c2 = LineChart()

    data = Reference(ws, min_col=11, min_row=3, max_row=ws.max_row, max_col=11)
    cats = Reference(ws, min_col=6, min_row=4, max_row=ws.max_row, max_col=6)
    c2.add_data(data, titles_from_data=True)
    c2.set_categories(cats)

    c2.y_axis.axId = 200
    c2.y_axis.title = "Index"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c2.y_axis.crosses = "max"
    c1 += c2

    c1.legend = None

    ws.add_chart(c1, "D15")


# Merge Test


def merge(path):
    # Loading work book
    wb = load_workbook(path)
    # Selecting active sheet
    ws = wb.active

    # Merge cells A and B test
    key_column = 2
    merge_columns = [2, 3, 4]
    start_row = 4
    max_row = ws.max_row
    key = None

    # Iterate all rows in `key_colum`
    for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
        if key != row_cells[0].value or row == max_row:
            if not key is None:
                for merge_column in merge_columns:
                    ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row - 1, end_column=merge_column)
                    ws.cell(row=start_row, column=merge_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                start_row = row
            key = row_cells[0].value
        if row == max_row:
            row += 1

    wb.save(path)
