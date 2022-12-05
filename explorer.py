import pandas as pd
from vyper.utils.tools import StatisticalTools as st
from vyper.base import Variable, VariableCollection
from statsmodels.stats.weightstats import DescrStatsW
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.drawing.text import CharacterProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from copy import deepcopy


class DataProfiler:
    def __init__(self,
                 df: pd.DataFrame,
                 dependent_variable: str = None,
                 variables: VariableCollection = None):
        if not variables:
            self.variables = VariableCollection()
        else:
            self.variables = variables
        self.data = df
        self.dependent_variable = dependent_variable
        # mapping dict to convert variable kind, into a processing method for data profiling
        self.processing_dt = {'continuous': 'continuous',
                              'categorical': 'categorical',
                              'categorical+other': 'categorical'
                              }

    def create_profile(self,
                       cts_cuts: dict = None,
                       distint_values_threshold: int = 10,
                       max_vals: int = 55,
                       max_cats: int = 55,
                       number_of_bins: int = 5,
                       blanks_as_na: bool = True,
                       add_missing_rate: bool = False,
                       ) -> pd.DataFrame:
        
        if not self.variables:
            for col in self.data.columns:
                if col != self.dependent_variable:
                    if col in ['OCCUPATION']:
                        self.variables[col] = Variable(name=col,
                                                    kind=st.classify_variable(self.data[col],
                                                                                distint_values_threshold = 30,
                                                                                max_values=max_vals,
                                                                                max_categories=max_cats))
                    elif col in ['DMA']:
                         self.variables[col] = Variable(name=col,
                                                    kind=st.classify_variable(self.data[col],
                                                                                distint_values_threshold = 400,
                                                                                max_values=max_vals,
                                                                                max_categories=500))
                    elif col in ['ETHNIC_GROUP_CODE1']:
                        self.variables[col] = Variable(name=col,
                                                    kind= 'categorical')

                    else:
                        self.variables[col] = Variable(name=col,
                                                    kind=st.classify_variable(self.data[col],
                                                                                max_values=max_vals,
                                                                                max_categories=max_cats))                                                        
        if not cts_cuts:
            cts_cuts = dict()
        result_dfs = []

        for variable in self.variables:
            _sub_df = None

            if variable.kind not in ['target_variable', 'exclude']:
                if self.dependent_variable:
                    selection = [variable.name, self.dependent_variable]
                else:
                    selection = [variable.name]

                # each loop inside the code should generate a _sub_df to be passed on to the profile segment function

                if variable.kind == 'continuous':
                    # TODO change to try except catch on key in cts_cuts to
                    #  reduce complexity if performance becomes issue.
                    if variable.name in cts_cuts:
                        cuts = cts_cuts[variable.name]
                        # # change the code if cts_cuts are supposed to be the bins and not the labels
                        # cuts = pd.cut(x=self.data.loc[:, variable.name], bins=cts_cuts[variable.name])
                    else:
                        cuts = st.get_breaks(self.data.loc[:, variable.name], nbins=number_of_bins,
                                             squash_extremes=False)
                    _sub_df = self.data.loc[:, selection]
                    _sub_df['cut_up_variable'] = cuts
                    if blanks_as_na:
                        _sub_df['cut_up_variable'] = _sub_df['cut_up_variable'].cat.add_categories('NA')
                        _sub_df['cut_up_variable'] = _sub_df['cut_up_variable'].fillna('NA')
                elif variable.kind in ['categorical', 'binary']:
                    all_cuts = self.data[variable.name].value_counts()
                    _sub_df = self.data[selection]
                    if blanks_as_na:
                        _sub_df[variable.name] = _sub_df[variable.name].fillna('NA')
                    if max_cats < len(all_cuts):
                        cats_to_flip = all_cuts.index[max_cats:].to_numpy()
                        _sub_df['cut_up_variable'] = _sub_df[variable.name]. \
                            apply(lambda x: '(Other)' if x in cats_to_flip else x)
                    else:
                        _sub_df['cut_up_variable'] = _sub_df[variable.name]

                profiled_df = st.profile_segment(subset=_sub_df,
                                                 subset_binned_name='cut_up_variable',
                                                 variable_name=variable.name,
                                                 dependent_variable=self.dependent_variable)
                # removes useless NA from profiled data which screws up calculation of weighted variance
                profiled_df = profiled_df[profiled_df['Count'] != 0]

                if self.dependent_variable:
                    profiled_df = self.calculate_weighted_variance(profiled_df)
                if add_missing_rate:
                    # profiled_df['Missing_Rate'] = profiled_df[profiled_df['Category'] == 'NA']['Percent'].values[0]
                    _sub_df = self.data[selection]
                    missing_rate = (_sub_df[variable.name].isna().sum() / len(_sub_df))
                    profiled_df['Missing_Rate'] = missing_rate

                result_dfs.append(profiled_df)

        return pd.concat(result_dfs)

    def create_custom_profile(self, variable_list: list) -> pd.DataFrame:
        """
        Def: Create profiling output of a dataframe with a given list of variables.
        :param variable_list: list, Variable names that the user want to create a profile on
        :return: pandas dataframe, Profiling result
        """
        result_dfs = list()
        for variable in variable_list:
            _sub_df = None

            if self.dependent_variable:
                selection = [variable, self.dependent_variable]
            else:
                selection = [variable]

            # each loop inside the code should generate a _sub_df to be passed on to the profile segment function
            subset_binned_name = 'cut_up_variable'

            # all_cuts = self.data[variable].value_counts()
            _sub_df = self.data[selection]

            _sub_df['cut_up_variable'] = _sub_df[variable]

            profiled_df = st.profile_segment(subset=_sub_df,
                                             subset_binned_name=subset_binned_name,
                                             variable_name=variable,
                                             dependent_variable=self.dependent_variable)

            if self.dependent_variable:
                profiled_df = self.calculate_weighted_variance(profiled_df)

            missing_rate = (_sub_df[variable].isna().sum() / len(_sub_df))
            profiled_df['Missing_Rate'] = missing_rate

            result_dfs.append(profiled_df)

        return pd.concat(result_dfs)

    @staticmethod
    def calculate_weighted_variance(df: pd.DataFrame,
                                    weight_col: str = 'Percent',
                                    index_col: str = 'index',
                                    output_col: str = 'Weighted_Variance') -> pd.DataFrame:
        """
        Def: Calculate weighted variance column of the profiling output.
        :param df: pd.DataFrame, Profiling output dataframe
        :param weight_col: Percent of the count of the bucket/total records of the dataset
        :param index_col: Index is average the DV within the bucket/ average of the DV within the whole dataset
        :param output_col: str, The column name of the calculated weighted variance
        :return: pd.DataFrame
        """
        w = DescrStatsW(data=df[index_col],
                        weights=df[weight_col],
                        ddof=0)

        df[output_col] = w.var
        return df

    # @staticmethod
    def create_var_profiling_ws(self,
                                wb,
                                sheet_name: str,
                                var_prof_df: pd.DataFrame = None,
                                sort_by_variance: str = None) -> None:
        """
        Def: Function that creates variable profiling tab in excel workbook.
        :param wb: Excel workbook object
        :param sheet_name: str, Sheet name of the profiling result excel sheet
        :param var_prof_df: Variable profiling dataframe
        :param sort_by_variance: Controls sort method for the variable profile, None if sorting by variable position,
            "desc" if sorting by descending weighted variance, otherwise sorting by ascending weighted variance
        """

        if var_prof_df is None:
            var_prof_df = self.create_profile(add_missing_rate=True)
            var_prof_df['Category'] = var_prof_df['Category'].astype(str)

        if sort_by_variance:
            var_asc = sort_by_variance != 'desc'
            var_prof_df.sort_values(by=['Weighted_Variance', 'Variable'], ascending=[var_asc, True],
                                    inplace=True)

        # create sheet
        all_var_profiling_ws = wb.create_sheet(sheet_name, 0)

        all_var_profiling_ws['B1'] = sheet_name
        all_var_profiling_ws['B1'].font = Font(size=18, underline="single", bold=True)
        all_var_profiling_ws.merge_cells('B1:D1')

        # make new sheet the active sheet we are working on
        all_var_profiling_ws = wb.active

        # remove gridlines from the sheet
        all_var_profiling_ws.sheet_view.showGridLines = False

        # set border line thickness
        thick = Side(border_style="thick", color="000000")
        thin = Side(border_style="thin", color="000000")

        # cells that need text inputted
        all_var_profiling_ws['B3'] = "Variable"
        all_var_profiling_ws['C3'] = "Category"
        all_var_profiling_ws['D3'] = "Count"
        all_var_profiling_ws['E3'] = "Percent"
        all_var_profiling_ws['F3'] = "mean_DV"
        all_var_profiling_ws['G3'] = "Index"
        all_var_profiling_ws['H3'] = "Missing_Rate"
        all_var_profiling_ws['I3'] = "Weighted_Variance"

        # format text cells
        for i in ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3']:
            all_var_profiling_ws[i].font = Font(bold=True)
            current_cell = all_var_profiling_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            current_cell.border = Border(
                top=thick,
                left=thick if i == 'B3' else thin,
                right=thick if i == 'I3' else thin,
                bottom=thin)

        # get max rows and cols of profiling df
        max_rows = var_prof_df.shape[0]

        # determine cell sizes
        input_rows = range(4, max_rows + 4)
        input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i in input_cols:
            if i == 'B':
                all_var_profiling_ws.column_dimensions[i].width = 35
            elif i == 'I':
                all_var_profiling_ws.column_dimensions[i].width = 20
            else:
                all_var_profiling_ws.column_dimensions[i].width = 15

        # copy variable profiling dataframe into excel workbook
        for i in input_cols:
            z = 0
            for j in input_rows:
                all_var_profiling_ws.row_dimensions[j].height = 25
                x = i + str(j)
                current_cell = all_var_profiling_ws[x]
                current_cell.border = Border(
                    top=thick if var_prof_df['Variable'].iloc[z] != var_prof_df['Variable'].iloc[z-1] else thin,
                    left=thick if i == 'B' else thin,
                    right=thick if i == 'I' else thin,
                    bottom=thick if j == max(input_rows) else thin)

                if var_prof_df['index'].iloc[z] > 150:
                    current_cell.fill = PatternFill("solid", fgColor="FFFF00")

                if var_prof_df['Variable'].iloc[z] != var_prof_df['Variable'].iloc[z-1]:
                    current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

                if var_prof_df['index'].iloc[z] > 150:
                    current_cell.fill = PatternFill("solid", fgColor="FFFF00")

                if i != 'B':
                    current_cell.alignment = Alignment(horizontal='center', vertical='center')
                if i == 'B':
                    all_var_profiling_ws[x] = var_prof_df['Variable'].iloc[z]
                elif i == 'C':
                    all_var_profiling_ws[x] = var_prof_df['Category'].iloc[z]
                elif i == 'D':
                    all_var_profiling_ws[x] = var_prof_df['Count'].iloc[z]
                elif i == 'E':
                    all_var_profiling_ws[x] = var_prof_df['Percent'].iloc[z]
                elif i == 'F':
                    all_var_profiling_ws[x] = var_prof_df['mean_DV'].iloc[z]
                elif i == 'G':
                    all_var_profiling_ws[x] = var_prof_df['index'].iloc[z]
                elif i == 'H':
                    all_var_profiling_ws[x] = var_prof_df['Missing_Rate'].iloc[z]
                elif i == 'I':
                    all_var_profiling_ws[x] = var_prof_df['Weighted_Variance'].iloc[z]
                else:
                    break
                z = z + 1

        # get list of unique variables in profile df
        unique_var = var_prof_df['Variable'].unique()

        # initialize start row
        start_row = 4

        # initialize start chart row
        start_chart_row = 4

        # loop through variables
        for i in unique_var:
            # get number of times the current variables is in profile df
            var_count = var_prof_df.loc[var_prof_df.Variable == i, 'Variable'].count()

            chart1 = BarChart()
            chart1.title = i
            chart1.y_axis.title = 'Count'
            data1 = Reference(worksheet=all_var_profiling_ws,
                              min_row=start_row,
                              max_row=(start_row + var_count - 1),
                              min_col=4)
            cats = Reference(all_var_profiling_ws, min_col=3, min_row=start_row, max_row=start_row + var_count - 1)
            chart1.add_data(data1, titles_from_data=False)
            chart1.set_categories(cats)
            chart1.y_axis.majorGridlines = None

            # generate line chart
            chart2 = LineChart()
            chart2.title = i

            data2 = Reference(worksheet=all_var_profiling_ws,
                              min_row=start_row,
                              max_row=start_row + var_count - 1,
                              min_col=7)
            chart2.add_data(data2, titles_from_data=False)
            chart2.set_categories(cats)
            chart2.y_axis.title = 'Index'

            chart2.y_axis.crosses = 'max'

            # generate combo chart (bar + line)
            c1 = deepcopy(chart1)
            c1.x_axis.crosses = 'min'
            c1.y_axis.crossAx = 500
            c2 = deepcopy(chart2)
            c2.x_axis.axId = 500
            c2.x_axis.crosses = 'autoZero'
            c2.y_axis.majorGridlines = None
            c2.y_axis.title = "Index"

            c2.y_axis.axId = 20
            c2.y_axis.CrossAx = 500
            c2.x_axis.delete = True
            c1.y_axis.crosses = 'min'
            cp = CharacterProperties(sz=1200)

            # edited from modules, keep this block in mind if output doesn't look right
            c1.y_axis.title.tx.rich.p[0].r[0].rPr = cp
            c2.y_axis.title.tx.rich.p[0].r[0].rPr = cp
            c1.title.tx.rich.p[0].r[0].rPr = cp

            c1 += c2
            c1.legend = None

            if var_count > 10:
                height = 10
                width = 15
            else:
                height = var_count
                width = 10
            c1.height = height * 0.907141
            c1.width = width

            chart_pos = "K" + str(start_chart_row)

            all_var_profiling_ws.add_chart(c1, chart_pos)

            start_row = start_row + var_count
            start_chart_row = start_chart_row + var_count

        for i in input_rows:
            all_var_profiling_ws['D' + str(i)].number_format = '#,###'
            all_var_profiling_ws['E' + str(i)].number_format = '0.00%'
            if var_prof_df['mean_DV'].max() <= 1:
                all_var_profiling_ws['F' + str(i)].number_format = '0.00%'
            else:
                all_var_profiling_ws['F' + str(i)].number_format = '"$"#,##0_);("$"#,##0)'
            all_var_profiling_ws['G' + str(i)].number_format = '#,###0'
            all_var_profiling_ws['I' + str(i)].number_format = '#,###0'
            all_var_profiling_ws['H' + str(i)].number_format = '0.00%'