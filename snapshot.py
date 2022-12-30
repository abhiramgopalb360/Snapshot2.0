import os
import string
import warnings

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.plotarea import DataTable
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from vyper.utils.tools import StatisticalTools as st

from explorer import DataProfiler  # 100522 customized explorer


warnings.simplefilter(action="ignore", category=pd.core.common.SettingWithCopyWarning)


class Snapshot:
    def __init__(
        self,
        profile_data: pd.DataFrame,
        segment_var: str,
        baseline: str = "",
        segments: list = [],
        filename: str = "Profile",
        continuous_path: str = "",
        nbins: int = 6,
        variable_order: list = [],
        include: list = [],
        exclude: list = [],
        continuous: list = [],
        exclude_other: bool = False,
        na_drop_threshold: float = 0.95,
    ) -> None:

        self.profile_data = profile_data
        self.segment_var = segment_var
        self.filename = filename
        self.nbins = nbins
        self.variable_order = variable_order
        self.include = include
        self.exclude = exclude
        self.continuous = continuous
        self.exclude_other = exclude_other
        self.na_drop_threshold = na_drop_threshold

        if not segments:
            segments = profile_data[segment_var].unique().tolist()
        else:
            all_segs = profile_data[segment_var].unique().tolist()
            for seg in segments:
                if seg not in all_segs:
                    raise ValueError(f"Segment '{seg}' not in '{segment_var}'")
        self.num_segments = len(segments)

        if baseline:
            if baseline not in segments:
                raise ValueError(f"Baseline '{baseline}' not in '{segment_var}'")
            segments.insert(0, segments.pop(segments.index(baseline)))

        self.mapping_dict = {}
        for i, seg in enumerate(segments):
            if i == 0:
                self.mapping_dict["Baseline"] = seg
            else:
                self.mapping_dict[f"Segment_{i}"] = seg

        # get custom bounds for continuous vars
        df_continuous = pd.read_csv(continuous_path)
        # Drop variables that do not have pre-defined bins
        df_continuous.dropna(inplace=True)
        # create bin dictionary for run_profiler()
        self.continuous_var_bounds = {}
        for index, row in df_continuous.iterrows():
            # Read in the row as string and convert to list of floats
            self.continuous_var_bounds[row["Attribute"]] = [float(i) for i in row["Bin"].split(sep=",")]

    def run_profiler(self) -> None:

        self.preprocess()

        for col in self.profile_data:
            try:
                self.profile_data[col] = pd.to_numeric(self.profile_data[col])
            except Exception:
                continue

        varclass = pd.DataFrame()
        for col in self.profile_data.columns:
            varclass = pd.concat([varclass.reset_index(drop=True), pd.DataFrame([st.classify_variable(self.profile_data[col])], columns=[col])], axis=1)
        varclass[varclass.columns in self.continuous] = "continuous"

        if self.include:
            self.exclude = varclass.columns[varclass.columns not in self.include]
        # Variables to exclude
        varclass[varclass.columns in self.exclude] = "exclude"

        if self.exclude_other:
            varclass[varclass.columns == "Categorical+other"] = "exclude"
        continuous_var_cuts = dict()
        self.varclass = varclass

        # binning continuous variable to bins in the binning file
        for variable in varclass.columns[varclass.iloc[0] == "continuous"]:
            if not self.continuous_var_bounds:
                self.continuous_var_bounds = dict()
            if variable.startswith("MT_") or variable.startswith("PROPENSITY_") or variable.startswith("LIKELY_") or variable == "TGT_PRE_MOVER_20_MODEL":
                continuous_var_cuts[variable] = pd.cut(self.profile_data[variable], bins=[0, 5, 25, 45, 65, 85, 99])
            elif variable in self.continuous_var_bounds:
                cut_bins = self.continuous_var_bounds[variable]
                continuous_var_cuts[variable] = pd.cut(self.profile_data[variable], bins=cut_bins)
            else:
                continuous_var_cuts[variable] = st.get_breaks(self.profile_data.loc[:, variable], nbins=self.nbins, squash_extremes=False)

        col_order = []
        index_order = []
        # TODO: array([2, 0, 1]) Always encode 0 as US
        # Change the US population as BASELINE
        for seg in np.sort(list(self.mapping_dict.keys())):
            profile_1 = DataProfiler(self.profile_data[self.profile_data[self.segment_var] == self.mapping_dict[seg]]).create_profile(
                number_of_bins=self.nbins, cts_cuts=continuous_var_cuts
            )

            if seg == "Baseline":
                profile = profile_1
                profile.columns = ["Variable", "Category", "Count " + str(self.mapping_dict[seg]), "Percent " + str(self.mapping_dict[seg])]
                for cols in profile.columns:
                    col_order.append(cols)

            else:
                profile_1.columns = ["Variable", "Category", "Count " + str(self.mapping_dict[seg]), "Percent " + str(self.mapping_dict[seg])]
                col_order.append("Count " + str(self.mapping_dict[seg]))
                col_order.append("Percent " + str(self.mapping_dict[seg]))
                profile = pd.merge(profile, profile_1, on=["Variable", "Category"], how="outer")
                profile[str(self.mapping_dict[seg]) + " vs " + self.mapping_dict["Baseline"]] = (profile["Percent " + str(self.mapping_dict[seg])] / profile["Percent " + self.mapping_dict["Baseline"]]) * 100
                index_order.append(str(self.mapping_dict[seg]) + " vs " + self.mapping_dict["Baseline"])

        profile = profile[col_order + index_order]

        for i in profile.columns:

            if "Count" in i:
                profile[i][profile[i].isna()] = 0
            if "Percent" in i:
                profile[i][profile[i].isna()] = 0

        # Defining PSI
        psi = {}
        uscol = ""
        NumPSI = []
        for i in profile.columns:
            if ("Percent" in i) & (self.mapping_dict["Baseline"] not in i):
                NumPSI.append(i)
            elif ("Percent" in i) & (self.mapping_dict["Baseline"] in i):
                uscol = i

        for index, row in profile.iterrows():
            if row["Variable"] not in psi:
                psi[row["Variable"]] = {}
            if row["Category"] not in psi[row["Variable"]]:
                psi[row["Variable"]][row["Category"]] = {}

            for i in NumPSI:

                try:
                    psi[row["Variable"]][row["Category"]][i + "_PSI"] = (row[uscol] - row[i]) * np.log(row[uscol] / row[i])
                    if psi[row["Variable"]][row["Category"]][i + "_PSI"] == np.inf:
                        psi[row["Variable"]][row["Category"]][i + "_PSI"] = 0
                    psi[row["Variable"]][i + "_Overall"] = 0

                except Exception:
                    psi[row["Variable"]][row["Category"]][i + "_PSI"] = 0

        overall = {}
        for variable in psi:

            for label in psi[variable]:
                if "Overall" in str(label):
                    continue

                for scores in psi[variable][label]:
                    if variable not in overall:
                        overall[variable] = {}
                    if scores not in overall[variable]:
                        overall[variable][scores] = 0
                    overall[variable][scores] += psi[variable][label][scores]

        # Variable overall contains all the overall PSI Scores
        self.overall = overall

        for col in profile.columns[profile.columns.str.contains("Percent")]:
            profile[col] = profile[col] / 100

        self.profile = profile

    def create_profile(self) -> None:
        wb = Workbook()
        del wb["Sheet"]

        profile = self.profile
        profile["Category"] = profile["Category"].astype(str)

        # create sheet
        all_var_profiling_ws = wb.create_sheet("Profile", 0)

        # make new sheet the active sheet we are working on
        all_var_profiling_ws = wb.active

        # remove gridlines from the sheet
        all_var_profiling_ws.sheet_view.showGridLines = False

        # set border line thickness
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # get max rows and cols of profiling df
        max_rows = profile.shape[0]
        max_cols = profile.shape[1]

        input_rows = range(4, max_rows + 4)
        input_cols = list(string.ascii_uppercase)[1: max_cols + 1]

        y = 0
        for i in input_cols:
            x = "3"
            x = i + x

            all_var_profiling_ws[x].font = Font(bold=True)
            current_cell = all_var_profiling_ws[x]
            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")

            if y == 0:
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif y == max_cols - 1:
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
            else:
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

            all_var_profiling_ws[x] = profile.columns[y]
            y = y + 1

        counter = []
        counter_val = 1

        for index, i in enumerate(input_cols[2::]):
            x = "2"
            x = i + x

            if index >= self.num_segments * 2:
                break

            if index % 2 == 1:
                all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
                all_var_profiling_ws[counter[-1]].font = Font(bold=True)
                current_cell = all_var_profiling_ws[counter[-1]]
                current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                continue

            if i == "D":
                all_var_profiling_ws[x] = "BASELINE"
                all_var_profiling_ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
                counter.append(x)
                continue
            all_var_profiling_ws[x] = "SEGMENT " + str(counter_val)
            counter_val += 1
            counter.append(x)
            # all_var_profiling_ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=5)
            # all_var_profiling_ws.merge_cells(f'{x}:D2')

        def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):

            """
            Apply styles to a range of cells as if they were a single cell.

            :param ws:  Excel worksheet instance
            :param range: An excel range to style (e.g. A1:F20)
            :param border: An openpyxl Border
            :param fill: An openpyxl PatternFill or GradientFill
            :param font: An openpyxl Font object
            """

            top = Border(top=border.top)
            left = Border(left=border.left)
            right = Border(right=border.right)
            bottom = Border(bottom=border.bottom)

            first_cell = ws[cell_range.split(":")[0]]
            if alignment:
                ws.merge_cells(cell_range)
                first_cell.alignment = alignment

            rows = ws[cell_range]
            if font:
                first_cell.font = font

            for cell in rows[0]:
                cell.border = cell.border + top
            for cell in rows[-1]:
                cell.border = cell.border + bottom

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = l.border + left
                r.border = r.border + right
                if fill:
                    for c in row:
                        c.fill = fill

        border = Border(top=thick, left=thick, right=thick, bottom=thick)

        for range1 in all_var_profiling_ws.merged_cells.ranges:
            style_range(all_var_profiling_ws, str(range1), border=border)

        y = 0
        for i in input_cols:
            z = 0
            for j in input_rows:
                all_var_profiling_ws.row_dimensions[j].height = 25
                x = str(j)
                x = i + x
                current_cell = all_var_profiling_ws[x]
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                    current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

                all_var_profiling_ws[x] = profile.iloc[z, y]

                if y == 0:
                    current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
                    if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                        current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                elif y == max_cols - 1:
                    current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
                    if profile["Variable"].iloc[z] != profile["Variable"].iloc[z - 1]:
                        current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)

                z = z + 1
            y = y + 1

        self.__allformat(all_var_profiling_ws)

        ws2 = wb.create_sheet("Allcategory")

        psi_df = pd.DataFrame(self.overall).transpose().reset_index()
        psi_df.rename({"index": "Category"}, axis=1, inplace=True)

        rows = dataframe_to_rows(psi_df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)

        self.__merge(wb)

        savepath = f"{self.filename}_Profile.xlsx"
        wb.save(savepath)

    def __allformat(self, ws) -> None:

        char = "C"
        for _ in range(self.num_segments):
            char = chr(ord(char) + 2)
            for col in ws[char]:
                col.number_format = "0%"

        start = chr(ord(char) + 1)
        for _ in range(1, self.num_segments):
            char = chr(ord(char) + 1)
            for col in ws[char]:
                col.number_format = "#,#0"

        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 32
        for c in range(ord("D"), ord(char) + 1):
            ws.column_dimensions[chr(c)].width = 16

        # Headers Alignment
        for row in ws.iter_rows(min_col=2, max_col=11, min_row=3, max_row=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Columns
        for row in ws.iter_rows(min_col=2, max_col=14):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

            # Color scale
        red = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
        yellow = PatternFill(start_color="00FFCC00", end_color="00FFCC00", fill_type="solid")
        green = PatternFill(start_color="0099CC00", end_color="0099CC00", fill_type="solid")
        # white = PatternFill(start_color="00EEEEEE", end_color="00EEEEEE", fill_type="solid")

        dxf1 = DifferentialStyle(fill=red)
        dxf2 = DifferentialStyle(fill=yellow)
        dxf3 = DifferentialStyle(fill=green)
        # dxf4 = DifferentialStyle(fill=white)

        rule1 = Rule(type="cellIs", operator="between", formula=[0.0, 85.0], dxf=dxf1)
        rule2 = Rule(type="cellIs", operator="between", formula=[85.0, 115.0], dxf=dxf2)
        rule3 = Rule(type="cellIs", operator="between", formula=[115.0, 1000000.0], dxf=dxf3)
        # rule4 = Rule(type="cellIs", operator="equal", formula=[0], dxf=dxf4)

        final_row = ws.max_row

        rule_string = f"{start}4:{char}{final_row}"

        ws.conditional_formatting.add(rule_string, rule1)
        ws.conditional_formatting.add(rule_string, rule2)
        ws.conditional_formatting.add(rule_string, rule3)

    def create_visual(self, plot_index=True, data_table=False, show_axes=True, show_na=True) -> None:

        self.plot_index = plot_index
        self.data_table = data_table
        self.show_axes = show_axes

        profile = self.profile
        lcn = str(profile.columns[-1])
        profile_sliced = profile[profile[lcn] > 0]
        wb = Workbook()

        def shorten_name(name):
            if len(name) >= 31:
                name = name[0:31]
            return name

        profile_sliced[profile_sliced["Variable"] == "VEHICLE_MAKE_4"]
        profile_sliced["Variable"] = profile_sliced["Variable"].apply(lambda x: shorten_name(x))
        profile_sliced["Category"] = profile_sliced["Category"].astype(str)

        all_var_profiling_ws = wb.create_sheet(profile_sliced["Variable"].iloc[0], 0)
        # make new sheet the active sheet we are working on
        all_var_profiling_ws = wb.active

        # remove gridlines from the sheet
        all_var_profiling_ws.sheet_view.showGridLines = False

        # set border line thickness
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # get max rows and cols of profiling df
        max_rows = profile_sliced.shape[0]
        max_cols = profile_sliced.shape[1]

        input_cols = list(string.ascii_uppercase)
        from openpyxl.utils.dataframe import dataframe_to_rows

        Allrows = dataframe_to_rows(profile)

        for z in range(0, len(profile_sliced)):
            current_cat = profile["Variable"].iloc[z]

            if z != 0:
                if profile_sliced["Variable"].iloc[z] != profile_sliced["Variable"].iloc[z - 1]:
                    # create sheet
                    all_var_profiling_ws = wb.create_sheet(profile_sliced["Variable"].iloc[z], 0)
                    # make new sheet the active sheet we are working on
                    all_var_profiling_ws = wb.active

                    for y in range(len(profile_sliced.columns)):
                        current_cell = all_var_profiling_ws.cell(row=3, column=y + 2)

                        current_cell.font = Font(bold=True)
                        current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                        current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                        current_cell.value = profile_sliced.columns[y]
                        if y == 0:
                            current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                        elif y == max_cols - 1:
                            current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
                        else:
                            current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
                else:
                    incol = 2
                    for y in range(len(profile_sliced.columns)):
                        current_cell = all_var_profiling_ws.cell(row=3, column=incol)
                        incol = incol + 1
                        current_cell.font = Font(bold=True)
                        current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                        current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                        current_cell.value = profile.columns[y]
                        if y == 0:
                            current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
                        elif y == max_cols - 1:
                            current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
                        else:
                            current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)

        del wb["Sheet"]
        for col in wb.sheetnames:
            temp_df = profile_sliced.loc[profile_sliced["Variable"] == col, :]
            # To code for thresold 3%
            percent_cols = []
            for i in temp_df.columns:
                if "Percent" in i:
                    percent_cols.append(i)

            if sum(temp_df["Category"].isin(["NA", "99", 99, "Z"])):
                # or min
                k = (temp_df[temp_df["Category"].isin(["NA"])][percent_cols] >= self.na_drop_threshold).any(axis=1).astype(bool)

                if sum(k):
                    # delete that sheet
                    std = wb[col]
                    wb.remove(std)
                    # print(f"{std} removed")
                    continue

            # temp_df["FLAG"] = (temp_df[percent_cols] >= 0.001).any(axis=1).astype(bool)
            temp_df.loc[:, "FLAG"] = (temp_df[percent_cols] >= 0.001).any(axis=1).astype(bool)
            temp_df = temp_df[temp_df["FLAG"]]

            if not show_na:
                temp_df = temp_df[~temp_df["Category"].isin(["NA", "99", 99, "Z"])]

            if len(temp_df) == 0:
                # delete that sheet
                std = wb[col]
                wb.remove(std)
                # print(f"{std} removed")
                continue
            # To code if temp_df contains NA in

            temp_df = temp_df.drop("FLAG", axis=1)

            #
            rows = dataframe_to_rows(temp_df)
            wb.active = wb[col]
            ws = wb.active

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    ws.cell(row=r_idx + 1, column=c_idx + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

            all_var_profiling_ws = ws
            all_var_profiling_ws.move_range("B3:Z3", cols=1)
            all_var_profiling_ws.delete_rows(2)
            all_var_profiling_ws.delete_cols(2)
            all_var_profiling_ws.insert_rows(2)

            counter = []
            counter_val = 1
            input_cols = list(string.ascii_uppercase)[1: max_cols + 1]

            for index, i in enumerate(input_cols[2::]):
                x = "2"
                x = i + x

                if index >= self.num_segments * 2:
                    break

                if index % 2 == 1:
                    all_var_profiling_ws.merge_cells(f"{counter[-1]}:{x}")
                    all_var_profiling_ws[counter[-1]].font = Font(bold=True)
                    current_cell = all_var_profiling_ws[counter[-1]]
                    current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                    current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                    continue

                if i == "D":
                    all_var_profiling_ws[x] = "BASELINE"
                    all_var_profiling_ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
                    counter.append(x)
                    continue
                all_var_profiling_ws[x] = "SEGMENT " + str(counter_val)
                counter_val += 1
                counter.append(x)

            # Merge cells A and B test
            key_column = 2
            merge_columns = [2]
            start_row = 4
            max_row = ws.max_row
            key = None

            # Iterate all rows in `key_colum`
            for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
                if key != row_cells[0].value or row == max_row:
                    if key is not None:
                        for merge_column in merge_columns:
                            ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row, end_column=merge_column)
                            ws.cell(row=start_row, column=merge_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        start_row = row
                    key = row_cells[0].value
                if row == max_row:
                    row += 1

            self.__visual(ws)
            self.__allformat(ws)

        # TODO: format Index page
        all_var_profiling_ws1 = wb.create_sheet("Index", 0)
        # make new sheet the active sheet we are working on
        wb.active = wb["Index"]
        all_var_profiling_ws1 = wb.active
        for i in wb.sheetnames:
            all_var_profiling_ws1.append([i])

        filesave = f"{self.filename}_Chart.xlsx"
        wb.save(filesave)

    def __visual(self, ws) -> None:
        def color_palette(n):
            if n == 2:
                return ["003f5c", "ffa600"]
            elif n == 3:
                return ["003f5c", "bc5090", "ffa600"]
            elif n == 4:
                return ["003f5c", "7a5195", "ef5675", "ffa600"]
            elif n == 5:
                return ["003f5c", "58508d", "bc5090", "ff6361", "ffa600"]
            else:
                box = []
                while len(box) < n:
                    box += ["003f5c", "7a5195", "ef5675", "ffa600"]
                return box[:n]

        c1 = BarChart()
        c1.height = 16  # default is 7.5
        c1.width = 30  # default is 15
        if self.data_table:
            c1.plot_area.dTable = DataTable()
            c1.plot_area.dTable.showHorzBorder = True
            c1.plot_area.dTable.showVertBorder = True
            c1.plot_area.dTable.showOutline = True
            c1.plot_area.dTable.showKeys = True
            c1.legend = None
        else:
            c1.dataLabels = DataLabelList()
            c1.dataLabels.showVal = True
            c1.legend.position = "b"

        colors = color_palette(self.num_segments)

        for seg in range(self.num_segments):
            x1 = seg * 2 + 5
            data = Reference(ws, min_col=x1, min_row=3, max_row=ws.max_row, max_col=x1)
            cats = Reference(ws, min_col=3, min_row=4, max_row=ws.max_row, max_col=3)
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(cats)
            c1.shape = 4
            c1.series[seg].graphicalProperties.solidFill = colors[seg]

        if self.show_axes:
            c1.x_axis.title = "Categories"
            c1.y_axis.title = "Percentage"
        c1.y_axis.majorGridlines = None
        c1.title = ws["B4"].value

        # Create a second chart
        if self.plot_index:
            c2 = LineChart()

            for seg in range(1, self.num_segments):
                x2 = seg + x1

                data = Reference(ws, min_col=x2, min_row=3, max_row=ws.max_row, max_col=x2)
                cats = Reference(ws, min_col=3, min_row=4, max_row=ws.max_row, max_col=3)
                c2.add_data(data, titles_from_data=True)
                c2.set_categories(cats)

            c2.y_axis.axId = 200
            c2.y_axis.title = "Index"

            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            c2.y_axis.crosses = "max"
            c1 += c2

        ws.add_chart(c1, "D15")

    def __merge(self, wb) -> Workbook:
        # Selecting active sheet
        ws = wb.active

        # Merge cells A and B test
        key_column = 2
        merge_columns = [2]
        start_row = 4
        max_row = ws.max_row
        key = None

        # Iterate all rows in `key_colum`
        for row, row_cells in enumerate(ws.iter_rows(min_col=key_column, min_row=start_row, max_col=key_column, max_row=max_row), start_row):
            if key != row_cells[0].value or row == max_row:
                if key is not None:
                    for merge_column in merge_columns:
                        ws.merge_cells(start_row=start_row, start_column=merge_column, end_row=row - 1, end_column=merge_column)
                        ws.cell(row=start_row, column=merge_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 1011
                    start_row = row
                key = row_cells[0].value
            if row == max_row:
                row += 1

        return wb

    def preprocess(self) -> None:
        # TODO add preprocessing for envision and acxiom fields
        self.profile_data = pd.DataFrame(self.profile_data)