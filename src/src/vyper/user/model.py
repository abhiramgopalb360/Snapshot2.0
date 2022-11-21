import math
import warnings
from statistics import mean
from typing import List, Optional

import numpy as np
import openpyxl
import pandas as pd
import statsmodels.api as sm
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, colors
from scipy.cluster.hierarchy import ward, fcluster, inconsistent
from scipy.spatial.distance import pdist
from scipy.stats.mstats import zscore
from sklearn.metrics import roc_auc_score, r2_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.feature_selection import RFE
from sklearn import linear_model

from vyper.base import Variable, VariableCollection
from vyper.user.explorer import DataProfiler
from vyper.utils.tools import DataTools as dt
from vyper.utils.tools import StatisticalTools as st

warnings.filterwarnings("ignore")


class Model:

    def __init__(self,
                 data: pd.DataFrame,
                 dependent_variable: str,
                 excluded_variables: set = None,
                 drop_if_dependent_na: bool = True,
                 na_drop_threshold: float = 0.95,
                 training_percentage: float = 0.7,
                 model_type: str = 'linear'):
        """
        Collection of tools for data exploration. This tool allows users to tackle modeling tasks throughout the entire
        process in a timely manner, including variable profiling & initiation, data preprocessing,
        model build & tuning, and out of time data scoring.

        :param data: pd.DataFrame, the input data frame
        :param dependent_variable: str, the name of the dependent variable
        :param excluded_variables: set, the names of the variables to be excluded
        :param drop_if_dependent_na: bool, option to drop the null values of the dependent variable or not
        :param na_drop_threshold: float, if percentage of null value of a variable is larger than this threshold,
        this variable will be excluded
        :param training_percentage: float, the percentage of data the user want to spilt as training data
        :param model_type: str, the type of the model the user want to run, default to 'linear'. If number of unique
        value of the dependent variable equals two, it will be overwrite as 'logistic'
        """
        self.na_drop_threshold = na_drop_threshold
        if not excluded_variables:
            self.excluded_variables = set()
        else:
            if not isinstance(excluded_variables, set):
                self.excluded_variables = set(excluded_variables)
            else:
                self.excluded_variables = excluded_variables

        self.data = data

        if self.excluded_variables:
            for var in self.excluded_variables:
                if var in self.data.columns:
                    self.data.drop(columns=var, inplace=True)

        self.dependent_variable = dependent_variable
        # TODO come back and give user access back
        self.variables = VariableCollection()

        self.training_percentage = training_percentage

        missing_dep_vars = self.data[dependent_variable].isna().sum()
        if missing_dep_vars > 0:
            if not drop_if_dependent_na:
                raise ValueError(f"{missing_dep_vars} dependent variables were null, drop_if_dependent_na"
                                 f"is set to false. Cannot proceed with model init")
            else:
                self.data = self.data.dropna(subset=[dependent_variable])
                warnings.warn(f"{missing_dep_vars} NAs were found in dependent variable; "
                              f"{missing_dep_vars} rows were dropped.")

        # TODO see if this should really be automatic
        self.auto_define_variables()

        self.training_data = None
        self.testing_data = None

        self.epsilon_dropped_columns = None

        self.corr_iv = None
        self.train_out = None
        self.test_out = None
        self.model_descriptor = None
        self.model_metric = None
        self.train_lift = None
        self.test_lift = None
        self.wb = None
        self.var_scores = None

        # TODO check with Abid on how to detect Epsilon datasets
        # TODO code separate data processing package for dealing specifically with Epsilon
        # if source_data.lower() == 'epsilon' or dt.find_eps_col(self.data):
        #     self.excluded_variables.update(dt.get_excluded_epsilon_columns())
        #     self.process_epsilon_columns()
        #     self.excluded_variables.update('ind')
        #     self.excluded_variables.update(self.epsilon_recoded_columns)

        if not model_type:
            if self.data[dependent_variable].nunique() == 2:
                self.model_type = 'logistic'
            else:
                self.model_type = 'linear'
        else:
            self.model_type = model_type

        self.epsilon_recoded_columns = None

        self.var_selected = None
        self.variable_reduction_method_values = None

        # underlying model class for storing statsmodel implementation
        self._model = None
        self.tni_processed = False

        self.original_variables = set(self.data.columns.to_list())

    def create_custom_profile(self, variable_list: list):
        """
        Def: create profiling output of a dataframe with a given list of variables
        :param variable_list: list, variable names that the user want to create a profile on
        :return: pandas dataframe, profiling result
        """
        return DataProfiler(self.get_training_data(), self.dependent_variable,
                            variables=self.variables).create_custom_profile(variable_list)

    def process_epsilon_columns(self):
        self.data, self.epsilon_recoded_columns = dt.epsilon_batch_recoder(dt(), self.data)

    def auto_define_variables(self,
                              categorical_threshold: int = 100,
                              string_threshold: int = 100,
                              overwrite: bool = False) -> None:
        """
        Def: This function will classify the type of each variable. There are five types of variables:
        'exclude', 'continuous', 'binary', 'categorical'
        (categorical variable exceeds number of max_categories). User defined excluded variables will be assigned as
        'excluded', User defined dependent variable will be assigned as 'target_variable'.

        :param categorical_threshold: int, if number of categories of a variable is larger than this threshold then
        classify as 'categorical+other'
        :param string_threshold: int, variable which is string in nature and number of distinct values is more than this
        threshold then classify as 'exclude'
        :param overwrite: bool, option of overwriting the variable kind or not
        """
        if self.variables:
            if not overwrite:
                raise AttributeError('define variables is called without overwrite, but model already has variables '
                                     'defined')
            else:
                self.variables = VariableCollection()

        for var in self.data.columns:
            if var != self.dependent_variable:
                if var in self.excluded_variables:
                    self.variables[var] = Variable(name=var,
                                                   kind='exclude')
                # self.variables[var] = Variable(name=var,
                #                                kind=st.
                #                                classify_variable(self.data.loc[:, var],
                #                                                  distinct_values_threshold=distinct_values_threshold,
                #                                                  max_values=max_values,
                #                                                  max_categories=max_categories,
                #                                                  missing_threshold=self.na_drop_threshold))
                self.variables[var] = Variable(name=var,
                                               kind=st.
                                               classify_variable_v2(vector=self.data.loc[:, var],
                                                                    categorical_threshold=categorical_threshold,
                                                                    string_threshold=string_threshold,
                                                                    missing_exclude_threshold=self.na_drop_threshold))
            else:
                self.variables[var] = Variable(name=var,
                                               kind='target_variable')

    @property
    def as_pandas(self):
        """
        placeholder for when other dataframes are used
        :return:
        """
        return self.data

    def get_training_data(self, random_state=1):
        if not isinstance(self.training_data, pd.DataFrame):
            self.generate_train_test_split(random_state=random_state)
        return self.training_data

    def get_testing_data(self, random_state=1):
        if not isinstance(self.testing_data, pd.DataFrame):
            self.generate_train_test_split(random_state=random_state)
        return self.testing_data

    def generate_train_test_split(self, random_state=1):
        self.training_data, self.testing_data = train_test_split(self.data,
                                                                 train_size=self.training_percentage,
                                                                 random_state=random_state
                                                                 )
        self.training_data.reset_index(inplace=True)
        self.testing_data.reset_index(inplace=True)

    def set_variable_types(self, variable_dt=dict) -> None:
        for k, v in variable_dt.items():
            self.variables[k].kind = v

    def tni_smart(self,  min_bin_size=0.1,
                  method='index', binary_var=True, bv_min_size=0.1, bv_min_incindex=0.1,
                  other_transform=True, bounds=.98, transformations=('Inv', 'Sqrt', 'Exp', 'Pw2', 'Log'),
                  all_cat=False, random_state=1) -> None:
        """
        Def: Transforms and imputes model variables by variable type for the given training data.
        Train and test data will be specified if not already done so. This will generate the TnI
        functions that can be used in tni_transform to process the testing or out of time data

        :param min_bin_size: Minimum size for Bins
        :param method: Method for Continuous Imputation; Must be ['mean', 'median', 'index', 'index_median']
        :param binary_var: Returns Significant Binary Variables if True
        :param bv_min_size: Minimum Binary Variable Size to be Significant
        :param bv_min_incindex: Minimum Y Index for Binary Variables to be Significant
        :param other_transform: Returns Transformations in transformation if True
        :param bounds: Bounds for Capping Continuous Variables
        :param transformations: Transformations to be returned;
            Must be a subset of: ('Inv', 'Sqrt', 'Exp', 'Pw2', 'Log')
        :param all_cat: Selects Categorical Variables based on each Category if True
        :param random_state: Random state for generating train/test split
        """
        # TODO handle of 'binary' variable type

        if self.tni_processed:
            self.generate_train_test_split(random_state=random_state)
#            self.auto_define_variables(overwrite=True)  # remove if auto_define_variables stops being auto defined

        binary_variables = list()
        categorical_variables = list()
        continuous_variables = list()
        mapping_dt = {'continuous': continuous_variables,
                      'categorical': categorical_variables,
                      'binary': binary_variables,
                      }

        for k, v in self.variables.items():
            if v.kind in mapping_dt:
                mapping_dt[v.kind].append(k)

        # other_columns = [i for i in self.variables if i not in (binary_variables +
        #                                                         continuous_variables + categorical_variables)]

        for col in binary_variables:
            self.tni_binary(col)

        for col in continuous_variables:
            self.tni_continuous(col,
                                min_bin_size=min_bin_size,
                                method=method,
                                binary_var=binary_var,
                                bv_min_size=bv_min_size,
                                bv_min_incindex=bv_min_incindex,
                                other_transform=other_transform,
                                bounds=bounds,
                                transformations=transformations)
        for col in categorical_variables:
            if all_cat:
                self.tni_categorical(col,
                                     min_bin_size=0,
                                     binary_var=True,
                                     bv_min_size=0,
                                     bv_min_incindex=0)
            else:
                self.tni_categorical(col,
                                     min_bin_size=min_bin_size,
                                     binary_var=binary_var,
                                     bv_min_size=bv_min_size,
                                     bv_min_incindex=bv_min_incindex)
        self.tni_processed = True

    def tni_binary(self,
                   variable_name: str,
                   missing_threshold: float = 0.0) -> None:
        """
        Def: Transform binary data columns

        :param variable_name: Binary Data Columns to be Transformed
        :param missing_threshold: Minimum Threshold for creating a binary variable for missing data
        """
        # TODO don't really understand the point of a missing threshold here.
        if self.get_training_data()[variable_name].isna().sum() > len(self.data) * missing_threshold:
            self.training_data.fillna(value={variable_name: 'TnImissing'})

        _dummy_df = pd.get_dummies(self.training_data[variable_name],
                                   prefix=variable_name,
                                   dummy_na=False,
                                   columns=[variable_name]
                                   )
        # self.training_data = self.training_data.join(_dummy_df)
        for x in _dummy_df:
            self.training_data[x] = _dummy_df[x]

        # get encoding for binary variables
        # TODO check for redundancy, since no transformations are really done on binary variables
        encoded_mapping = {i: i.split(f"{variable_name}_")[1] for i in _dummy_df.columns}
        self.variables[variable_name].set_tni_encoded_column_mapping(encoded_mapping)

    def tni_categorical(self, x, min_bin_size=0.1, binary_var=True, bv_min_size=0.1, bv_min_incindex=0.1) -> None:
        """
        Def: Transform categorical data columns to binary format based on its values, so that regression algo can
        directly leverage.

        :param x: Categorical Data Column to be Transformed
        :param min_bin_size: Minimum size for Bins (Note: In cate var, each value represent a unique bin)
        :param binary_var: Returns Significant Binary Variables if True (Note: Binary var created based on bins)
        :param bv_min_size: Minimum Binary Variable Size to be Significant
        :param bv_min_incindex: Minimum Y Index for Binary Variables to be Significant

        """
        testdata = self.get_training_data().loc[:, [x, self.dependent_variable]]

        # TODO evaluate usefulness of missing flag
        # if missing_flag:
        #     # if testdata['x_royalias'].isna().mean() > 0:  # do we need mean here or count would do?
        #     #
        #     testdata['x_missing_flag'] = [1 if ct == True else 0 for ct in list(pd.isnull(testdata['x_royalias']))]
        #     #     lookup_table = pd.concat([lookup_table, pd.DataFrame(
        #     #         {'var_name': x + '_missing_flag', 'original_name': x, 'type': 'missing_flag'}, index=[1])])
        #     #     testdata = testdata.rename(columns={'x_missing_flag': x + '_missing_flag'})
        #     # else:
        #     #     if not bar:
        #     #         print(
        #     #
        #     'there is no missing in independent variable, no inputing was done, no missing flag was created')
        #     pass
        # TODO remove redundant renaming
        testdata['x_royalias'] = testdata[x]
        testdata['y_royalias'] = testdata[self.dependent_variable]
        testdata['x_royalias'].fillna('TnImissing', inplace=True)  # to avoid issue when joining by this var
        testlookup = testdata[['y_royalias', 'x_royalias']].groupby(['x_royalias'], dropna=False).mean().reset_index()
        testlookup = testlookup.rename(columns={'y_royalias': 'y_royindex'})
        testlookup = testlookup.sort_values(by='y_royindex', ascending=True)
        testlookup['x_royassign'] = list(range(1, (len(testlookup) + 1)))
        # testlookup.drop(['y_royindex1'], axis=1, inplace=True)
        self.variables[x].set_tni_value_mapping({var: assigned for var, assigned in zip(testlookup['x_royalias'],
                                                                                        testlookup['x_royassign'])})

        testdata = testdata.merge(testlookup, how='left', on='x_royalias')

        # TODO has no handling in original code
        if len(testlookup) == 1:
            raise ValueError('With bounds, only one unique value')

        else:
            counter2 = 0
            tab = testdata['x_royassign'].value_counts().sort_index()
            cut_point = np.full((len(tab) - 1), np.nan)

            for ii in range(0, (len(tab) - 1)):
                counter1 = tab.iloc[ii]
                counter2 = counter1 + counter2
                if counter2 / sum(testdata['x_royassign'].isna()) >= min_bin_size:
                    cut_point[ii] = int(tab.index[ii])
                    counter2 = 0
            del counter1, counter2, ii
            cut_point = np.array([x for x in list(cut_point) if x is not np.nan])
            if sum(testdata['x_royassign'] > cut_point[-1]) / sum(testdata['x_royassign'].notnull()) < min_bin_size:
                cut_point = cut_point[:-1]

        cut_point = [-np.inf] + list(cut_point) + [np.inf]
        cut_point = list(set(cut_point))
        cut_point.sort()

        if min_bin_size == 0:
            cut_labels = testlookup['x_royalias'].values
        else:
            cut_labels = ['(' + str(m) + '_' + str(n) + ']' for m, n in zip(cut_point[:-1], cut_point[1:])]

        testdata['x_roybins_bi'] = pd.cut(testdata['x_royassign'], cut_point, right=True, labels=cut_labels)
        testdata['x_roybins_bi'] = testdata['x_roybins_bi'].astype(str)
        # save cutter for easy reproduction later on
        self.variables[x].get_or_create_tni().cutter = {'bins': cut_point,
                                                        'right': True,
                                                        'labels': cut_labels}

        # declared here for use when assigning variable attributes
        replacement_dt = dict()
        if binary_var:
            temp_x_roybins_bi = testdata['x_roybins_bi']
            original_columns = set(testdata.columns.to_list())
            testdata = pd.get_dummies(testdata, columns=['x_roybins_bi'])

            # replacements declared here to go into variable metadata
            # TODO possibly not needed since information is available as part of column name
            dummy_columns = set(testdata.columns.to_list()) - original_columns
            replacement_dt = {i: i.replace('x_roybins', x) for i in dummy_columns}
            testdata = testdata.rename(columns=replacement_dt)

            _encoded_bins = {v: v.split('_bi_')[1] for k, v in replacement_dt.items()}

            testdata['x_roybins_bi'] = temp_x_roybins_bi
            del temp_x_roybins_bi

            drop2 = []

            for ii in range((len(testdata.columns) - len(testdata['x_roybins_bi'].unique()) - 1),
                            (len(testdata.columns) - 1)):
                if testdata.iloc[:, ii].mean() < bv_min_size:
                    drop2.append(ii)
                else:
                    if abs(testdata['y_royindex'][testdata.iloc[:, ii] == 1].mean() /
                           testdata['y_royindex'].mean() - 1) < bv_min_incindex:
                        drop2.append(ii)

            if len(drop2) == 0:
                warnings.warn('all binary variable taken are significant')
            else:
                warnings.warn(
                    str(len(testdata['x_roybins_bi'].unique()) - len(
                        drop2)) + ' significant binary variable was taken')

                # remove items from _encoded dict bin for storage in variable class
                for col in list(testdata.columns[drop2]):
                    _encoded_bins.pop(col, None)
                testdata = testdata.drop(list(testdata.columns[drop2]), axis='columns')

            self.variables[x].get_or_create_tni()._categorical['drop_columns'] = drop2

            del drop2, ii
        bin_col = f"{x}_bins"
        tni_columns = {'x_royassign': f"{x}_tni_assign",
                       'x_roybins_bi': bin_col,
                       'y_royindex': f"{x}_dependent_variable_index"}
        testdata = testdata.rename(columns=tni_columns)

        self.variables[x].set_tni_column_mapping({'tni_processed': f"{x}_tni_assign",
                                                  'tni_assigned_bin': bin_col})

        # select only useful columns to keep
        main_columns = [v for k, v in tni_columns.items()]
        binary_columns = [col for col in testdata.columns if f'{x}_bi_' in col]
        testdata = testdata[main_columns + binary_columns]

        self.variables[x].set_tni_encoded_column_mapping(_encoded_bins)

        # merging on index only right now, change to join in future if not using pandas df
        # self.training_data = self.training_data.join(testdata)
        for x in testdata:
            self.training_data[x] = testdata[x]

    def tni_continuous(self,
                       x,
                       min_bin_size=0.1,
                       method='index',
                       binary_var=True,
                       bv_min_size=0.1,
                       bv_min_incindex=0.1,
                       other_transform=True,
                       bounds=.98,
                       transformations=('Inv', 'Sqrt', 'Exp', 'Pw2', 'Log')) -> None:
        """
        Def: Transform continuous data columns

        :param x: Continuous Data Column to be Transformed
        :param min_bin_size: Minimum size for Bins
        :param method: Method for Continuous Imputation; Must be ['mean', 'median', 'index', 'index_median']
        :param binary_var: Returns Significant Binary Variables if True (Note: Binary var created based on bins)
        :param bv_min_size: Minimum Binary Variable Size to be Significant
        :param bv_min_incindex: Minimum Y Index for Binary Variables to be Significant
        :param other_transform: Returns Transformations in transformation if True
        :param bounds: Bounds for Capping Continuous Variables
        :param transformations: Transformations to be returned; Must be a subset of:
            ('Inv', 'Sqrt', 'Exp', 'Pw2', 'Log')
        """
        testdata = self.get_training_data().loc[:, [x, self.dependent_variable]]

        testdata['x_royalias'] = testdata[x]
        testdata['y_royalias'] = testdata[self.dependent_variable]
        if method not in ['index', 'index_med', 'mean', 'median']:
            raise SyntaxError(f'method needs to be index, mean or median,'
                              f'instead it is {method}')
        bnds = st.quantile_function([x for x in list(testdata['x_royalias']) if not math.isnan(x)], bounds=bounds)

        self.variables[x].get_or_create_tni().upper_bound = bnds[1]
        self.variables[x].get_or_create_tni().lower_bound = bnds[0]

        testdata['x_royalias'] = pd.Series(
            np.minimum(np.maximum(testdata['x_royalias'], bnds[0]), bnds[1]))

        tab = testdata['x_royalias'].value_counts(sort=False)
        n = len(testdata['x_royalias'].isna())

        i = 0
        j = 0
        cut_point = []

        b = tab.cumsum() / n

        while i < 1:
            a = list(b[b > min_bin_size + i].index)
            if len(a) >= 1:
                cut_point.insert(j, a[0])  # THIS COULD CAUSE ISSUES ON DIFFERENT DATASET
                i = b[a[0]]
                j = j + 1
            else:
                i = 1
                j = j - 1

        cut_point = cut_point[0:j]

        cut_point = [-np.inf] + cut_point + [bnds[1]]
        cut_point = list(set(cut_point))
        cut_point.sort()

        cut_labels = ['(' + str(m) + '_' + str(n) + ']' for m, n in zip(cut_point[:-1], cut_point[1:])]
        testdata['x_roybins_bi'] = pd.cut(testdata['x_royalias'], bins=cut_point, right=True, labels=cut_labels)

        self.variables[x].get_or_create_tni().cutter = {'bins': cut_point,
                                                        'right': True,
                                                        'labels': cut_labels}

        testdata['x_roybins_bi_factor'] = testdata['x_roybins_bi']
        testdata['x_roybins_bi'] = testdata['x_roybins_bi'].astype(str)

        testdata = testdata.merge(
            testdata.groupby(['x_roybins_bi'], dropna=False)[['y_royalias']].mean().reset_index().rename(
                columns={'y_royalias': 'y_royindex'}), how='left', on='x_roybins_bi')

        # TODO loc instead of slice
        testdata['y_roydistance'] = abs(
            testdata['y_royindex'] - testdata['y_royindex'][testdata['x_royalias'].isna()].mean())
        testdata['y_roydistance'][testdata['x_royalias'].isna()] = max(testdata['y_roydistance']) + 1

        # moved variable tni_assignment into one single statement
        _filler = None
        if testdata['x_royalias'].isna().mean() > 0:
            if method == 'index':
                _filler = testdata['x_royalias'][
                    testdata['y_roydistance'] == testdata['y_roydistance'][
                        testdata['y_roydistance'] > 0].min()].mean()

                testdata['x_royalias'][testdata['x_royalias'].isna()] = _filler

            if method == 'index_med':
                # TODO replaces with list. Is this correct behaviour
                _filler = st.quantile_function(
                    testdata['x_royalias'][
                        testdata['y_roydistance'] == testdata['y_roydistance'][testdata['y_roydistance'] > 0].min()],
                    .5)

            if method == 'mean':
                _filler = testdata['x_royalias'].mean()

            if method == 'median':
                _filler = testdata['x_royalias'].median()
            testdata['x_roybins_bi'][testdata['x_roybins_bi'].isna()] = 'missing'
            drop1 = ['y_roydistance', 'y_royalias']
        else:
            if 'med' in method:
                _filler = testdata['x_royalias'].median()
            else:
                _filler = testdata['x_royalias'].mean()

            drop1 = ['y_roydistance', 'y_royalias', 'x_royalias', 'x_missing_flag']

        # moved substituing missing values into a single fucntion call here for cleaner code
        testdata['x_royalias'][testdata['x_royalias'].isna()] = _filler
        self.variables[x].get_or_create_tni().value_mapping[None] = _filler

        if binary_var:
            temp_x_roybins_bi = testdata['x_roybins_bi']
            testdata = pd.get_dummies(testdata, columns=['x_roybins_bi'])
            testdata['x_roybins_bi'] = temp_x_roybins_bi
            del temp_x_roybins_bi

            drop2 = []

            for ii in range((len(testdata.columns) - len(testdata['x_roybins_bi'].unique()) - 1),
                            (len(testdata.columns) - 1)):
                if testdata.iloc[:, ii].mean() < bv_min_size:
                    drop2.append(ii)

                else:
                    if abs(testdata['y_royindex'][testdata.iloc[:, ii] == 1].mean() /
                           testdata['y_royindex'].mean() - 1) < bv_min_incindex:
                        drop2.append(ii)

            if len(drop2) != 0:
                testdata = testdata.drop(list(testdata.columns[drop2]), axis='columns')
            del drop2, ii

        testdata = testdata[[col for col in testdata.columns if col not in ['x_roybins_bi']]]

        if other_transform:

            if 'Pw2' in transformations:
                trn = testdata['x_royalias'] ** 2
                if not np.isinf(trn).any() and not trn.isnull().any():
                    testdata['x_roytran_TnI_pw2'] = testdata['x_royalias'] ** 2
                    self.variables[x].get_or_create_tni().column_mapping['tni_transformation_pw2'] = x + '_TnI_pw2'

            if str(0) not in testdata['x_royalias'] and 'Inv' in transformations:
                trn = 1 / testdata['x_royalias']
                if not np.isinf(trn).any() and not trn.isnull().any():
                    testdata['x_roytran_TnI_inv'] = 1 / testdata['x_royalias']
                    self.variables[x].get_or_create_tni().column_mapping['tni_transformation_inv'] = x + '_TnI_inv'

            if bnds[0] > 0 and 'Sqrt' in transformations:
                trn = testdata['x_royalias'] ** .5
                if not np.isinf(trn).any() and not trn.isnull().any():
                    testdata['x_roytran_TnI_sqrt'] = testdata['x_royalias'] ** .5
                    self.variables[x].get_or_create_tni().column_mapping['tni_transformation_sqrt'] = x + '_TnI_sqrt'

            if bnds[0] > 0 and 'Log' in transformations:
                trn = np.log(testdata['x_royalias'])
                if not np.isinf(trn).any() and not trn.isnull().any():
                    testdata['x_roytran_TnI_log'] = np.log(testdata['x_royalias'])
                    self.variables[x].get_or_create_tni().column_mapping['tni_transformation_log'] = x + '_TnI_log'

            if 'Exp' in transformations:
                trn = np.exp((testdata['x_royalias'] - testdata['x_royalias'].min()) / (
                            testdata['x_royalias'].max() - testdata['x_royalias'].min()))
                if not np.isinf(trn).any() and not trn.isnull().any():
                    testdata['x_roytran_TnI_exp'] = np.exp((testdata['x_royalias'] - testdata['x_royalias'].min()) / (
                            testdata['x_royalias'].max() - testdata['x_royalias'].min()))
                    self.variables[x].get_or_create_tni().column_mapping['tni_transformation_exp'] = x + '_TnI_exp'

            testdata.columns = [c if 'x_roytran' not in c else c.replace('x_roytran', x) for c in testdata.columns]

        self.variables[x].get_or_create_tni().column_mapping['tni_assigned_bin'] = x + '_bins'

        testdata = testdata.rename(columns={'x_roybins_bi_factor': (x + '_bins')})

        idx = [testdata.columns.get_loc(cols) for cols in testdata.columns if cols.startswith('x_roybins')]

        # TODO check if necessary, bin information should be preserved within bin assignment
        # for ii in idx:
        #     rge = [testdata['x_royalias'][testdata.iloc[:, ii] == 1].min(),
        #            testdata['x_royalias'][testdata.iloc[:, ii] == 1].max()]
        #     if rge[0] == bnds[0]:
        #         up = -np.Inf
        #     else:
        #         up = testdata['x_royalias'][(testdata.iloc[:, ii] == 0) & (testdata['x_royalias'] < rge[0])].max()
        #
        #     # lookup_table_bin = lookup_table_bin.append(
        #     #
        #     {'var_name': testdata.columns[ii].replace('x_roybins', x), 'lower_bound': up, 'upper_bound': rge[1]},
        #     #     ignore_index=True)
        #     del rge, up
        self.variables[x].get_or_create_tni().column_mapping['tni_processed'] = x + '_TnI_processed'

        testdata = testdata.rename(columns={'x_royalias': (x + '_TnI_processed')})

        # TODO this depends on missing flag. Should remove
        # if testdata['x_missing_flag'].mean() > 0:
        #     # lookup_table=lookup_table({'var_name':
        #     (x+'_missing_flag'),'original_name':x,'type':'missing_flag'},ignore_index=True)
        #     lookup_table = lookup_table.append(
        #         {'var_name': (x + '_missing_flag'), 'original_name': x, 'type': 'missing_flag'}, ignore_index=True)
        #     testdata = testdata.rename(columns={'x_missing_flag': (x + '_missing_flag')})

        for ii in idx:
            _bin_col_name = testdata.columns[ii].replace('x_roybins', x)
            self.variables[x].get_or_create_tni(). \
                encoded_column_mapping[_bin_col_name] = _bin_col_name.split('_bi_')[1]

        testdata.columns = [c.replace('x_roybins', x) if 'x_roybins' in c else c for c in testdata.columns]

        if 'y_royindex' in testdata.columns:
            testdata = testdata.rename(columns={'y_royindex': f"{x}_dependent_var_index"})

        testdata = testdata[[c for c in testdata.columns if c not in drop1]]
        testdata = testdata.drop(columns=[x, self.dependent_variable])
        # self.training_data = self.training_data.join(testdata)
        for x in testdata:
            self.training_data[x] = testdata[x]

    def tni_transform(self, oot_data: pd.DataFrame = None) -> Optional[pd.DataFrame]:
        """
        Def: Transform data based on tni_smart transformations. If data is given, the given data will be
        transformed based on the training data from tni_smart and returned. If no data is given,
        the test data inside the model will be transformed instead

        :param oot_data: out of time data to transform, optional
        :return: TnI transformed out of time data if supplied, else None
        """
        if oot_data is not None:
            output_df = oot_data
        else:
            output_df = self.get_testing_data()

        for var in self.variables:
            if oot_data is not None and var.name not in [self.variables[v].name for v in self.var_selected]:
                continue

            # regenerate numerical variables
            if var.kind == 'continuous':
                _column_mapping = var.get_or_create_tni().column_mapping
                processed_tni_col = _column_mapping['tni_processed']
                # replace na with imputed value
                output_df[processed_tni_col] = output_df[var.name]
                output_df.fillna(value={
                    _column_mapping['tni_processed']: var.get_or_create_tni().value_mapping.get(None, None)},
                    inplace=True)

                # replace transformed variables
                # TODO look into creating common functions for transformations
                if 'tni_transformation_inv' in _column_mapping:
                    output_df[_column_mapping['tni_transformation_inv']] = 1 / \
                                                                                   output_df[processed_tni_col]

                if 'tni_transformation_sqrt' in _column_mapping:
                    output_df[_column_mapping['tni_transformation_sqrt']] = np.sqrt(
                        output_df[processed_tni_col])

                if 'tni_transformation_log' in _column_mapping:
                    output_df[_column_mapping['tni_transformation_log']] = \
                        np.log(np.abs(output_df[processed_tni_col]))

                if 'tni_transformation_exp' in _column_mapping:
                    output_df[_column_mapping['tni_transformation_exp']] = \
                        np.exp((output_df[processed_tni_col] - var.get_or_create_tni().lower_bound) /
                               (var.get_or_create_tni().upper_bound - var.get_or_create_tni().lower_bound))

                if 'tni_transformation_pw2' in _column_mapping:
                    output_df[_column_mapping['tni_transformation_pw2']] = \
                        output_df[processed_tni_col] ** 2

                # recreate binary variables
                if var.get_or_create_tni().cutter:
                    output_df[_column_mapping['tni_assigned_bin']] = \
                        pd.cut(output_df[processed_tni_col], **var.get_or_create_tni().cutter)
                    output_df = pd.get_dummies(output_df, columns=[_column_mapping['tni_assigned_bin']])
                    # fill in for columns not in test set
                    for col in var.get_or_create_tni().encoded_column_mapping.keys():
                        if col not in output_df.columns:
                            output_df[col] = 0

            if var.kind == 'categorical':
                # tni processed name defined here for easy code reading
                tni_processed_name = var.tni().column_mapping['tni_processed']
                # fill in missing values
                # safeguard to make sure train test split is done
                output_df = output_df.fillna({var.name: 'TnImissing'})

                # creates clone of source table to be flipped into tni assigned variables
                output_df[tni_processed_name] = output_df.loc[:, var.name]

                # maps values for tni_assign
                # value mapping contains dict for mapping categorical variable into tni assigned variable
                output_df.replace({tni_processed_name: var.tni().value_mapping}, inplace=True)

                # drop dummy columns that were removed during tni train
                output_df = pd.get_dummies(output_df, columns=[var.name])
                for col in var.tni()._categorical['drop_columns']:
                    if col in output_df.columns:
                        output_df.drop(columns=col, inplace=True)

                # TODO should probably move this into shared function
                # regenerate remaining dummy variables if not in testing subset.
                original_tni_dummies = list(var.get_or_create_tni().encoded_column_mapping.keys())
                for og in original_tni_dummies:
                    if og not in output_df.columns:
                        output_df[og] = 0

            if var.kind == 'binary':
                # replace null values
                output_df = output_df.fillna({var.name: 'TnImissing'})

                # recreate binary variables
                output_df = pd.get_dummies(output_df, columns=[var.name])

                # TODO performance enhancement possible to avoid iterating through entire column list
                # TODO should probably move this into shared function
                # regenerate remaining dummy variables if not in testing subset.
                original_tni_dummies = list(var.get_or_create_tni().encoded_column_mapping.keys())
                for og in original_tni_dummies:
                    if og not in output_df.columns:
                        output_df[og] = 0

        if oot_data is not None:
            return output_df
        else:
            self.testing_data = output_df

    def score_oot_data(self, oot_data):
        """
        Def: score the inputted dataframe, requires tni_smart to have been run

        :param oot_data: Out of time dataframe
        :return: tuple of pd.Dataframe, the prediction result bucketed into 10 deciles and incremental lift
        """
        return self.predict(self.tni_transform(oot_data=oot_data), self._model, self.dependent_variable)

    def variable_reduction(self,
                           excluded: list = (),
                           cluster_type='hclust',
                           wt_corr_dv=1,
                           wt_univ_reg=1,
                           wt_inf_val=1,
                           wt_clust=1,
                           wt_lasso=1,
                           selection=50) -> None:
        """
        Def: Variable reduction takes tni_smart object on the training dataset and incorporates 4 statistical methods
            (Correlation with Dependent Variable, Univariate Regression, Weight of Evidence for Logistic, Clustering)
            to enable variable reduction process. Independent variables having high Pearson correlation (0.8) with
            other independent variables are earmarked as potential multicollinearity candidates. Zero variance variables
            are removed. Pearson correlation is provided for correlation with dependent variable. For clustering,
            1 independent variable is chosen per cluster. The function produces an object containing all the
            evaluations. It will weight the difference methods based on the weights given by user and calculate the
            weighted avg score for variables and use the score to select the top variables, with the number of variables
            to be selected defined by the user.
        Outputs:
        Every variable is updated with one of the following status:
        , 1- Excluded by user
        , 2- Zero variance/near Zero variance
        , 3- High correlation with other independent variables
        , 5- Whether selected or not
        Method_values – Scores from various methods considered contains the following:
        , variable
        , Correlation with DV: corr_dv = Square of the Pearson correlation value and corr_direction: 1: +ive, 0: -ive
        , Univariate regression: univ_reg = Logistic – AUC, Linear – R_squared
        , Information value: inf_val (Only for logistic)
        , Cluster: cluster = cluster the variable is part of and cluster_correlation
        , Lasso: Lasso has been disabled and needs to be optimized.

        var_Selected gets updated
        :param excluded: list, variables to be excluded
        :param cluster_type: str, default hclust
        :param wt_corr_dv: weight given to correlation with dependent variable method
        :param wt_univ_reg: weight given to univariate regression method
        :param wt_inf_val: weight given to information value method, only for logistic
        :param wt_clust: weight given to clustering method
        :param wt_lasso: weight given to lasso method
        :param selection: number of var to be selected,
            var_Selected gets updated by the final selection of top variables
        """
        model_type = self.model_type
        # removes original variables
        current_variables = self.variables.get_active_variables_tni()
        current_variables = set(current_variables) - self.original_variables
        current_variables = set(list(current_variables) + list(self.variables.get_binary_variables()))

        for variable in excluded:
            if variable in current_variables:
                current_variables.remove(variable)
                self.variables[variable].set_variable_selection_status(1)
            else:
                warnings.warn(f'user excluded variable: {variable} is not found in data')

        # zero variance check
        for variable in current_variables.copy():
            if self.variables[variable].kind == 'binary':
                # use one of the binary columns to check for zero variance
                binary_col = list(self.variables[variable].tni().encoded_column_mapping.keys())[0]
                if np.var(self.get_training_data()[binary_col]) == 0:
                    current_variables.remove(variable)
                    self.variables[variable].set_variable_selection_status(2)
            elif np.var(self.get_training_data()[self.variables[variable].get_transformed_tni_variable_name()]) == 0:
                current_variables.remove(variable)
                self.variables[variable].set_variable_selection_status(2)
            # Check for near-zero variance
            if ((self.get_training_data()[variable].value_counts(normalize=True)) * 100 > 95).any():
                current_variables.remove(variable)
                self.variables[variable].set_variable_selection_status(2)

        if len(current_variables) == 0:
            raise Exception('No active variables')

        _evaluated_columns = []

        self._evaluated_columns = _evaluated_columns
        for var in current_variables:
            # pick up all binary variables
            if self.variables[var].tni().encoded_column_mapping:
                if self.variables[var].kind == 'binary':
                    # only take first listed value column for binary variables
                    _evaluated_columns.append(list(self.variables[var].tni().encoded_column_mapping.keys())[0])
                else:
                    _evaluated_columns += list(self.variables[var].tni().encoded_column_mapping.keys())

            # adds tni_transformed variables
            if self.variables[var].tni().column_mapping:
                _evaluated_columns.append(var)
                # _evaluated_columns.append(self.variables[var].tni().column_mapping['tni_processed'])

        # correlation with dv check
        _evaluated_columns = list(set(_evaluated_columns))
        data_w_dv = self.training_data[_evaluated_columns + [self.dependent_variable]]
        data_wout_dv = data_w_dv.drop([self.dependent_variable], axis=1)

        # Correlation with DV and among IV
        corr_result = data_w_dv.corr(method='pearson')
        corr_DV = corr_result.drop([self.dependent_variable])[self.dependent_variable].to_frame() ** 2

        # Correlation with DV + DF Prep
        corr_DV.columns = ['corr_dv']
        corr_DV['variable'] = corr_DV.index
        # TODO check if this change broke anything
        corr_DV['corr_direction'] = (corr_result.drop([self.dependent_variable])[self.dependent_variable].to_frame()
                                     >= 0).astype(int)

        self.corr_iv = corr_result.drop([self.dependent_variable]).drop([self.dependent_variable], axis=1)
        corr_IV1 = corr_result.drop([self.dependent_variable]).drop([self.dependent_variable], axis=1) ** 2
        np.fill_diagonal(corr_IV1.values, 0)

        corr_IV1 = (corr_IV1 > 0.64).sum().to_frame()
        corr_IV1.columns = ['High_IV_Corr']
        corr_IV1['variable'] = corr_IV1.index

        for var in corr_IV1[corr_IV1['High_IV_Corr'] > 0]['variable'].to_list():
            self.variables[var].set_variable_selection_status(3)

        # univariate regression
        univ_reg = pd.DataFrame()
        univ_reg['variable'] = data_wout_dv.columns.values
        if model_type == 'logistic':
            AUC = []
            for column in data_wout_dv.columns.values:
                try:
                    model = sm.Logit(data_w_dv[self.dependent_variable],
                                     data_wout_dv[column]).fit(disp=0)
                except Exception:
                    model = sm.Logit(data_w_dv[self.dependent_variable],
                                     data_wout_dv[column]).fit(disp=0, method='bfgs')
                fitted_values = model.predict()
                AUC.append(roc_auc_score(data_w_dv[self.dependent_variable], fitted_values))
            univ_reg['univ_reg'] = AUC
        else:
            R2 = []
            for column in data_wout_dv.columns.values:
                model = sm.OLS(data_w_dv[self.dependent_variable], data_wout_dv[column]).fit()
                R2.append(model.rsquared)
            univ_reg['univ_reg'] = R2

        # Information value

        if model_type == 'logistic':
            # Information Value Algorithm
            inf_val = st.iv_woe(data_w_dv, self.dependent_variable, bins=10, show_woe=False, force_bin=3)

        # Cluster
        data_wout_dv_scaled = MinMaxScaler().fit_transform(data_wout_dv)
        data_wout_dv_scaled = pd.DataFrame(data=data_wout_dv_scaled, columns=data_wout_dv.columns)
        data_wout_dv_scaled.index = data_wout_dv.index
        data_wout_dv_tr = data_wout_dv_scaled.transpose()
        if cluster_type == 'hclust':
            clust = pd.DataFrame()
            clust['variable'] = data_wout_dv.columns.values
            Z = ward(pdist(data_wout_dv_tr))
            clust['cluster'] = fcluster(Z, t=selection, criterion='maxclust')
            R = inconsistent(Z)
            clust_corr = R[:, 3]
            clust_corr = np.insert(clust_corr, 0, 0., axis=0)
            clust['cluster_correlation'] = clust_corr
        # lasso commented out temporarily for performance

        # Lasso
        # if model_type == 'linear':
        #     lasso_model = sm.OLS(data_w_dv[self.dependent_variable],
        #                          data_wout_dv_scaled).fit_regularized(alpha=0, L1_wt=1)
        # else:
        #     lasso_model = sm.Logit(data_w_dv[self.dependent_variable],
        #                            data_wout_dv_scaled).fit_regularized(alpha=0, L1_wt=1)
        #
        # lasso1 = lasso_model.params.to_frame().pow(2).reset_index()
        # lasso1.columns = ['variable', 'lasso']

        # Accumulating method scores
        method_values = pd.merge(corr_DV, univ_reg[['variable', 'univ_reg']], on='variable')
        if model_type == 'logistic':
            method_values = pd.merge(method_values, inf_val[['variable', 'inf_val']], on='variable')
        method_values = pd.merge(method_values, clust[['variable', 'cluster', 'cluster_correlation']], on='variable')

        # lasso commented out temporarily for performance
        # method_values = pd.merge(method_values, lasso1[['variable', 'lasso']], on='variable')

        # Selection function call
        var_scores = self.select_variable(method_values, model_type, selection, wt_corr_dv, wt_univ_reg, wt_inf_val,
                                          wt_clust,
                                          wt_lasso)
        self.var_scores = var_scores
        # TODO change selection to use Vyper variables class
        self.var_selected = var_scores.loc[(var_scores['selected'] > 0), 'variable'].to_list()
        for var in self.var_selected:
            self.variables[var].set_variable_selection_status(5)

        # Retro fitting dictionary breakdown of 'lookup' table for now.
        # TODO come back to recode accumulating method scores section after module 4 requirements are out.
        self.variable_reduction_method_values = method_values
        # self.variable_reduction_method_values.set_index('variable', inplace=True)
        # self.variable_reduction_method_values['variable'] = self.variable_reduction_method_values.index.copy()
        # for k, v in self.variable_reduction_method_values.to_dict('index').items():
        #     self.variables[k].set_variable_metrics(v)

    def rfe_selection(self,
                      use_all_variables: bool = False,
                      num_features: int = 20) -> None:
        """
        Def: Conduct rfe on the variables selected by variable_reduction. var_selected gets updated with the
        subset of the variables that survive rfe. It can also start over from all active variables.
        :param use_all_variables: use all active variables, not just previously selected
        :param num_features: The top number of features to select using RFE
        """
        model_type = self.model_type
        if use_all_variables:
            current_variables = self.variables.get_active_variables_tni()
            current_variables = set(current_variables) - self.original_variables
            current_variables = set(list(current_variables) + list(self.variables.get_binary_variables()))
        else:
            current_variables = self.var_selected

        data_train = self.get_training_data()
        dv = self.dependent_variable
        # getting the data without dependent variable
        data_wout_dv = data_train[current_variables]

        if model_type == 'linear':
            selector = RFE(linear_model.LinearRegression(), n_features_to_select=num_features)
            selector = selector.fit(data_wout_dv, data_train[dv])
        else:
            selector = RFE(linear_model.LogisticRegression(), n_features_to_select=num_features)
            selector = selector.fit(data_wout_dv, data_train[dv])

        selected_variables = pd.DataFrame(list(data_wout_dv.columns[selector.support_]), columns=['variables'])
        final_selection = selected_variables['variables']
        self.var_selected = list(final_selection)

    def lasso_selection(self,
                        use_all_variables: bool = False,
                        alpha: float = 0.001,
                        l1_wt: float = 1) -> None:
        """
        Def: Conduct lasso on the variables selected by variable_reduction. var_selected gets updated with the
        subset of the variables that survive lasso. It can also start over from all active variables.
        :param use_all_variables: use all active variables, not just previously selected
        :param alpha: The weight multiplying the l1 penalty term.
        :param l1_wt: The fraction of the penalty given to the L1 penalty term. Must be between 0 and 1 (inclusive).
        If 0, the fit is a ridge fit, if 1 it is a lasso fit.
        """
        model_type = self.model_type
        if use_all_variables:
            current_variables = self.variables.get_active_variables_tni()
            current_variables = set(current_variables) - self.original_variables
            current_variables = set(list(current_variables) + list(self.variables.get_binary_variables()))
        else:
            current_variables = self.var_selected

        # TODO clean up integration to directly refer to self.objects
        data_train = self.get_training_data()
        dv = self.dependent_variable
        # TODO change lasso from scaled to original data.
        data_wout_dv = data_train[current_variables]
        scaler = MinMaxScaler()
        data_wout_dv_scaled = scaler.fit_transform(data_wout_dv)
        data_wout_dv_scaled = pd.DataFrame(data=data_wout_dv_scaled, columns=data_wout_dv.columns)
        data_wout_dv_scaled.index = data_train.index
        if model_type == 'linear':
            lasso_model = sm.OLS(data_train[dv], data_wout_dv_scaled).fit_regularized(alpha=alpha, L1_wt=l1_wt)
        else:
            lasso_model = sm.Logit(data_train[dv], data_wout_dv_scaled).fit_regularized(alpha=alpha, L1_wt=l1_wt)
        selected_variables = abs(lasso_model.params).to_frame()
        selected_variables.columns = ['coefficient']
        selected_variables['variable'] = selected_variables.index
        final_selection = selected_variables.loc[(selected_variables['coefficient'] > 0), 'variable']
        self.var_selected = list(final_selection)

    @staticmethod
    def select_variable(method_values,
                        model_type,
                        selection=50,
                        wt_corr_dv=1,
                        wt_univ_reg=1,
                        wt_inf_val=1,
                        wt_clust=1,
                        wt_lasso=1):
        """
        Def: used by variable reduction to do the ranking across variables
        :param method_values:
        :param model_type: str, default linear, if logistic wt_inf_val will be used
        :param selection: number of var to be selected,
            var_Selected gets updated by the final selection of top variables
        :param wt_corr_dv: weight given to correlation with dv method
        :param wt_univ_reg: weight given to univariate regression method
        :param wt_inf_val: weight given to information value method, only for logistic
        :param wt_clust: weight given to clustering method
        :param wt_lasso: weight given to lasso method

        :return: pd.Dataframe, variable scores

        """
        # TODO does this really need to be a separate function?
        var_ranks = method_values.drop(columns=['cluster', 'cluster_correlation', 'corr_direction', 'variable']).rank(
            method='min', ascending=False)
        var_ranks['clust'] = method_values[['cluster', 'cluster_correlation']].groupby('cluster')[
            'cluster_correlation'].rank(method='first').rank(method='min')
        var_scores = (var_ranks <= selection).astype(int)
        var_scores['corr_dv'] = var_scores['corr_dv'] * wt_corr_dv
        if model_type == 'logistic':
            var_scores['inf_val'] = var_scores['inf_val'] * wt_inf_val
        var_scores['univ_reg'] = var_scores['univ_reg'] * wt_univ_reg
        var_scores['clust'] = var_scores['clust'] * wt_clust

        # lass commented out temporarily for performance
        # var_scores['lasso'] = var_scores['lasso'] * wt_lasso

        var_scores['selected'] = (var_scores.sum(axis=1).rank(method='first', ascending=False) <= selection).astype(int)
        var_scores['variable'] = method_values.variable
        return var_scores

    def extract_modelling_variables(self,
                                    var: str,
                                    ) -> List[str]:
        """
        variable object/name doesn't always reflect the variables you should model with.
        e.g for categorical/binary variables, binary vars should be used.
        for numerical variable, tni_transformed should be used
        :param var: name of variable to go look in self.variables for

        :return: list of column names that should be modeled on for that particular variable
        """

        variable_object = self.variables[var]

        if variable_object.kind in {'binary', 'categorical'}:
            # take transformed binary variables
            # converts dict keys to list object
            return list(variable_object.get_or_create_tni().encoded_column_mapping.keys())

        elif variable_object.kind == 'continuous':
            # return tni_transformed for continous variables
            return [variable_object.get_or_create_tni().column_mapping['tni_processed']]

    def fit(self,
            pval_cutoff=0.05,
            vif_cutoff=4,
            report_all_vars=True):
        """
        Def : Conduct model fit and test validations on the variables in var_selected. Model is evaluated based on
        correlation, p_value and vif. Prerequisite: test dataset needs to be created before running fit. It reports
        model descriptors, model metrics, train and test lift profiles. It also outputs train and test datasets with
        predicted scores and deciles.
        :param pval_cutoff: float, the cutoff p-value used to evaluate the variable
        :param vif_cutoff: int, the cutoff value of variance inflation factor (VIF) used to evaluate the variable
        :param report_all_vars: bool, when output the dataset whether output all variables in the model or not
        :return: dict, result of the model trained. Output as:
        , train_out - predicted result of training data
        , test_out - predicted result of test data
        , model_descriptor - variable, estimate, standard_error, t_z_val (Linear – t value; Logistic z value),
            p_val, vif, standardized_coeff, importance, corr_dv_direction: Positive or negative correlation as
            determined by variable reduction, corr_model: Positive or negative correlation as per model fit,
            evaluation: 1 if variable is suitable; 0 if not suitable
        , model_metric - Linear: adj_r_sq, aic; Logistic: train_auc, test_auc
        , train_lift - incremental lift bucketed into 10 deciles of training data
        , test_lift - incremental lift bucketed into 10 deciles of testing data
        """
        if not self.var_selected:
            warnings.warn('variable reduction not yet run')

            # Train Test Split check
        print('Train mean response: ', mean(self.get_training_data()[self.dependent_variable]))
        print('Test mean response: ', mean(self.get_testing_data()[self.dependent_variable]))
        print('Deviation in response: ' + '{:.2%}'.format((mean(self.get_testing_data()[self.dependent_variable]) /
                                                           mean(self.get_training_data()[self.dependent_variable])) -
                                                          1))
        # selected_variables =
        # [name for name, object in self.variables.items() if object.variable_selection_status == 5]
        # Data initialization

        model_vars = self.var_selected

        train_wout_dv = self.get_training_data()[model_vars]
        # test_wout_dv = self.get_testing_data()[model_vars]
        train_constant_add = sm.add_constant(train_wout_dv)
        # test_constant_add = sm.add_constant(test_wout_dv)

        # Model Build
        if self.model_type == 'logistic':
            model = sm.Logit(self.get_training_data()[self.dependent_variable], train_constant_add).fit(disp=0)
        else:
            model = sm.OLS(self.get_training_data()[self.dependent_variable], train_constant_add).fit()

        # Standardized Coefficients
        if self.model_type == 'logistic':
            beta_coef = sm.Logit(self.get_training_data()[self.dependent_variable],
                                 sm.add_constant(zscore(train_wout_dv))).fit(disp=0)
        else:
            beta_coef = sm.OLS(zscore(self.get_training_data()[self.dependent_variable]),
                               sm.add_constant(zscore(train_wout_dv))).fit()

        # Model Descriptor Summary
        model_descriptor = pd.DataFrame()
        model_descriptor['variable'] = model.params.to_frame().index
        model_descriptor.iloc[0, 0] = 'intercept'
        model_descriptor['estimate'] = model.params.to_list()
        model_descriptor['standard_error'] = model.bse.to_list()
        model_descriptor['t_z_val'] = model.tvalues.to_list()
        model_descriptor['p_val'] = model.pvalues.to_list()
        model_descriptor['vif'] = [variance_inflation_factor(train_constant_add.values, i) for i in
                                   range(train_constant_add.shape[1])]
        model_descriptor.iloc[0, 5] = 0.0
        if self.model_type == 'logistic':
            model_descriptor['standardized_coeff'] = beta_coef.params.to_list()
        else:
            model_descriptor['standardized_coeff'] = beta_coef.params
        model_descriptor.iloc[0, 6] = 0.0
        model_descriptor[
            'importance'] = model_descriptor.standardized_coeff.abs() / model_descriptor.standardized_coeff.abs().sum()

        model_descriptor = pd.merge(model_descriptor,
                                    self.variable_reduction_method_values[['variable', 'corr_direction']], how='left',
                                    on='variable')
        # def cust_mapper(row):
        #     # TODO not the best way to access hidden dict
        #     if self.variables.bool_get(row.variable):
        #         metrics_dt = self.variables[row.variable].metrics
        #         return {**row.to_dict(), **metrics_dt}
        #     else:
        #         return {**row.to_dict(), **Variable('empty').metrics}
        # model_descriptor = model_descriptor.apply(cust_mapper, axis=1, result_type='expand')

        model_descriptor = model_descriptor.rename(columns={'corr_direction': 'corr_dv_direction'})
        # TODO check if this broke anything, formatting fix
        model_descriptor['corr_model'] = (model_descriptor['estimate'] > 0).astype(int)
        model_descriptor['evaluation'] = ((model_descriptor['corr_model'] == model_descriptor['corr_dv_direction']) & (
                model_descriptor['p_val'] <= pval_cutoff) & (model_descriptor['vif'] <= vif_cutoff)).astype(int)

        # Lift evaluation
        (train_out, train_lift) = self.predict(data=self.get_training_data()[list(model_vars) +
                                                                             [self.dependent_variable]],
                                               model_obj=model,
                                               dv=self.dependent_variable,
                                               report_all_vars=report_all_vars)

        (test_out, test_lift) = self.predict(data=self.get_testing_data(),
                                             model_obj=model,
                                             dv=self.dependent_variable, report_all_vars=report_all_vars)

        if self.model_type == 'logistic':
            train_lift = train_lift.rename(columns={'value': 'response'})
            test_lift = test_lift.rename(columns={'value': 'response'})

        # Model Metrics
        if self.model_type == 'logistic':
            model_metric = {"train_auc": roc_auc_score(self.get_training_data()[self.dependent_variable],
                                                       train_out['y_pred']),
                            "test_auc": roc_auc_score(self.get_testing_data()[self.dependent_variable],
                                                      test_out['y_pred'])}
        else:
            model_metric = {"adj_r_sq": model.rsquared_adj, "aic": model.aic}

        self._model = model
        self.train_out = train_out
        self.test_out = test_out
        self.model_descriptor = model_descriptor
        self.model_metric = model_metric
        self.train_lift = train_lift
        self.test_lift = test_lift

        return {'train_out': train_out,
                'test_out': test_out,
                'model_descriptor': model_descriptor,
                'model_metric': model_metric,
                'train_lift': train_lift,
                'test_lift': test_lift}

    def get_adj_rsq(self, y_true, y_pred):
        """
        Get the adjusted R squared for the given model and predictions
        :param y_true: The true y values
        :param y_pred: The predicted y values
        :return: The adjusted R squared
        """
        r2 = r2_score(y_true, y_pred)
        # params contains constant
        return 1 - (1-r2)*(len(y_true)-1)/(len(y_true)-(len(self._model.params)-1)-1)

    @staticmethod
    def predict(data,
                model_obj,
                dv='y',
                report_all_vars=True):
        """
        Def: make predictions on train/test/OOT data and score, Tni transformation needs to be done

        :param data: result of tni transformed data, same variables as the Model trained
        :param model_obj: model object generated in the fit function
        :param dv: str, column name of dependent variable
        :param report_all_vars: bool, when output the dataset whether output all variables in the model or not
        :return: tuple of pd.Dataframe, the prediction result bucketed into 10 deciles and incremental lift
        """
        # TODO rework this function not keen on static method in main model class
        if report_all_vars:
            data_out = data
        else:
            data_out = data[model_obj.params.index[1:]]
            data_out[dv] = data[dv]

        data1 = sm.add_constant(data[model_obj.params.index[1:]])

        # error out if test data has categorical values that did not show up in train
        cat_error = ~data1.applymap(np.isreal).all(axis=0)
        if len(cat_error[cat_error].index.values) > 0:
            bad_cols = ', '.join(cat_error[cat_error].index.values)
            raise ValueError('Categorical values show up in test but not train for columns: {}'.format(bad_cols))

        data_out['y_pred'] = model_obj.predict(data1)
        data_out['decile'] = 10 - (pd.qcut(data_out['y_pred'], 10, labels=False, duplicates='drop'))

        lift = None
        if dv in data.columns:
            lift = data_out.groupby(['decile']).agg(number_obs=('y_pred', 'count'), value=(dv, sum),
                                                    incremental_lift=(dv, 'mean'))
            lift['decile'] = lift.index
            lift['incremental_lift'] = lift['incremental_lift'] / (lift['incremental_lift'].mean())
        return data_out, lift

    # TODO either recreate this worksheet or remove entirely
    # tni_lookup_obj is no longer generated
    def create_tni_model_var_descr_ws(self, tni_lookup_obj) -> None:
        """
        Def: function to create the tni model variable description worksheet

        :param tni_lookup_obj: the output dataframe from Module 2
        """
        tni_model_var_descr_ws = self.wb.create_sheet("TnI Model Variable Descr.", 0)

        # make new sheet the active sheet we are working on
        tni_model_var_descr_ws = self.wb.active

        # remove gridlines from the sheet
        tni_model_var_descr_ws.sheet_view.showGridLines = False

        # headers of table -- content & style
        tni_model_var_descr_ws['B3'] = 'Model Variable'
        tni_model_var_descr_ws['C3'] = 'Original Variable'
        tni_model_var_descr_ws['D3'] = 'Transformation Type'
        tni_model_var_descr_ws['E3'] = 'Transformation Code'

        thin = Side(border_style="thin", color="000000")

        tni_model_var_descr_ws.column_dimensions['B'].width = 40
        tni_model_var_descr_ws.column_dimensions['C'].width = 30
        tni_model_var_descr_ws.column_dimensions['D'].width = 20
        tni_model_var_descr_ws.column_dimensions['E'].width = 60

        for i in 'BCDE':
            tni_model_var_descr_ws[i + '3'].font = Font(bold=True)
            tni_model_var_descr_ws[i + '3'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            current_cell = tni_model_var_descr_ws[i + '3']
            current_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")

        # Variable names & Transformation type
        for i in range(0, len(self.model_descriptor) - 1):
            for j in range(0, len(tni_lookup_obj['lookup']['lookup_table'])):
                if tni_lookup_obj['lookup']['lookup_table']['var_name'][j] == self.model_descriptor['variable'][i + 1]:
                    tni_model_var_descr_ws['B' + str(4 + i)] = tni_lookup_obj['lookup']['lookup_table']['var_name'][j]
                    tni_model_var_descr_ws['C' + str(4 + i)] = \
                        tni_lookup_obj['lookup']['lookup_table']['original_name'][j]
                    tni_model_var_descr_ws['D' + str(4 + i)] = tni_lookup_obj['lookup']['lookup_table']['type'][j]
                    # print(i, tni_output_df['lookup']['lookup_table']['var_name'][j], 'added')
            for k in 'BCDE':
                tni_model_var_descr_ws[k + str(4 + i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #
    def create_model_corr_matrix_ws(self) -> None:
        """
        Def: function to create the model correlation matrix worksheet
        """
        model_corr_matrix_ws = self.wb.create_sheet("Model Correlation", 0)

        # make new sheet the active sheet we are working on
        model_corr_matrix_ws = self.wb.active

        model_eval = self.model_descriptor

        # remove gridlines from the sheet
        model_corr_matrix_ws.sheet_view.showGridLines = False

        # set 1st and 3rd row dimensions for the sheet
        model_corr_matrix_ws.row_dimensions[3].height = 25
        model_corr_matrix_ws.row_dimensions[1].height = 30

        data = self.corr_iv

        cols_to_keep = list(model_eval['variable'])
        cols_to_keep.remove('intercept')
        data = data[cols_to_keep][data.index.isin(cols_to_keep)]
        data = data[data.index]

        # set border line thickenss
        # thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        letter_col_name_map = {}
        for i, col in enumerate(list(data.columns)):
            letter_col_name_map[openpyxl.utils.cell._get_column_letter(i + 2)] = col

        # set column widths
        for i in list(letter_col_name_map.keys()):
            model_corr_matrix_ws.column_dimensions[i].width = 30

        # create title text
        model_corr_matrix_ws['A3'] = "Pearson Correlation"
        # model_corr_matrix_ws['A7'] = "Client and Project ID:"
        # model_corr_matrix_ws['A8'] = "Model Name:"
        model_corr_matrix_ws['A9'] = "Var Name:"
        model_corr_matrix_ws['A9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        model_corr_matrix_ws['A9'].font = Font(size=11, underline="single", bold=True)
        model_corr_matrix_ws['A3'].font = Font(size=18, underline="single", bold=True)
        model_corr_matrix_ws['A3'].fill = PatternFill("solid", fgColor="A9C4FE")

        # cells_text = ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3']

        # iterrate over data frame created above
        for index, (col, row) in enumerate(data.iterrows()):
            i = index + 10

            for col_number in range(len(data)):
                col_letter = openpyxl.utils.cell._get_column_letter(col_number + 2)

                model_corr_matrix_ws[col_letter + str(i)] = row[letter_col_name_map[col_letter]]
                model_corr_matrix_ws[col_letter + str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col_letter, col_name in letter_col_name_map.items():
            model_corr_matrix_ws[col_letter + '9'] = col_name
            model_corr_matrix_ws[col_letter + '9'].fill = PatternFill("solid", fgColor="A9C4FE")
            model_corr_matrix_ws[col_letter + '9'].font = Font(size=11, bold=True)
            model_corr_matrix_ws[col_letter + '9'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
        model_corr_matrix_ws.column_dimensions['A'].width = 30

        for i, row_name in enumerate(list(data.columns)):
            index = i + 10
            model_corr_matrix_ws['A' + str(index)] = row_name
            model_corr_matrix_ws['A' + str(index)].fill = PatternFill("solid", fgColor="A9C4FE")
            model_corr_matrix_ws['A' + str(index)].font = Font(size=11, bold=True)
            model_corr_matrix_ws['A' + str(index)].border = Border(top=thin, left=thin, right=thin, bottom=thin)

    #
    def create_model_perf_resp_ws(self, oot_out: pd.DataFrame = None) -> None:
        """
        Def: function to create the model performance tables for response models

        :param oot_out: scored out of time dataset output from model_out_of_time function
        """
        model_perf_resp_ws = self.wb.create_sheet("Model Perf. (Resp. Model)", 0)

        # make new sheet the active sheet we are working on
        model_perf_resp_ws = self.wb.active

        # remove gridlines from the sheet
        model_perf_resp_ws.sheet_view.showGridLines = False

        # auc data
        auc = self.model_metric

        # training data model performance
        train_perf = self.train_lift

        # testing data model performance
        test_perf = self.test_lift

        # set row dimensions for the sheet
        model_perf_resp_ws.row_dimensions[3].height = 25
        model_perf_resp_ws.row_dimensions[4].height = 25
        model_perf_resp_ws.row_dimensions[5].height = 25
        model_perf_resp_ws.row_dimensions[19].height = 25
        model_perf_resp_ws.row_dimensions[20].height = 25
        model_perf_resp_ws.row_dimensions[21].height = 25

        # set border line thickenss
        # thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # set column widths
        cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O']
        for i in cols:
            model_perf_resp_ws.column_dimensions[i].width = 20

        # create title text
        model_perf_resp_ws['B1'] = "Model Performance"
        model_perf_resp_ws['B1'].font = Font(size=18, underline="single", bold=True)
        model_perf_resp_ws.merge_cells('B1:E1')

        # training data model performance table setup
        model_perf_resp_ws['B3'] = 'Training'
        model_perf_resp_ws['B3'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('B3:C3')

        model_perf_resp_ws['B4'] = 'Decile'
        # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
        model_perf_resp_ws.merge_cells('B4:B5')

        model_perf_resp_ws['C4'] = 'Incremental'
        model_perf_resp_ws['C4'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('C4:G4')

        model_perf_resp_ws['H4'] = 'Cumulative'
        model_perf_resp_ws['H4'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('H4:L4')

        model_perf_resp_ws['C5'] = "Total Records"
        model_perf_resp_ws['D5'] = "Responders"
        model_perf_resp_ws['E5'] = "Non-Responders"
        model_perf_resp_ws['F5'] = "Resp. Rate"
        model_perf_resp_ws['G5'] = "Inc. Lift"
        model_perf_resp_ws['H5'] = "Cum. Total Records"
        model_perf_resp_ws['I5'] = "Cum. Responders"
        model_perf_resp_ws['J5'] = "Cum. Non-Responders"
        model_perf_resp_ws['K5'] = "Cum. Resp. Rate"
        model_perf_resp_ws['L5'] = "Cum. Lift"

        model_perf_resp_ws['N4'] = "AUC ="
        model_perf_resp_ws['O4'] = auc['train_auc']
        #########################################################################

        # testing data model performance table set up
        model_perf_resp_ws['B19'] = 'Testing'
        model_perf_resp_ws['B19'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('B19:C19')

        model_perf_resp_ws['B20'] = 'Decile'
        # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
        model_perf_resp_ws.merge_cells('B20:B21')

        model_perf_resp_ws['C20'] = 'Incremental'
        model_perf_resp_ws['C20'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('C20:G20')

        model_perf_resp_ws['H20'] = 'Cumulative'
        model_perf_resp_ws['H20'].font = Font(size=18, bold=True)
        model_perf_resp_ws.merge_cells('H20:L20')

        model_perf_resp_ws['C21'] = "Total Records"
        model_perf_resp_ws['D21'] = "Responders"
        model_perf_resp_ws['E21'] = "Non-Responders"
        model_perf_resp_ws['F21'] = "Resp. Rate"
        model_perf_resp_ws['G21'] = "Inc. Lift"
        model_perf_resp_ws['H21'] = "Cum. Total Records"
        model_perf_resp_ws['I21'] = "Cum. Responders"
        model_perf_resp_ws['J21'] = "Cum. Non-Responders"
        model_perf_resp_ws['K21'] = "Cum. Resp. Rate"
        model_perf_resp_ws['L21'] = "Cum. Lift"

        model_perf_resp_ws['N20'] = "AUC ="
        model_perf_resp_ws['O20'] = auc['test_auc']

        #########################################################################################
        # format text cells

        cells_text = ['B3', 'C3', 'B4', 'B5', 'C4', 'D4', 'E4',
                      'F4', 'G4', 'C5', 'D5', 'E5', 'F5',
                      'G5', 'H4', 'I4', 'J4', 'K4', 'L4',
                      'H5', 'I5', 'J5', 'K5', 'L5',
                      'N4', 'O4']
        for i in cells_text:
            model_perf_resp_ws[i].font = Font(bold=True)
            current_cell = model_perf_resp_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        input_rows = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
        input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for i in input_cols:
            z = 1
            for j in input_rows:
                x = str(j)
                x = i + x
                # print(x)
                current_cell = model_perf_resp_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # decile
                if i == 'B':
                    model_perf_resp_ws[x] = train_perf['decile'][z]
                    z = z + 1
                # total records
                elif i == 'C':
                    model_perf_resp_ws[x] = train_perf['number_obs'][z]
                    z = z + 1
                # responders
                elif i == 'D':
                    model_perf_resp_ws[x] = train_perf['response'][z]
                    z = z + 1
                # inc. lift
                elif i == 'G':
                    model_perf_resp_ws[x] = train_perf['incremental_lift'][z] * 100
                    z = z + 1
                # non-responders
                elif i == 'E':
                    c = 'C' + str(j)
                    d = 'D' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[c].value - model_perf_resp_ws[d].value
                # resp. rate
                elif i == 'F':
                    c = 'C' + str(j)
                    d = 'D' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[d].value / model_perf_resp_ws[c].value
                    model_perf_resp_ws[x].number_format = '0.00%'
                # cum. total responders
                elif i == 'H':
                    if j == 6:
                        model_perf_resp_ws[x] = model_perf_resp_ws['C6'].value
                    else:
                        c = 'C' + str(j)
                        h = 'H' + str(j - 1)
                        model_perf_resp_ws[x] = model_perf_resp_ws[h].value + model_perf_resp_ws[c].value
                # cum. responders
                elif i == 'I':
                    if j == 6:
                        model_perf_resp_ws[x] = model_perf_resp_ws['D6'].value
                    else:
                        d = 'D' + str(j)
                        ii = 'I' + str(j - 1)
                        model_perf_resp_ws[x] = model_perf_resp_ws[ii].value + model_perf_resp_ws[d].value
                # cum. non-responders
                elif i == 'J':
                    h = 'H' + str(j)
                    ii = 'I' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[h].value - model_perf_resp_ws[ii].value
                    # cum. resp. rate
                elif i == 'K':
                    ii = 'I' + str(j)
                    h = 'H' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[ii].value / model_perf_resp_ws[h].value
                    model_perf_resp_ws[x].number_format = '0.00%'
                    # cum. lift
                elif i == 'L':
                    k = 'K' + str(j)
                    model_perf_resp_ws[x] = (model_perf_resp_ws[k].value / model_perf_resp_ws['K15'].value) * 100
                else:
                    break

        # create total column values for incremental data
        tot_col = ['B', 'C', 'D', 'E', 'F', 'G']
        row = 16
        for i in tot_col:
            x = i + str(row)
            current_cell = model_perf_resp_ws[x]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i == 'B':
                model_perf_resp_ws[x] = 'Total'
            elif i == 'C':
                model_perf_resp_ws[x] = train_perf['number_obs'].sum()
            elif i == 'D':
                model_perf_resp_ws[x] = train_perf['response'].sum()
            elif i == 'E':
                model_perf_resp_ws[x] = model_perf_resp_ws['C16'].value - model_perf_resp_ws['D16'].value
            elif i == 'F':
                model_perf_resp_ws[x] = model_perf_resp_ws['D16'].value / model_perf_resp_ws['C16'].value
                model_perf_resp_ws[x].number_format = '0.00%'
            elif i == 'G':
                model_perf_resp_ws[x] = (model_perf_resp_ws['F16'].value / model_perf_resp_ws['F16'].value) * 100
            else:
                break

        #########################################################################
        # now need to do the same thing for the testing decile table
        ###########################################################################

        cells_text = ['B19', 'C19', 'B20', 'B21', 'C20', 'D20', 'E20',
                      'F20', 'G20', 'C21', 'D21', 'E21', 'F21',
                      'G21', 'H20', 'I20', 'J20', 'K20', 'L20',
                      'H21', 'I21', 'J21', 'K21', 'L21',
                      'N20', 'O20']
        for i in cells_text:
            model_perf_resp_ws[i].font = Font(bold=True)
            current_cell = model_perf_resp_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        input_rows = [22, 23, 24, 25, 26, 27, 28, 29, 30, 31]
        input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        for i in input_cols:
            z = 1
            for j in input_rows:
                x = str(j)
                x = i + x
                current_cell = model_perf_resp_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # decile
                if i == 'B':
                    model_perf_resp_ws[x] = test_perf['decile'][z]
                    z = z + 1
                # total records
                elif i == 'C':
                    model_perf_resp_ws[x] = test_perf['number_obs'][z]
                    z = z + 1
                # responders
                elif i == 'D':
                    model_perf_resp_ws[x] = test_perf['response'][z]
                    z = z + 1
                # inc. lift
                elif i == 'G':
                    model_perf_resp_ws[x] = test_perf['incremental_lift'][z] * 100
                    z = z + 1
                # non-responders
                elif i == 'E':
                    c = 'C' + str(j)
                    d = 'D' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[c].value - model_perf_resp_ws[d].value
                # resp. rate
                elif i == 'F':
                    c = 'C' + str(j)
                    d = 'D' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[d].value / model_perf_resp_ws[c].value
                    model_perf_resp_ws[x].number_format = '0.00%'
                # cum. total responders
                elif i == 'H':
                    if j == 22:
                        model_perf_resp_ws[x] = model_perf_resp_ws['C22'].value
                    else:
                        c = 'C' + str(j)
                        h = 'H' + str(j - 1)
                        model_perf_resp_ws[x] = model_perf_resp_ws[h].value + model_perf_resp_ws[c].value
                # cum. responders
                elif i == 'I':
                    if j == 22:
                        model_perf_resp_ws[x] = model_perf_resp_ws['D22'].value
                    else:
                        d = 'D' + str(j)
                        ii = 'I' + str(j - 1)
                        model_perf_resp_ws[x] = model_perf_resp_ws[ii].value + model_perf_resp_ws[d].value
                # cum. non-responders
                elif i == 'J':
                    h = 'H' + str(j)
                    ii = 'I' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[h].value - model_perf_resp_ws[ii].value
                    # cum. resp. rate
                elif i == 'K':
                    ii = 'I' + str(j)
                    h = 'H' + str(j)
                    model_perf_resp_ws[x] = model_perf_resp_ws[ii].value / model_perf_resp_ws[h].value
                    model_perf_resp_ws[x].number_format = '0.00%'
                    # cum. lift
                elif i == 'L':
                    k = 'K' + str(j)
                    model_perf_resp_ws[x] = (model_perf_resp_ws[k].value / model_perf_resp_ws['K31'].value) * 100
                else:
                    break
        # create total column values for incremental data
        tot_col = ['B', 'C', 'D', 'E', 'F', 'G']
        row = 32
        for i in tot_col:
            x = i + str(row)
            current_cell = model_perf_resp_ws[x]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i == 'B':
                model_perf_resp_ws[x] = 'Total'
            elif i == 'C':
                model_perf_resp_ws[x] = test_perf['number_obs'].sum()
            elif i == 'D':
                model_perf_resp_ws[x] = test_perf['response'].sum()
            elif i == 'E':
                model_perf_resp_ws[x] = model_perf_resp_ws['C32'].value - model_perf_resp_ws['D32'].value
            elif i == 'F':
                model_perf_resp_ws[x] = model_perf_resp_ws['D32'].value / model_perf_resp_ws['C32'].value
                model_perf_resp_ws[x].number_format = '0.00%'
            elif i == 'G':
                model_perf_resp_ws[x] = (model_perf_resp_ws['F32'].value / model_perf_resp_ws['F32'].value) * 100
            else:
                break

        #########################################################################
        # now need to do the same thing for the out of time decile table
        ###########################################################################
        if oot_out is not None:

            model_perf_resp_ws.row_dimensions[35].height = 25
            model_perf_resp_ws.row_dimensions[36].height = 25
            model_perf_resp_ws.row_dimensions[37].height = 25

            # oot data model performance table setup
            model_perf_resp_ws['B35'] = 'Out of Time'
            model_perf_resp_ws['B35'].font = Font(size=18, bold=True)
            model_perf_resp_ws.merge_cells('B35:C35')

            model_perf_resp_ws['B36'] = 'Decile'
            # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
            model_perf_resp_ws.merge_cells('B36:B37')

            model_perf_resp_ws['C36'] = 'Incremental'
            model_perf_resp_ws['C36'].font = Font(size=18, bold=True)
            model_perf_resp_ws.merge_cells('C36:G36')

            model_perf_resp_ws['H36'] = 'Cumulative'
            model_perf_resp_ws['H36'].font = Font(size=18, bold=True)
            model_perf_resp_ws.merge_cells('H36:L36')

            model_perf_resp_ws['C37'] = "Total Records"
            model_perf_resp_ws['D37'] = "Responders"
            model_perf_resp_ws['E37'] = "Non-Responders"
            model_perf_resp_ws['F37'] = "Resp. Rate"
            model_perf_resp_ws['G37'] = "Inc. Lift"
            model_perf_resp_ws['H37'] = "Cum. Total Records"
            model_perf_resp_ws['I37'] = "Cum. Responders"
            model_perf_resp_ws['J37'] = "Cum. Non-Responders"
            model_perf_resp_ws['K37'] = "Cum. Resp. Rate"
            model_perf_resp_ws['L37'] = "Cum. Lift"

            model_perf_resp_ws['N36'] = "AUC ="
            model_perf_resp_ws['O36'] = roc_auc_score(oot_out[0][self.dependent_variable], oot_out[0]['y_pred'])

            oot_lift = oot_out[1]

            cells_text = ['B35', 'C35', 'B36', 'B37', 'C36', 'D36', 'E36',
                          'F36', 'G36', 'C37', 'D37', 'E37', 'F37',
                          'G37', 'H36', 'I36', 'J36', 'K36', 'L36',
                          'H37', 'I37', 'J37', 'K37', 'L37',
                          'N36', 'O36']
            for i in cells_text:
                model_perf_resp_ws[i].font = Font(bold=True)
                current_cell = model_perf_resp_ws[i]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            input_rows = [38, 39, 40, 41, 42, 43, 44, 45, 46, 47]
            input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            for i in input_cols:
                z = 0
                for j in input_rows:
                    x = str(j)
                    x = i + x
                    current_cell = model_perf_resp_ws[x]
                    current_cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    # decile
                    if i == 'B':
                        model_perf_resp_ws[x] = oot_lift['decile'].iloc[z]
                        # print(oot_lift['decile'].iloc[z])
                        z = z + 1
                    # total records
                    elif i == 'C':
                        model_perf_resp_ws[x] = oot_lift['number_obs'].iloc[z]
                        z = z + 1
                    # responders
                    elif i == 'D':
                        model_perf_resp_ws[x] = oot_lift['value'].iloc[z]
                        z = z + 1
                    # inc. lift
                    elif i == 'G':
                        model_perf_resp_ws[x] = oot_lift['incremental_lift'].iloc[z] * 100
                        z = z + 1
                    # non-responders
                    elif i == 'E':
                        c = 'C' + str(j)
                        d = 'D' + str(j)
                        model_perf_resp_ws[x] = model_perf_resp_ws[c].value - model_perf_resp_ws[d].value
                    # resp. rate
                    elif i == 'F':
                        c = 'C' + str(j)
                        d = 'D' + str(j)
                        model_perf_resp_ws[x] = model_perf_resp_ws[d].value / model_perf_resp_ws[c].value
                        model_perf_resp_ws[x].number_format = '0.00%'
                    # cum. total responders
                    elif i == 'H':
                        if j == 38:
                            model_perf_resp_ws[x] = model_perf_resp_ws['C38'].value
                        else:
                            c = 'C' + str(j)
                            h = 'H' + str(j - 1)
                            model_perf_resp_ws[x] = model_perf_resp_ws[h].value + model_perf_resp_ws[c].value
                    # cum. responders
                    elif i == 'I':
                        if j == 38:
                            model_perf_resp_ws[x] = model_perf_resp_ws['D38'].value
                        else:
                            d = 'D' + str(j)
                            ii = 'I' + str(j - 1)
                            model_perf_resp_ws[x] = model_perf_resp_ws[ii].value + model_perf_resp_ws[d].value
                    # cum. non-responders
                    elif i == 'J':
                        h = 'H' + str(j)
                        ii = 'I' + str(j)
                        model_perf_resp_ws[x] = model_perf_resp_ws[h].value - model_perf_resp_ws[ii].value
                        # cum. resp. rate
                    elif i == 'K':
                        ii = 'I' + str(j)
                        h = 'H' + str(j)
                        model_perf_resp_ws[x] = model_perf_resp_ws[ii].value / model_perf_resp_ws[h].value
                        model_perf_resp_ws[x].number_format = '0.00%'
                        # cum. lift
                    elif i == 'L':
                        k = 'K' + str(j)
                        model_perf_resp_ws[x] = (model_perf_resp_ws[k].value / model_perf_resp_ws['K47'].value) * 100
                    else:
                        break
            # create total column values for incremental data
            tot_col = ['B', 'C', 'D', 'E', 'F', 'G']
            row = 48
            for i in tot_col:
                x = i + str(row)
                current_cell = model_perf_resp_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                if i == 'B':
                    model_perf_resp_ws[x] = 'Total'
                elif i == 'C':
                    model_perf_resp_ws[x] = oot_lift['number_obs'].sum()
                elif i == 'D':
                    model_perf_resp_ws[x] = oot_lift['value'].sum()
                elif i == 'E':
                    model_perf_resp_ws[x] = model_perf_resp_ws['C48'].value - model_perf_resp_ws['D48'].value
                elif i == 'F':
                    model_perf_resp_ws[x] = model_perf_resp_ws['D48'].value / model_perf_resp_ws['C48'].value
                    model_perf_resp_ws[x].number_format = '0.00%'
                elif i == 'G':
                    model_perf_resp_ws[x] = (model_perf_resp_ws['F48'].value / model_perf_resp_ws['F48'].value) * 100
                else:
                    break

        for i in range(6, 17):
            model_perf_resp_ws['C' + str(i)].number_format = '#,###'
            model_perf_resp_ws['D' + str(i)].number_format = '#,###'
            model_perf_resp_ws['E' + str(i)].number_format = '#,###'
            model_perf_resp_ws['F' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['G' + str(i)].number_format = '#,###'
            model_perf_resp_ws['H' + str(i)].number_format = '#,###'
            model_perf_resp_ws['I' + str(i)].number_format = '#,###'
            model_perf_resp_ws['J' + str(i)].number_format = '#,###'
            model_perf_resp_ws['K' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['L' + str(i)].number_format = '#,###'

        for i in range(22, 33):
            model_perf_resp_ws['C' + str(i)].number_format = '#,###'
            model_perf_resp_ws['D' + str(i)].number_format = '#,###'
            model_perf_resp_ws['E' + str(i)].number_format = '#,###'
            model_perf_resp_ws['F' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['G' + str(i)].number_format = '#,###'
            model_perf_resp_ws['H' + str(i)].number_format = '#,###'
            model_perf_resp_ws['I' + str(i)].number_format = '#,###'
            model_perf_resp_ws['J' + str(i)].number_format = '#,###'
            model_perf_resp_ws['K' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['L' + str(i)].number_format = '#,###'

        for i in range(38, 49):
            model_perf_resp_ws['C' + str(i)].number_format = '#,###'
            model_perf_resp_ws['D' + str(i)].number_format = '#,###'
            model_perf_resp_ws['E' + str(i)].number_format = '#,###'
            model_perf_resp_ws['F' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['G' + str(i)].number_format = '#,###'
            model_perf_resp_ws['H' + str(i)].number_format = '#,###'
            model_perf_resp_ws['I' + str(i)].number_format = '#,###'
            model_perf_resp_ws['J' + str(i)].number_format = '#,###'
            model_perf_resp_ws['K' + str(i)].number_format = '0.00%'
            model_perf_resp_ws['L' + str(i)].number_format = '#,###'

    #
    def create_model_predictor_summary_ws(self, tni_lookup_obj=None, column_mapping: dict = None) -> None:
        """
        Def: function to create model predictor summary ws tab

        :param tni_lookup_obj: tni lookup object output from module 2
        :param column_mapping: dictionary of model_fit_out names to excel workbook names
        """
        if column_mapping is None:
            column_mapping = {'Variable Name': 'variable',
                              'Estimate': 'estimate',
                              'Importance': 'importance',
                              'VIF': 'vif', 'P Val': 'p_val',
                              'T Val': 't_z_val', 'std error': 'standard_error',
                              'std coeff': 'standardized_coeff'
                              }
        # get model eval table
        model_eval = self.model_descriptor

        # if tni lookup obj is entered than grab lookup table
        if tni_lookup_obj is not None:
            lookup = tni_lookup_obj['lookup']['lookup_table']

            # make lookup dictionary, not currently used
            # type_lookup_dict = lookup.set_index('var_name').T.to_dict()

        model_pred_summary_ws = self.wb.create_sheet("Model Predictor Summary", 0)

        # make new sheet the active sheet we are working on
        model_pred_summary_ws = self.wb.active

        # remove gridlines from the sheet
        model_pred_summary_ws.sheet_view.showGridLines = False

        # set 1st and 3rd row dimensions for the sheet
        model_pred_summary_ws.row_dimensions[3].height = 25
        model_pred_summary_ws.row_dimensions[1].height = 30

        # set border line thickenss
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # set column widths
        cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for i in cols:
            model_pred_summary_ws.column_dimensions[i].width = 30

        # create title text
        model_pred_summary_ws['B1'] = "Model Predictors Summary"
        model_pred_summary_ws['B1'].font = Font(size=18, underline="single", bold=True)
        model_pred_summary_ws.merge_cells('B1:D1')

        cells_text_map = {
            'B': "Original Var. Name",
            'C': "Variable Name",
            'D': "Var. Data Type",
            'E': "Sign (+/-)",
            'F': "Estimate",
            'G': "Importance",
            'H': "VIF",
            'I': "Description",
            'J': "P Val",
            'K': "T Val",
            'L': 'std error',
            'M': "std coeff"
        }

        cells_text = ['B3', 'C3', 'D3', 'E3', 'F3', 'G3', 'H3', 'I3', 'J3', 'K3', 'L3', 'M3']

        # format text cells
        for i in cells_text:
            model_pred_summary_ws[i].font = Font(bold=True)
            current_cell = model_pred_summary_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            if i == 'B3':
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif i == 'C3' or i == 'D3' or i == 'E3' or i == 'F3' or i == 'G3' or i == 'H3':
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            else:
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)

        round_decimal_places = 3
        remove_intercept = True
        decide_sign_based_on = 'Estimate'
        name_column = 'Variable Name'

        data = model_eval.copy()

        # Get columns from the dataframe which are values of the column_mapping dictionary
        fltrd_data = data[list(column_mapping.values())]
        # Rename columns to required naming using the keys of the column_mapping dictionary
        fltrd_data.columns = list(column_mapping.keys())

        # Add a new column for Sign with '+' default
        fltrd_data['Sign (+/-)'] = '+'
        # Fill the sign column with '-' where 'Coeff.' id negative
        fltrd_data.loc[fltrd_data[decide_sign_based_on] < 0, 'Sign (+/-)'] = '-'

        # if tni lookup obj is entered, then include necessary original var names and datatype columns
        if tni_lookup_obj is not None:
            # Add a new column for var data type
            fltrd_data['Var. Data Type'] = model_eval['variable'].replace(
                lookup[['type', 'var_name']].set_index('var_name').T.to_dict('list'))
            fltrd_data['Original Var. Name'] = model_eval['variable'].replace(
                lookup[['original_name', 'var_name']].set_index('var_name').T.to_dict('list'))

        # Round all the numbers in the dataframe to 'round_decimal_places' decimal places
        fltrd_data = fltrd_data.round(round_decimal_places)

        if remove_intercept:
            fltrd_data = fltrd_data[fltrd_data[name_column] != '(Intercept)']

        # iterrate over data frame created above
        for index, row in fltrd_data.iterrows():
            i = index + 3
            for col_letter in cells_text_map:
                # condition for if column not present
                if not cells_text_map[col_letter] in list(fltrd_data.columns):
                    current_cell = model_pred_summary_ws[col_letter + str(i)]

                    if col_letter == 'B':
                        current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
                    elif col_letter == 'M':
                        current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
                    else:
                        current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    if index == len(fltrd_data):
                        if col_letter == 'B':
                            current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thick)
                        elif col_letter == 'M':
                            current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thick)
                        else:
                            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
                    continue

                current_cell = model_pred_summary_ws[col_letter + str(i)]
                if col_letter == 'B':
                    current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
                elif col_letter == 'M':
                    current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
                else:
                    current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                if index == len(fltrd_data):
                    if col_letter == 'B':
                        current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thick)
                    elif col_letter == 'M':
                        current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thick)
                    else:
                        current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
                # write value from df to current cell
                model_pred_summary_ws[col_letter + str(i)] = row[cells_text_map[col_letter]]

        # write value from df to cell in sheet
        for col_letter in cells_text_map:
            model_pred_summary_ws[col_letter + '3'] = cells_text_map[col_letter]

    #
    def var_prof_subset(self, var_prof_df: pd.DataFrame, tni_lookup_obj=None) -> pd.DataFrame:
        """
        Def: function used to subset profiling dataframe to only include those that were selected in the model.

        :param var_prof_df: variable profile dataframe
        :param tni_lookup_obj: tni lookup object from module 2, this is only needed if the model fit variables are
                        different than the var_prof_df variables (i.e. variables were transformed after profiling)
        :return: subsetted variables profiling dataframe of only model variables.
        """
        # if tni lookup object is entered into the function, then the user probably transformed the model variables
        # after profiling, this will link back the model variables to there original profiling name
        if tni_lookup_obj is not None:
            var_list = []
            for i in range(0, len(self.model_descriptor) - 1):
                var_i = self.model_descriptor['variable'][i + 1]
                for j in range(0, len(tni_lookup_obj['lookup']['lookup_table'])):
                    if tni_lookup_obj['lookup']['lookup_table']['var_name'][j] == var_i:
                        var_list.append(tni_lookup_obj['lookup']['lookup_table']['original_name'][j])
                        var_list.append(tni_lookup_obj['lookup']['lookup_table']['var_name'][j])

                        # if no tni_lookup_obj is entered then the function will assume that the variables were not
                        # transformed after profiling and
        # there is a direct link to the model variables and profiling df variables
        else:
            var_list = list(self.variables.get_active_variables())

        # subset var prof df to only be the variables from the model
        model_var_prof = var_prof_df[var_prof_df['Variable'].isin(var_list)]
        # print(model_var_prof)
        return model_var_prof

    #
    def create_var_profiling_ws(self, var_prof_df: pd.DataFrame, sheet_name: str = "Model Var. Profiling",
                                sort_by_variance: str = None) -> None:
        """
        Def: function that creates variable profiling tab in excel workbook

        :param var_prof_df: variable profiling dataframe
        :param sheet_name: str, sheet name of the profiling excel sheet
        :param sort_by_variance: controls sort method for the variable profile, None if sorting by variable position,
            "desc" if sorting by descending weighted variance, otherwise sorting by ascending weighted variance
        """
        DataProfiler(self.get_training_data(), self.dependent_variable,
                     variables=self.variables).create_var_profiling_ws(wb=self.wb, var_prof_df=var_prof_df,
                                                                       sheet_name=sheet_name,
                                                                       sort_by_variance=sort_by_variance)

    #
    def create_data_summary_resp_ws(self):
        """
        Def: function to create the data summary response worksheet of the model playbook
        """
        # create sheet
        data_summary_resp_ws = self.wb.create_sheet("Data Summary (Resp.)", 0)

        # make new sheet the active sheet we are working on
        data_summary_resp_ws = self.wb.active

        # remove gridlines from the sheet
        data_summary_resp_ws.sheet_view.showGridLines = False

        # set row dimensions for the sheet
        rows = [1, 3, 4, 5, 6]
        for i in rows:
            if i == 1:
                data_summary_resp_ws.row_dimensions[i].height = 25
            else:
                data_summary_resp_ws.row_dimensions[i].height = 40

        # set border line thickenss
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # set column widths
        cols = ['B', 'C', 'D', 'E']
        for i in cols:
            data_summary_resp_ws.column_dimensions[i].width = 15

        # create title text
        data_summary_resp_ws['B1'] = "Title of Model/Data"  # change to what?
        data_summary_resp_ws['B1'].font = Font(size=18, underline="single", bold=True)
        data_summary_resp_ws.merge_cells('B1:D1')

        # cells that need text inputted
        cells_text = ['B3', 'C3', 'D3', 'E3', 'B4', 'B5', 'B6']
        data_summary_resp_ws['B3'] = "Data"
        data_summary_resp_ws['C3'] = "Training"
        data_summary_resp_ws['D3'] = "Testing"
        data_summary_resp_ws['E3'] = "Total"
        data_summary_resp_ws['B4'] = "DV Target"
        data_summary_resp_ws['B5'] = "DV Non-Target"
        data_summary_resp_ws['B6'] = "Total"

        # format text cells
        for i in cells_text:
            data_summary_resp_ws[i].font = Font(bold=True)
            current_cell = data_summary_resp_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            if i == 'B3':
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif i == 'C3' or i == 'D3':
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            elif i == 'E3':
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
            elif i == 'B4' or i == 'B5':
                current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
            else:
                current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thick)

        # cells that have values inputted into them
        cells_values = ['C4', 'C5', 'C6', 'D4', 'D5', 'D6', 'E4', 'E5', 'E6']
        # use inputted model summary data to input values
        data_summary_resp_ws['C4'] = self.train_lift.iloc[:, 1].sum()
        data_summary_resp_ws['C6'] = self.train_lift.iloc[:, 0].sum()
        data_summary_resp_ws['D4'] = self.test_lift.iloc[:, 1].sum()
        data_summary_resp_ws['D6'] = self.test_lift.iloc[:, 0].sum()
        data_summary_resp_ws['C5'] = data_summary_resp_ws['C6'].value - data_summary_resp_ws['C4'].value
        data_summary_resp_ws['D5'] = data_summary_resp_ws['D6'].value - data_summary_resp_ws['D4'].value
        data_summary_resp_ws['E4'] = data_summary_resp_ws['C4'].value + data_summary_resp_ws['D4'].value
        data_summary_resp_ws['E5'] = data_summary_resp_ws['C5'].value + data_summary_resp_ws['D5'].value
        data_summary_resp_ws['E6'] = data_summary_resp_ws['E4'].value + data_summary_resp_ws['E5'].value

        # format value cells
        for i in cells_values:
            current_cell = data_summary_resp_ws[i]
            current_cell.number_format = '#,###'
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="DDDDDD")
            if i == 'C4' or i == 'C5':
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            elif i == 'C6' or i == 'D6':
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
            elif i == 'D4' or i == 'D5':
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            elif i == 'E4' or i == 'E5':
                current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
            else:
                current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thick)

        # generate chart
        chart = BarChart()
        data = Reference(worksheet=data_summary_resp_ws, min_row=3, max_row=6, min_col=3, max_col=4)
        cats = Reference(data_summary_resp_ws, min_col=2, min_row=4, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        data_summary_resp_ws.add_chart(chart, "G3")

    #
    def create_model_perf_rev_ws(self, oot_out: pd.DataFrame = None) -> None:
        """
        Def: function to create the model performance tables for response models

        :param oot_out: scored out of time dataset output from the score_oot_data function
        """

        model_perf_rev_ws = self.wb.create_sheet("Model Perf. (Rev. Model)", 0)

        # make new sheet the active sheet we are working on
        model_perf_rev_ws = self.wb.active

        # remove gridlines from the sheet
        model_perf_rev_ws.sheet_view.showGridLines = False

        # adj. r-squared and AIC
        lin_metrics = self.model_metric

        # training data model performance
        train_perf = self.train_lift

        # testing data model performance
        test_perf = self.test_lift

        # set row dimensions for the sheet
        model_perf_rev_ws.row_dimensions[3].height = 25
        model_perf_rev_ws.row_dimensions[4].height = 25
        model_perf_rev_ws.row_dimensions[5].height = 25
        model_perf_rev_ws.row_dimensions[19].height = 25
        model_perf_rev_ws.row_dimensions[20].height = 25
        model_perf_rev_ws.row_dimensions[21].height = 25

        # set border line thickenss
        # thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # set column widths
        cols = ['C', 'D', 'E', 'F', 'G', 'H', 'J']
        for i in cols:
            model_perf_rev_ws.column_dimensions[i].width = 20

        # create title text
        model_perf_rev_ws['B1'] = "Model Performance"
        model_perf_rev_ws['B1'].font = Font(size=18, underline="single", bold=True)
        model_perf_rev_ws.merge_cells('B1:E1')

        # training data model performance table setup
        model_perf_rev_ws['B3'] = 'Training'
        model_perf_rev_ws['B3'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('B3:C3')

        model_perf_rev_ws['B4'] = 'Decile'
        # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
        model_perf_rev_ws.merge_cells('B4:B5')

        model_perf_rev_ws['C4'] = 'Incremental'
        model_perf_rev_ws['C4'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('C4:E4')

        model_perf_rev_ws['F4'] = 'Cumulative'
        model_perf_rev_ws['F4'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('F4:H4')

        model_perf_rev_ws['C5'] = "Records"
        model_perf_rev_ws['D5'] = "Value"
        model_perf_rev_ws['E5'] = "Lift"
        model_perf_rev_ws['F5'] = "Cum. Records"
        model_perf_rev_ws['G5'] = "Cum. Value"
        model_perf_rev_ws['H5'] = "Cum. Lift"

        model_perf_rev_ws['J4'] = "Adj. R-Squared ="
        model_perf_rev_ws['K4'] = lin_metrics['adj_r_sq']

        model_perf_rev_ws['J5'] = 'AIC ='
        model_perf_rev_ws['K5'] = lin_metrics['aic']
        #########################################################################

        # testing data model performance table set up
        model_perf_rev_ws['B19'] = 'Testing'
        model_perf_rev_ws['B19'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('B19:C19')

        model_perf_rev_ws['B20'] = 'Decile'
        # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
        model_perf_rev_ws.merge_cells('B20:B21')

        model_perf_rev_ws['C20'] = 'Incremental'
        model_perf_rev_ws['C20'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('C20:E20')

        model_perf_rev_ws['F20'] = 'Cumulative'
        model_perf_rev_ws['F20'].font = Font(size=18, bold=True)
        model_perf_rev_ws.merge_cells('F20:H20')

        model_perf_rev_ws['C21'] = "Records"
        model_perf_rev_ws['D21'] = "Value"
        model_perf_rev_ws['E21'] = "Lift"
        model_perf_rev_ws['F21'] = "Cum. Records"
        model_perf_rev_ws['G21'] = "Cum. Value"
        model_perf_rev_ws['H21'] = "Cum. Lift"

        model_perf_rev_ws['J20'] = "Adj. R-Squared ="
        model_perf_rev_ws['K20'] = self.get_adj_rsq(self.testing_data[self.dependent_variable], self.test_out['y_pred'])

        #########################################################################################
        # format text cells

        cells_text = ['B3', 'C3', 'B4', 'B5', 'C4', 'D4', 'E4',
                      'C5', 'D5', 'E5', 'F4', 'G4',
                      'H4', 'F5', 'G5', 'H5', 'J4', 'K4',
                      'J5', 'K5']
        for i in cells_text:
            model_perf_rev_ws[i].font = Font(bold=True)
            current_cell = model_perf_rev_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        input_rows = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
        input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        tot_val = 0
        for i in input_cols:
            z = 1
            for j in input_rows:
                x = str(j)
                x = i + x
                # print(x)
                current_cell = model_perf_rev_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # decile
                if i == 'B':
                    model_perf_rev_ws[x] = train_perf['decile'][z]
                    z = z + 1
                # total records
                elif i == 'C':
                    model_perf_rev_ws[x] = train_perf['number_obs'][z]
                    z = z + 1
                # value
                elif i == 'D':
                    model_perf_rev_ws[x] = train_perf['value'][z]
                    z = z + 1
                # incremental lift
                elif i == 'E':
                    model_perf_rev_ws[x] = train_perf['incremental_lift'][z] * 100
                    z = z + 1
                # cum. records
                elif i == 'F':
                    if j == 6:
                        model_perf_rev_ws[x] = model_perf_rev_ws['C6'].value
                    else:
                        c = 'C' + str(j)
                        f = 'F' + str(j - 1)
                        model_perf_rev_ws[x] = model_perf_rev_ws[c].value + model_perf_rev_ws[f].value
                # cum. value
                elif i == 'G':
                    if j == 6:
                        model_perf_rev_ws[x] = model_perf_rev_ws['D6'].value
                        tot_val = model_perf_rev_ws['C6'].value * model_perf_rev_ws['D6'].value
                    else:
                        d = 'D' + str(j)
                        c = 'C' + str(j)
                        tot_val = tot_val + (model_perf_rev_ws[d].value * model_perf_rev_ws[c].value)
                        f = 'F' + str(j)

                        model_perf_rev_ws[x] = tot_val / model_perf_rev_ws[f].value
                        # cum. lift
                elif i == 'H':
                    g = 'G' + str(j)
                    model_perf_rev_ws[x] = (model_perf_rev_ws[g].value / model_perf_rev_ws['G15'].value) * 100
                else:
                    break

        # create total column values for incremental data
        tot_col = ['B', 'C', 'D', 'E']
        row = 16
        for i in tot_col:
            x = i + str(row)
            current_cell = model_perf_rev_ws[x]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i == 'B':
                model_perf_rev_ws[x] = 'Total'
            elif i == 'C':
                model_perf_rev_ws[x] = train_perf['number_obs'].sum()
            elif i == 'D':
                model_perf_rev_ws[x] = model_perf_rev_ws['G15'].value
                # use inputted model summary data to input values
            elif i == 'E':
                model_perf_rev_ws[x] = (model_perf_rev_ws['D16'].value / model_perf_rev_ws['D16'].value) * 100
            else:
                break

        #########################################################################
        # now need to do the same thing for the testing decile table
        ###########################################################################
        cells_text = ['B19', 'C19', 'B20', 'B21', 'C20', 'D20', 'E20',
                      'C21', 'D21', 'E21', 'F20', 'G20',
                      'H20', 'F21', 'G21', 'H21', 'J20', 'K20']
        for i in cells_text:
            model_perf_rev_ws[i].font = Font(bold=True)
            current_cell = model_perf_rev_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        input_rows = [22, 23, 24, 25, 26, 27, 28, 29, 30, 31]
        input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        tot_val = 0
        for i in input_cols:
            z = 1
            for j in input_rows:
                x = str(j)
                x = i + x
                # print(x)
                current_cell = model_perf_rev_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # decile
                if i == 'B':
                    model_perf_rev_ws[x] = test_perf['decile'][z]
                    z = z + 1
                # total records
                elif i == 'C':
                    model_perf_rev_ws[x] = test_perf['number_obs'][z]
                    z = z + 1
                # value
                elif i == 'D':
                    model_perf_rev_ws[x] = test_perf['value'][z]
                    z = z + 1
                # incremental lift
                elif i == 'E':
                    model_perf_rev_ws[x] = test_perf['incremental_lift'][z] * 100
                    z = z + 1
                # cum. records
                elif i == 'F':
                    if j == 22:
                        model_perf_rev_ws[x] = model_perf_rev_ws['C22'].value
                    else:
                        c = 'C' + str(j)
                        f = 'F' + str(j - 1)
                        model_perf_rev_ws[x] = model_perf_rev_ws[c].value + model_perf_rev_ws[f].value
                # cum. value
                elif i == 'G':
                    if j == 22:
                        model_perf_rev_ws[x] = model_perf_rev_ws['D22'].value
                        tot_val = model_perf_rev_ws['C22'].value * model_perf_rev_ws['D22'].value
                    else:
                        d = 'D' + str(j)
                        c = 'C' + str(j)
                        tot_val = tot_val + (model_perf_rev_ws[d].value * model_perf_rev_ws[c].value)
                        f = 'F' + str(j)
                        model_perf_rev_ws[x] = tot_val / model_perf_rev_ws[f].value
                # cum. lift
                elif i == 'H':
                    g = 'G' + str(j)
                    model_perf_rev_ws[x] = (model_perf_rev_ws[g].value / model_perf_rev_ws['G31'].value) * 100
                else:
                    break
        # create total column values for incremental data
        tot_col = ['B', 'C', 'D', 'E']
        row = 32
        for i in tot_col:
            x = i + str(row)
            current_cell = model_perf_rev_ws[x]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            if i == 'B':
                model_perf_rev_ws[x] = 'Total'
            elif i == 'C':
                model_perf_rev_ws[x] = test_perf['number_obs'].sum()
            elif i == 'D':
                model_perf_rev_ws[x] = model_perf_rev_ws['G31'].value
            elif i == 'E':
                model_perf_rev_ws[x] = (model_perf_rev_ws['D32'].value / model_perf_rev_ws['D32'].value) * 100
            else:
                break

        #########################################################################
        # now need to do the same thing for the out of time decile table
        ###########################################################################

        if oot_out is not None:
            model_perf_rev_ws.row_dimensions[35].height = 25
            model_perf_rev_ws.row_dimensions[36].height = 25
            model_perf_rev_ws.row_dimensions[37].height = 25

            # testing data model performance table set up
            model_perf_rev_ws['B35'] = 'Out of Time'
            model_perf_rev_ws['B35'].font = Font(size=18, bold=True)
            model_perf_rev_ws.merge_cells('B35:C35')

            model_perf_rev_ws['B36'] = 'Decile'
            # model_perf_resp_ws['B3'].font = Font(size = 18, bold = True)
            model_perf_rev_ws.merge_cells('B36:B37')

            model_perf_rev_ws['C36'] = 'Incremental'
            model_perf_rev_ws['C36'].font = Font(size=18, bold=True)
            model_perf_rev_ws.merge_cells('C36:E36')

            model_perf_rev_ws['F36'] = 'Cumulative'
            model_perf_rev_ws['F36'].font = Font(size=18, bold=True)
            model_perf_rev_ws.merge_cells('F36:H36')

            model_perf_rev_ws['C37'] = "Records"
            model_perf_rev_ws['D37'] = "Value"
            model_perf_rev_ws['E37'] = "Lift"
            model_perf_rev_ws['F37'] = "Cum. Records"
            model_perf_rev_ws['G37'] = "Cum. Value"
            model_perf_rev_ws['H37'] = "Cum. Lift"

            model_perf_rev_ws['J36'] = "Adj. R-Squared ="
            model_perf_rev_ws['K36'] = self.get_adj_rsq(oot_out[0][self.dependent_variable], oot_out[0]['y_pred'])

            oot_lift = oot_out[1]

            cells_text = ['B35', 'C35', 'B36', 'B37', 'C36', 'D36', 'E36',
                          'C37', 'D37', 'E37', 'F36', 'G36',
                          'H36', 'F37', 'G37', 'H37', 'J36', 'K36']
            for i in cells_text:
                model_perf_rev_ws[i].font = Font(bold=True)
                current_cell = model_perf_rev_ws[i]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            input_rows = [38, 39, 40, 41, 42, 43, 44, 45, 46, 47]
            input_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
            tot_val = 0
            for i in input_cols:
                z = 0
                for j in input_rows:
                    x = str(j)
                    x = i + x
                    # print(x)
                    current_cell = model_perf_rev_ws[x]
                    current_cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                    # decile
                    if i == 'B':
                        model_perf_rev_ws[x] = oot_lift['decile'].iloc[z]
                        z = z + 1
                    # total records
                    elif i == 'C':
                        model_perf_rev_ws[x] = oot_lift['number_obs'].iloc[z]
                        z = z + 1
                    # value
                    elif i == 'D':
                        model_perf_rev_ws[x] = oot_lift['value'].iloc[z]
                        z = z + 1
                    # incremental lift
                    elif i == 'E':
                        model_perf_rev_ws[x] = oot_lift['incremental_lift'].iloc[z] * 100
                        z = z + 1
                    # cum. records
                    elif i == 'F':
                        if j == 38:
                            model_perf_rev_ws[x] = model_perf_rev_ws['C38'].value
                        else:
                            c = 'C' + str(j)
                            f = 'F' + str(j - 1)
                            model_perf_rev_ws[x] = model_perf_rev_ws[c].value + model_perf_rev_ws[f].value
                    # cum. value
                    elif i == 'G':
                        if j == 38:
                            model_perf_rev_ws[x] = model_perf_rev_ws['D38'].value
                            tot_val = model_perf_rev_ws['C38'].value * model_perf_rev_ws['D38'].value
                        else:
                            d = 'D' + str(j)
                            c = 'C' + str(j)
                            tot_val = tot_val + (model_perf_rev_ws[d].value * model_perf_rev_ws[c].value)
                            f = 'F' + str(j)
                            model_perf_rev_ws[x] = tot_val / model_perf_rev_ws[f].value
                    # cum. lift
                    elif i == 'H':
                        g = 'G' + str(j)
                        model_perf_rev_ws[x] = (model_perf_rev_ws[g].value / model_perf_rev_ws['G47'].value) * 100
                    else:
                        break
            # create total column values for incremental data
            tot_col = ['B', 'C', 'D', 'E']
            row = 48
            for i in tot_col:
                x = i + str(row)
                current_cell = model_perf_rev_ws[x]
                current_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                if i == 'B':
                    model_perf_rev_ws[x] = 'Total'
                elif i == 'C':
                    model_perf_rev_ws[x] = oot_lift['number_obs'].sum()
                elif i == 'D':
                    model_perf_rev_ws[x] = model_perf_rev_ws['G47'].value
                elif i == 'E':
                    model_perf_rev_ws[x] = (model_perf_rev_ws['D48'].value / model_perf_rev_ws['D48'].value) * 100
                else:
                    break

        for i in range(6, 17):
            model_perf_rev_ws['C' + str(i)].number_format = '#,###'
            model_perf_rev_ws['D' + str(i)].number_format = '#,###'
            model_perf_rev_ws['E' + str(i)].number_format = '#,###'
            model_perf_rev_ws['F' + str(i)].number_format = '#,###'
            model_perf_rev_ws['G' + str(i)].number_format = '#,###'
            model_perf_rev_ws['H' + str(i)].number_format = '#,###'

        for i in range(22, 33):
            model_perf_rev_ws['C' + str(i)].number_format = '#,###'
            model_perf_rev_ws['D' + str(i)].number_format = '#,###'
            model_perf_rev_ws['E' + str(i)].number_format = '#,###'
            model_perf_rev_ws['F' + str(i)].number_format = '#,###'
            model_perf_rev_ws['G' + str(i)].number_format = '#,###'
            model_perf_rev_ws['H' + str(i)].number_format = '#,###'

        for i in range(22, 33):
            model_perf_rev_ws['C' + str(i)].number_format = '#,###'
            model_perf_rev_ws['D' + str(i)].number_format = '#,###'
            model_perf_rev_ws['E' + str(i)].number_format = '#,###'
            model_perf_rev_ws['F' + str(i)].number_format = '#,###'
            model_perf_rev_ws['G' + str(i)].number_format = '#,###'
            model_perf_rev_ws['H' + str(i)].number_format = '#,###'

        for i in range(38, 49):
            model_perf_rev_ws['C' + str(i)].number_format = '#,###'
            model_perf_rev_ws['D' + str(i)].number_format = '#,###'
            model_perf_rev_ws['E' + str(i)].number_format = '#,###'
            model_perf_rev_ws['F' + str(i)].number_format = '#,###'
            model_perf_rev_ws['G' + str(i)].number_format = '#,###'
            model_perf_rev_ws['H' + str(i)].number_format = '#,###'

    #
    def create_data_summary_rev_ws(self, width_bin: int) -> None:
        """
        Def: function to create the data summary revenue worksheet of the model playbook

        :param width_bin: width of bins for data summary chart
        """
        dv = self.dependent_variable

        # create sheet
        data_summary_rev_ws = self.wb.create_sheet("Data Summary (Rev.)", 0)

        # make new sheet the active sheet we are working on
        data_summary_rev_ws = self.wb.active

        # remove gridlines from the sheet
        data_summary_rev_ws.sheet_view.showGridLines = False

        # set row dimensions for the sheet
        rows = [1, 3, 4, 5, 6]
        for i in rows:
            if i == 1:
                data_summary_rev_ws.row_dimensions[i].height = 25
            else:
                data_summary_rev_ws.row_dimensions[i].height = 40

        # set border line thickenss
        thick = Side(border_style="thick", color="000000")
        # medium = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")

        # set column widths
        cols = ['B', 'C', 'D', 'E']
        for i in cols:
            data_summary_rev_ws.column_dimensions[i].width = 15

        # create title text
        data_summary_rev_ws['B1'] = "Title of Model/Data"
        data_summary_rev_ws['B1'].font = Font(size=18, underline="single", bold=True)
        data_summary_rev_ws.merge_cells('B1:D1')

        # cells that need text inputted
        cells_text = ['B3', 'C3', 'D3', 'E3', 'B4', 'B5']
        data_summary_rev_ws['B3'] = "Data"
        data_summary_rev_ws['C3'] = "Training"
        data_summary_rev_ws['D3'] = "Testing"
        data_summary_rev_ws['E3'] = "Total"
        data_summary_rev_ws['B4'] = "Quantity"
        data_summary_rev_ws['B5'] = "Avg. Revenue"

        # format text cells
        for i in cells_text:
            data_summary_rev_ws[i].font = Font(bold=True)
            current_cell = data_summary_rev_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="A9C4FE")
            if i == 'B3':
                current_cell.border = Border(top=thick, left=thick, right=thin, bottom=thin)
            elif i == 'C3' or i == 'D3':
                current_cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            elif i == 'E3':
                current_cell.border = Border(top=thick, left=thin, right=thick, bottom=thin)
            elif i == 'B4':
                current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
            else:
                current_cell.border = Border(top=thin, left=thick, right=thin, bottom=thick)

        # cells that have values inputted into them
        cells_values = ['C4', 'C5', 'D4', 'D5', 'E4', 'E5']

        training = self.train_out
        testing = self.test_out
        train_desc = training[dv].describe()
        test_desc = testing[dv].describe()
        full_df = training.append(testing)
        # grouped = df.groupby( "ModelingSample_Flag")
        # grouped_y = grouped["Y"].describe()
        # print(grouped_y)

        # use inputted model summary data to input values
        data_summary_rev_ws['C4'] = train_desc[0]
        data_summary_rev_ws['C5'] = train_desc[1]
        data_summary_rev_ws['D4'] = test_desc[0]
        data_summary_rev_ws['D5'] = test_desc[1]
        data_summary_rev_ws['E4'] = data_summary_rev_ws['C4'].value + data_summary_rev_ws['D4'].value
        data_summary_rev_ws['E5'] = full_df[dv].mean()

        # format value cells
        for i in cells_values:
            current_cell = data_summary_rev_ws[i]
            current_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_cell.fill = PatternFill("solid", fgColor="DDDDDD")
            if i[1] == '4':
                current_cell.number_format = '#,###'
            elif i[1] == '5':
                current_cell.number_format = '$#,###.00'

            if i == 'C4' or i == 'D4':
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            elif i == 'C5' or i == 'D5':
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
            elif i == 'E4':
                current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
            else:
                current_cell.border = Border(top=thin, left=thin, right=thick, bottom=thick)

        # split Y into bins
        n_bins = int((full_df[dv].max() - full_df[dv].min()) // width_bin + 1)
        bin_split = full_df[dv].groupby(
            pd.cut(full_df[dv], np.arange(full_df[dv].min(), full_df[dv].max() + width_bin, width_bin),
                   include_lowest=True))

        # input for chart and hide in white fonts
        alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        data_summary_rev_ws['B7'] = "Bins"
        data_summary_rev_ws['B8'] = "Range"
        data_summary_rev_ws['B9'] = "Quantity"

        for i in range(0, n_bins):
            if i <= 23:
                data_summary_rev_ws[alphabet[i + 2] + '7'] = i + 1
                data_summary_rev_ws[alphabet[i + 2] + '9'] = list(bin_split.count())[i]
            else:
                data_summary_rev_ws[alphabet[(i + 2) // 26 - 1] + alphabet[(i + 2) % 26] + '7'] = i + 1
                data_summary_rev_ws[alphabet[(i + 2) // 26 - 1] + alphabet[(i + 2) % 26] + '9'] = \
                    list(bin_split.count())[i]

        data_summary_rev_ws['C8'] = '[' + str(full_df[dv].min()) + ', ' + str(full_df[dv].min() + width_bin) + ']'
        for i in range(1, n_bins):
            if full_df[dv].min() + width_bin * (i + 1) < full_df[dv].max():
                if i <= 23:
                    data_summary_rev_ws[alphabet[i + 2] + '8'] = '(' + str(
                        full_df[dv].min() + width_bin * i) + ', ' + str(
                        full_df[dv].min() + width_bin * (i + 1)) + ']'
                else:
                    data_summary_rev_ws[alphabet[(i + 2) // 26 - 1] + alphabet[(i + 2) % 26] + '8'] = '(' + str(
                        full_df[dv].min() + width_bin * i) + ', ' + str(full_df[dv].min() + width_bin * (i + 1)) + ']'
            else:
                if i <= 23:
                    data_summary_rev_ws[alphabet[i + 2] + '8'] = '(' + str(
                        full_df[dv].min() + width_bin * i) + ', ' + str(
                        full_df[dv].max()) + ']'
                else:
                    data_summary_rev_ws[alphabet[(i + 2) // 26 - 1] + alphabet[(i + 2) % 26] + '8'] = '(' + str(
                        full_df[dv].min() + width_bin * i) + ', ' + str(full_df[dv].max()) + ']'

        for r in range(7, 10):
            data_summary_rev_ws['B' + str(r)].font = Font(color=colors.WHITE)
            for c in range(0, n_bins):
                if c <= 23:
                    data_summary_rev_ws[alphabet[c + 2] + str(r)].font = Font(color=colors.WHITE)
                else:
                    data_summary_rev_ws[alphabet[(c + 2) // 26 - 1] + alphabet[(c + 2) % 26] + str(r)].font = Font(
                        color=colors.WHITE)

        # generate chart
        chart = BarChart()
        data = Reference(worksheet=data_summary_rev_ws,
                         min_row=8,
                         max_row=9,
                         min_col=3,
                         max_col=2 + n_bins)
        cats = Reference(data_summary_rev_ws, min_col=2, max_col=2, min_row=9, max_row=10)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        data_summary_rev_ws.add_chart(chart, "G3")

        # bins info table
        data_summary_rev_ws['P3'] = "Min Revenue"
        data_summary_rev_ws['P4'] = "Max Revenue"
        data_summary_rev_ws['P5'] = "Bin Width"
        data_summary_rev_ws['P6'] = "# of Bins"
        data_summary_rev_ws['Q3'] = full_df[dv].min()
        data_summary_rev_ws['Q4'] = full_df[dv].max()
        data_summary_rev_ws['Q5'] = width_bin
        data_summary_rev_ws['Q6'] = n_bins

        # format the bins info table
        data_summary_rev_ws.column_dimensions['P'].width = 15
        for i in range(3, 7):
            data_summary_rev_ws['P' + str(i)].font = Font(bold=True)
            data_summary_rev_ws['P' + str(i)].fill = PatternFill("solid", fgColor="A9C4FE")
            data_summary_rev_ws['P' + str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            data_summary_rev_ws['Q' + str(i)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            data_summary_rev_ws['P' + str(i)].alignment = Alignment(horizontal='left', vertical='center')
            data_summary_rev_ws['Q' + str(i)].alignment = Alignment(horizontal='right', vertical='center')
            if i == 3 or i == 4:
                data_summary_rev_ws['Q' + str(i)].number_format = '$#,###.00'
            else:
                data_summary_rev_ws['Q' + str(i)].number_format = '#,###'

    #
    def create_model_playbook(self,
                              playbook_name: str,
                              file_save_loc: str,
                              width_bin: int = None,
                              oot_data: pd.DataFrame = None,
                              sort_by_variance: str = None) -> None:
        """
        Def: function that combines the individual sheet functions into a single playbook and outputs it to its inputted
            location. This will generate Data Summary, Model Variable Profiling, Model Predictor Summary,
            Model Performance, and Model Correlation sheets.

        :param playbook_name: name for the playbook, what is entered here will be inputted in the titles for each sheet
        :param file_save_loc: file save location
        :param width_bin: the width of the bins for model summary chart (only needed if linear model is being created)
        :param oot_data: out of time dataframe
        :param sort_by_variance: controls sort method for the variable profile, None if sorting by variable position,
            "desc" if sorting by descending weighted variance, otherwise sorting by ascending weighted variance
        """
        if self.model_type == 'linear' and width_bin is None:
            raise Exception("The 'width_bin' parameter needs a value when model type is linear")

        self.wb = Workbook()
        pred_oot_data = None
        model_type = self.model_type
        var_prof_df = self.create_custom_profile(list(self.variables.get_active_variables()))

        if oot_data is not None:
            pred_oot_data = self.score_oot_data(oot_data)

        if model_type == 'logistic':
            self.create_model_corr_matrix_ws()
            if oot_data is not None:
                self.create_model_perf_resp_ws(pred_oot_data)
            else:
                self.create_model_perf_resp_ws()

            self.create_model_predictor_summary_ws()
            if var_prof_df is not None:
                model_var_prof_df = self.var_prof_subset(var_prof_df)
                self.create_var_profiling_ws(model_var_prof_df, sort_by_variance=sort_by_variance)
            self.create_data_summary_resp_ws()
        elif model_type == 'linear':
            self.create_model_corr_matrix_ws()
            if oot_data is not None:
                self.create_model_perf_rev_ws(pred_oot_data)
            else:
                self.create_model_perf_rev_ws()

            self.create_model_predictor_summary_ws()
            if var_prof_df is not None:
                model_var_prof_df = self.var_prof_subset(var_prof_df)
                self.create_var_profiling_ws(model_var_prof_df, sort_by_variance=sort_by_variance)
            self.create_data_summary_rev_ws(width_bin)

        extra_sheet = self.wb.get_sheet_by_name('Sheet')
        self.wb.remove_sheet(extra_sheet)

        for i in range(0, len(self.wb.sheetnames)):
            self.wb.active = i
            # for ws in wb.worksheets:
            # ws = wb[sheet]
            self.wb.active['B1'] = playbook_name + " " + self.wb.active.title
            self.wb.active['B1'].font = Font(size=18, underline="single", bold=True)
        # wb.active.merge_cells('B1:E1')
        self.wb.active = 0
        self.wb.save(file_save_loc)

    def clear_data(self) -> None:
        """
        Clear out data variables. Mainly for saving the model with pickle.
        """
        self.data = self.data.iloc[0:0]
        self.training_data = self.training_data.iloc[0:0]
        self.testing_data = self.testing_data.iloc[0:0]
        self.train_out = self.train_out.iloc[0:0]
        self.test_out = self.test_out.iloc[0:0]

    def generate_python_tni(self, in_df_name: str, out_df_name: str = 'out_df') -> str:
        """
        Generates python code to transform a DataFrame to the model's TnI process without the model object
        :param in_df_name: string name of the input dataframe
        :param out_df_name: string name of the output dataframe
        :return: string of the generated code
        """
        base_vars = list(set([self.variables[v].name for v in self.var_selected]))
        output_string = '{0} = {1}.copy()\n\n'.format(out_df_name, in_df_name)

        for og_name in base_vars:
            tni = self.variables[og_name].tni()
            col_map = tni.column_mapping
            val_replace = tni.value_mapping
            col_code = OrderedDict()

            if tni.variable_type == 'continuous':
                proc_name = col_map['tni_processed']
                na_fill = val_replace[None]
                col_code['base'] = "{0}['{1}'] = {0}['{2}'].fillna({3})".format(out_df_name, proc_name, og_name,
                                                                                na_fill)
                if 'tni_transformation_sqrt' in col_map:
                    col_name = col_map['tni_transformation_sqrt']
                    col_code[col_name] = "{0}['{1}'] = np.sqrt({0}['{2}'])".format(out_df_name, col_name, proc_name)

                if 'tni_transformation_log' in col_map:
                    col_name = col_map['tni_transformation_log']
                    col_code[col_name] = "{0}['{1}'] = np.log(np.abs({0}['{2}']))".format(out_df_name, col_name,
                                                                                          proc_name)

                if 'tni_transformation_exp' in col_map:
                    col_name = col_map['tni_transformation_exp']
                    lower_b = tni.lower_bound
                    upper_b = tni.upper_bound
                    col_code[col_name] = "{0}['{1}'] = np.exp(({0}['{2}'] - {3}) / ({4} - {3}))".format(out_df_name,
                                                                                                        col_name,
                                                                                                        proc_name,
                                                                                                        lower_b,
                                                                                                        upper_b)

                if 'tni_transformation_pw2' in col_map:
                    col_name = col_map['tni_transformation_pw2']
                    col_code[col_name] = "{0}['{1}'] = {0}['{2}'] ** 2".format(out_df_name, col_name, proc_name)

                if tni.cutter:
                    cut_col = col_map['tni_assigned_bin']
                    col_code['cut_s'] = "{0}['{1}'] = pd.cut({0}['{2}'], **{3})".format(out_df_name, cut_col, proc_name,
                                                                                        tni.cutter)
                    for col in tni.encoded_column_mapping.keys():
                        col_code[col] = "if '{0}' not in {1}.columns:\n    {1}['{0}'] = 0".format(col, out_df_name)

            if tni.variable_type == 'categorical':
                proc_name = col_map['tni_processed']
                col_code['base'] = "{0}['{1}'] = {0}['{1}'].fillna('TnImissing')".format(out_df_name, og_name)
                col_code['base'] += "\n{0}['{1}'] = {0}['{2}'].replace({3})".format(out_df_name, proc_name, og_name,
                                                                                    val_replace)
                col_code['base'] += "\n{0} = pd.get_dummies({0}, columns=['{1}'])".format(out_df_name, og_name)
                col_code['base'] += "\nfor col in {}:".format(tni._categorical['drop_columns'])
                col_code['base'] += "\n    if col in {}.columns:".format(out_df_name)
                col_code['base'] += "\n        {}.drop(columns=col, inplace=True)".format(out_df_name)

                col_code['base'] += "\nog_tni_dummies = {}".format(list(tni.encoded_column_mapping.keys()))
                col_code['base'] += "\nfor og in og_tni_dummies:"
                col_code['base'] += "\n    if og not in {}.columns:".format(out_df_name)
                col_code['base'] += "\n        {}[og] = 0".format(out_df_name)

            if tni.variable_type == 'binary':
                col_code['base'] = "{0}['{1}'] = {0}['{1}'].fillna('TnImissing')".format(out_df_name, og_name)
                col_code['base'] += "\n{0} = pd.get_dummies({0}, columns=['{1}'])".format(out_df_name, og_name)
                col_code['base'] += "\nog_tni_dummies = {}".format(list(tni.encoded_column_mapping.keys()))
                col_code['base'] += "\nfor og in og_tni_dummies:"
                col_code['base'] += "\n    if og not in {}.columns:".format(out_df_name)
                col_code['base'] += "\n        {}[og] = 0".format(out_df_name)

            for k, v in col_code.items():
                if k in ['base', 'cut_s', 'cut_e'] + self.var_selected:
                    output_string += v + '\n'

            output_string += '\n'

        model_cols = ", ".join(["'{}'".format(c) for c in self.var_selected])
        output_string += "{0} = {0}[[{1}]]".format(out_df_name, model_cols)
        return output_string

    def generate_python_score(self, df_name: str, y_name: str = 'y_pred'):
        """
        Generates python code to score the TnI processed dataframe outside of the model
        :param df_name: string name of the dataframe
        :param y_name: string name to save the score to
        :return: string python code
        """
        result_sr = self._model.params
        output_string = "{0}['{1}'] = ".format(df_name, y_name)
        val_string = ''
        for col, val in result_sr.items():
            if col:
                if col == 'const':
                    val_string += " + {}".format(str(val))
                else:
                    val_string += " + ({0} * {1}['{2}'])".format(str(val), df_name, col)

        val_string = val_string[3:]

        if self.model_type == 'linear':
            output_string += val_string
        else:
            output_string += '1 / (1 + np.exp(-1.0 * ({0})))'.format(val_string)
        return output_string

    def generate_python_code(self,
                             in_df_name: str,
                             out_df_name: str = 'out_df',
                             y_name: str = 'y_pred') -> str:
        """
        Generates python code for both the data transformations and the scoring of the model
        :param in_df_name: string name in the input dataframe
        :param out_df_name: string name of the output dataframe
        :param y_name: string name of the output score column
        :return: string python code
        """
        output_string = self.generate_python_tni(in_df_name=in_df_name, out_df_name=out_df_name)
        output_string += "\n\n"
        output_string += self.generate_python_score(df_name=out_df_name, y_name=y_name)
        return output_string
